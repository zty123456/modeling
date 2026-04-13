"""Component classification and color palette for Excel output.

Returns short op-level labels (e.g. ``attn_norm``, ``attn.score``,
``moe.gate.mm``, ``ffn.gate_proj``).  Architecture-specific patterns for
DeepSeek MLA/MoE are tried first; generic fallbacks cover LLaMA, Qwen,
Mistral, Mixtral, and other standard transformers models.
"""
from __future__ import annotations

from typing import Optional

from openpyxl.styles import PatternFill

FILL_COLORS = {
    "attn_norm":    "E8F5E9",
    "ffn_norm":     "E8F5E9",
    "final_norm":   "E8F5E9",
    "attn.":        "E3F2FD",
    "moe.gate":     "FFF3E0",
    "moe.shared":   "FFF8E1",
    "moe.expert":   "FCE4EC",
    "ffn":          "F3E5F5",
    "embedding":    "ECEFF1",
    "lm_head":      "ECEFF1",
    "add":          "FFFFFF",
}


def extract_layer_idx(module_path: str) -> str:
    """Return the transformer-block index from a dotted module path."""
    parts = module_path.split(".")
    block_containers = {"layers", "blocks", "h", "layer"}
    for i, part in enumerate(parts):
        if part in block_containers and i + 1 < len(parts):
            try:
                return str(int(parts[i + 1]))
            except ValueError:
                pass
    return ""


def classify_component(module_path: str, func_name: str) -> str:
    """Map a module path to a human-readable component category.

    Short-name convention (matches the original remote style):
      attn_norm / ffn_norm / final_norm
      attn.q_a_proj / attn.score / attn.softmax / attn.rope / attn.<op>
      moe.gate.<op> / moe.shared.<proj> / moe.experts.<op>
      ffn.gate_proj / ffn.up_proj / ffn.down_proj / ffn.silu / ffn.<op>
      embedding / lm_head
      <op>   ← catch-all: bare aten short name
    """
    s = module_path.lower()
    fn_parts = func_name.split(".")
    op_short = fn_parts[1] if len(fn_parts) >= 2 else func_name

    # ── Norm layers ──────────────────────────────────────────────────────────
    if "input_layernorm" in s:
        return "attn_norm"
    if "post_attention_layernorm" in s:
        return "ffn_norm"
    # Generic pre/post norm names used by Llama / Qwen / Mistral
    if any(tok in s for tok in ("layernorm", "rmsnorm", "rms_norm", "layer_norm")):
        if "pre" in s or "input" in s:
            return "attn_norm"
        if "post" in s:
            return "ffn_norm"
        if "attn" not in s and "mlp" not in s and "expert" not in s:
            return "final_norm"

    # ── MLA projections (DeepSeek) ────────────────────────────────────────
    if "q_a_proj" in s:
        return "attn.q_a_proj"
    if "q_a_layernorm" in s:
        return "attn.q_norm"
    if "q_b_proj" in s:
        return "attn.q_b_proj"
    if "kv_a_proj" in s:
        return "attn.kv_a_proj"
    if "kv_a_layernorm" in s:
        return "attn.kv_norm"
    if "kv_b_proj" in s:
        return "attn.kv_b_proj"

    # ── Standard attention projections ───────────────────────────────────
    if "q_proj" in s:
        return "attn.q_proj"
    if "k_proj" in s:
        return "attn.k_proj"
    if "v_proj" in s:
        return "attn.v_proj"
    if "o_proj" in s:
        return "attn.o_proj"
    if "out_proj" in s:
        return "attn.o_proj"

    # ── Attention compute ────────────────────────────────────────────────
    if "self_attn" in s or "attention" in s or (
            "attn" in s and "norm" not in s):
        if "rotary" in s or "rope" in s:
            return "attn.rope"
        if op_short in ("matmul", "mm", "bmm"):
            return "attn.score"
        if "softmax" in op_short or "safe_softmax" in op_short:
            return "attn.softmax"
        return f"attn.{op_short}"

    # ── MoE shared expert ────────────────────────────────────────────────
    if "shared_expert" in s or ("shared" in s and "mlp" in s):
        if "gate_proj" in s:
            return "moe.shared.gate_proj"
        if "up_proj" in s:
            return "moe.shared.up_proj"
        if "down_proj" in s:
            return "moe.shared.down_proj"
        return f"moe.shared.{op_short}"

    # ── MoE gate / router ────────────────────────────────────────────────
    _path_parts = s.split(".")
    _bare_gate = (
            "gate" in _path_parts
            and "gate_proj" not in s
            and "gate_up" not in s
    )

    if _bare_gate or (
            "gate" in s and "experts" not in s and "up" not in s
            and "moe" in s):
        return f"moe.gate.{op_short}"

    # ── MoE routed experts ────────────────────────────────────────────────
    if "experts" in s or "expert" in s:
        return f"moe.experts.{op_short}"

    # ── Dense FFN / MLP ──────────────────────────────────────────────────
    if "mlp" in s or "moe" in s or "ffn" in s or "feed_forward" in s:
        if "gate_proj" in s:
            return "ffn.gate_proj"
        if "up_proj" in s:
            return "ffn.up_proj"
        if "down_proj" in s:
            return "ffn.down_proj"
        if "fc1" in s:
            return "ffn.fc1"
        if "fc2" in s:
            return "ffn.fc2"
        if op_short in ("silu", "gelu", "relu", "gelu_new"):
            return "ffn.silu"
        if op_short == "mul":
            return "ffn.mul"
        return f"ffn.{op_short}"

    # ── Embedding / head / final norm ────────────────────────────────────
    if "embed" in s:
        return "embedding"
    if "norm" in s:
        return "final_norm"
    if "lm_head" in s:
        return "lm_head"

    return op_short


def get_fill(component: str) -> Optional[PatternFill]:
    for prefix, color in FILL_COLORS.items():
        if component.startswith(prefix):
            return PatternFill(start_color=color, end_color=color, fill_type="solid")
    return None
