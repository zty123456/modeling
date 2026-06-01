from __future__ import annotations

import concurrent.futures
import itertools
import logging
import math
import multiprocessing
import os
import time
from copy import copy
from concurrent.futures import ProcessPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Generator, Iterable

import pandas as pd
from tqdm import tqdm

# Unit conventions:
# - Search filtering uses decimal GB (/1e9) for consistency with max_memory_gb parameter
# - report_to_dict() uses GiB (1024**3) via MemBreakdown.to_gb()
# - Difference: 1000 bytes / 1e9 = 1.00 decimal GB, 1000 bytes / 1024**3 = 0.93 GiB (~7%)

from zrt.hardware.registry import load as load_hw
from zrt.training.io.config_loader import _parse_model, _parse_system, _parse_strategy
from zrt.training.models.memory import memory_breakdown
from zrt.training.search.estimator import estimate
from zrt.training.search.report import report_summary
from zrt.training.spec.model import ModelSpec
from zrt.training.spec.report import TrainingReport
from zrt.training.spec.strategy import (
    CPKind,
    MuonConfig,
    OptKind,
    PPSched,
    RecomputePolicy,
    Strategy,
    TPOverlap,
)
from zrt.training.spec.system import GPU, SystemSpec
from zrt.training.topology import comm_domain_report, format_comm_domain_entry

logger = logging.getLogger(__name__)

_WORKER_MODEL_CACHE: Dict[Tuple[str, Optional[str]], ModelSpec] = {}
_WORKER_HW_CACHE: Dict[Tuple[str, int], SystemSpec] = {}
_WORKER_GRAPH_CACHE: Dict[Tuple[Any, ...], Any] = {}

_MODELS_DIR = Path(__file__).parent.parent / "configs" / "models"
_DEFAULT_POD_PACKING_AXES = ("tp", "cp")
_COMM_DOMAIN_AXES = ("ep", "pp", "dp", "tp", "cp")

RAW_DATA_HEADERS = [
    "组号", "model", "hw", "micro_batch", "seq_len", "total_token", "zero_stage",
    "pp_schedule", "cp_kind", "vpp_chunks", "tp_overlap", "ep_overlap",
    "dualbatch", "dp_overlap_in_bubble", "recompute", "optimizer", "quant_preset",
    "world_size", "global_batch", "tp", "cp", "pp", "ep", "dp",
    "ep_comm_domain", "pp_comm_domain", "dp_comm_domain", "tp_comm_domain",
    "cp_comm_domain", "compute_time_ms", "fwd_compute_ms", "bwd_compute_ms",
    "exposed_comm_ms", "tp_total_ms", "tp_exposed_ms", "cp_total_ms",
    "cp_exposed_ms", "ep_total_ms", "ep_exposed_ms", "pp_total_ms",
    "pp_exposed_ms", "pp_hidden_ms", "dp_total_ms", "dp_exposed_ms",
    "optimizer_compute_ms", "optimizer_comm_ms", "optimizer_exposed_ms",
    "recompute_time_ms", "recompute_time_raw_ms", "step_time_ms",
    "pipeline_time_ms", "mfu", "mfu_native", "hfu", "bubble_fraction",
    "bubble_time_ms", "tokens_per_sec", "weights_gb", "grads_gb",
    "opt_state_gb", "activations_gb", "comm_buffers_gb", "memory_gb",
]

ANALYSIS_HEADERS = [
    "组号", "硬件+seq", "TP", "PP", "DP", "EP", "CP", "重计算",
    "单卡迭代时间", "集群吞吐", "集群吞吐归一化", "计算占比",
    "TP通信占比", "EP通信占比", "PP通信占比", "DP通信占比",
    "CP通信占比", "优化器占比", "空泡占比", "fw_time", "bw_time",
    "recompute_time", "计算时间", "TP通信时间(未掩盖)",
    "EP通信时间(未掩盖)", "PP通信时间(未掩盖)", "DP通信时间(未掩盖)",
    "CP通信时间(未掩盖)", "TP通信时间(掩盖)", "EP通信时间(掩盖)",
    "PP通信时间(掩盖)", "DP通信时间(掩盖)", "CP通信时间(掩盖)",
    "优化器计算时间", "优化器通信时间", "优化器时间(掩盖)", "空泡时间",
]

RAW_TO_ANALYSIS = {
    "TP": "tp",
    "PP": "pp",
    "DP": "dp",
    "EP": "ep",
    "CP": "cp",
    "重计算": "recompute",
    "单卡迭代时间": "step_time_ms",
    "集群吞吐": "tokens_per_sec",
    "fw_time": "fwd_compute_ms",
    "bw_time": "bwd_compute_ms",
    "recompute_time": "recompute_time_ms",
    "TP通信时间(未掩盖)": "tp_total_ms",
    "EP通信时间(未掩盖)": "ep_total_ms",
    "PP通信时间(未掩盖)": "pp_total_ms",
    "DP通信时间(未掩盖)": "dp_total_ms",
    "CP通信时间(未掩盖)": "cp_total_ms",
    "TP通信时间(掩盖)": "tp_exposed_ms",
    "EP通信时间(掩盖)": "ep_exposed_ms",
    "PP通信时间(掩盖)": "pp_exposed_ms",
    "DP通信时间(掩盖)": "dp_exposed_ms",
    "CP通信时间(掩盖)": "cp_exposed_ms",
    "优化器计算时间": "optimizer_compute_ms",
    "优化器通信时间": "optimizer_comm_ms",
    "优化器时间(掩盖)": "optimizer_exposed_ms",
    "空泡时间": "bubble_time_ms",
}

PERCENT_ANALYSIS_HEADERS = {
    "计算占比", "TP通信占比", "EP通信占比", "PP通信占比",
    "DP通信占比", "CP通信占比", "优化器占比", "空泡占比",
}
REPORT_HEADER_FILL_COLOR = "4F81BD"
REPORT_EVEN_GROUP_FILL_COLOR = "C6EFCE"


def _ceil_nodes_for_world_size(world_size: int, gpus_per_node: int) -> int:
    if gpus_per_node <= 0:
        raise ValueError("gpus_per_node must be positive")
    return max(1, math.ceil(world_size / gpus_per_node))


def _inferred_gpus_per_node(hw: Any) -> int:
    """Search allocation unit inferred from the innermost hardware tier."""
    try:
        n = int(hw.interconnect.tiers[0].link.num_devices)
    except (AttributeError, IndexError, TypeError, ValueError):
        n = 0
    return n if n > 0 else 8


def _reject_gpus_per_node_config(config_or_grid: Dict[str, Any]) -> None:
    if "gpus_per_node" in config_or_grid:
        raise ValueError(
            "gpus_per_node has been removed from search configs; express "
            "hardware topology via the hardware YAML interconnect.tiers"
        )


def _exact_global_batch_from_total_token(total_token: int | float, seq_len: int | float) -> int:
    if seq_len <= 0:
        raise ValueError("seq_len must be positive when deriving global_batch from total_token")
    total = int(total_token)
    seq = int(seq_len)
    if total != float(total_token) or seq != float(seq_len):
        raise ValueError("total_token and seq_len must be integers")
    if total % seq != 0:
        raise ValueError(
            f"total_token must be divisible by seq_len in exact-token mode: "
            f"total_token={total}, seq_len={seq}"
        )
    return total // seq


def _apply_total_token_batch_rule(
    config: Dict[str, Any],
    *,
    default_seq_len: int,
    model: ModelSpec | None = None,
) -> None:
    total_token = config.get("total_token")
    seq_len = config.get("seq_len", default_seq_len)
    if total_token is not None and total_token > 0:
        config["seq_len"] = int(seq_len)
        config["global_batch"] = _exact_global_batch_from_total_token(total_token, seq_len)
    if model is not None and seq_len is not None:
        model.seq_len = int(seq_len)


def _warn_if_partial_allocation(system: SystemSpec) -> None:
    if system.idle_gpus > 0:
        logger.warning(
            "Search allocation has idle GPUs: world_size=%s allocated_gpus=%s "
            "idle_gpus=%s nodes=%s gpus_per_node=%s",
            system.world_size,
            system.allocated_gpus,
            system.idle_gpus,
            system.nodes,
            system.gpus_per_node,
        )


def _system_from_hw(
    hw: Any,
    *,
    nodes: int,
    gpus_per_node: int,
    world_size_override: int | None,
    host_mem_gb: float,
    warn_partial: bool = True,
) -> SystemSpec:
    system = SystemSpec(
        gpu=GPU(
            name=hw.name,
            flops_bf16=hw.compute.bf16_tflops,
            flops_fp8=hw.compute.fp8_tops or hw.compute.bf16_tflops * 2,
            flops_fp4=hw.compute.fp4_tops,
            hbm_gb=hw.memory.capacity_gb,
            hbm_bw_gbps=hw.memory.hbm_bandwidth_gbps,
            cube_tflops=hw.compute.cube_bf16_tflops,
            vector_tflops=hw.compute.vector_bf16_tflops,
            overlap_ratio=dict(hw.compute.overlap_ratio),
            compute_efficiency=hw.compute.compute_efficiency,
            mem_bw_efficiency=hw.memory.mem_bw_efficiency,
        ),
        interconnect=hw.interconnect,
        nodes=nodes,
        gpus_per_node=gpus_per_node,
        world_size_override=world_size_override,
        host_mem_gb=host_mem_gb,
    )
    if warn_partial:
        _warn_if_partial_allocation(system)
    return system


def _normalize_pod_packing_axes(value: Any) -> tuple[str, ...]:
    if value is None:
        return _DEFAULT_POD_PACKING_AXES
    if isinstance(value, str):
        parts = [p.strip().lower() for p in value.split(",")]
    else:
        parts = [str(p).strip().lower() for p in value]
    axes = tuple(p for p in parts if p)
    allowed = {"tp", "cp", "pp", "dp"}
    invalid = [axis for axis in axes if axis not in allowed]
    if invalid:
        raise ValueError(
            f"Unsupported pod_packing_axes {invalid}; expected subset of {sorted(allowed)}"
        )
    return axes or _DEFAULT_POD_PACKING_AXES


def _pod_packing_group_size(
    axes: tuple[str, ...],
    *,
    tp: int,
    cp: int,
    pp: int,
    dp: int,
) -> int:
    values = {"tp": tp, "cp": cp, "pp": pp, "dp": dp}
    group_size = 1
    for axis in axes:
        group_size *= values[axis]
    return group_size


def _passes_pod_packing(
    *,
    tp: int,
    cp: int,
    pp: int,
    dp: int,
    target_ws: int,
    system: SystemSpec | None,
    other_config: Dict[str, Any] | None,
) -> bool:
    cfg = other_config or {}
    if system is None:
        raise ValueError("_passes_pod_packing requires system for tier-aware checks")
    if target_ws % system.gpus_per_node == 0:
        return True
    axes = _normalize_pod_packing_axes(cfg.get("pod_packing_axes"))
    if _pod_packing_group_size(axes, tp=tp, cp=cp, pp=pp, dp=dp) <= 1:
        return True

    try:
        from zrt.training.topology.process_groups import build_process_groups

        # EP shares physical ranks, so tier assignment is independent of EP.
        strategy = Strategy(tp=tp, cp=cp, pp=pp, dp=dp, ep=1)
        groups = build_process_groups(target_ws, strategy, system)
    except ValueError:
        return False

    outermost = len(system.interconnect.tiers) - 1
    degrees = {"tp": tp, "cp": cp, "pp": pp, "dp": dp}
    for axis in axes:
        if degrees[axis] <= 1:
            continue
        assignment = groups.tier.get(axis.upper())
        if assignment is None:
            return False
        tier = system.interconnect.tiers[assignment.primary_tier]
        if assignment.primary_tier >= outermost and tier.link.num_devices == 0:
            return False
    return True


def _load_model_spec(model_name: str, quant_preset: Optional[str] = None) -> ModelSpec:
    import yaml
    path = _MODELS_DIR / f"{model_name}.yaml"
    if not path.exists():
        raise FileNotFoundError(
            f"Model config not found: {path}. "
            f"Available: {[p.stem for p in _MODELS_DIR.glob('*.yaml')]}"
        )
    with open(path, encoding="utf-8") as f:
        d = yaml.safe_load(f)
    if quant_preset:
        d["quant_preset"] = quant_preset
    return _parse_model(d)


def _make_system_from_config(config: Dict, *, warn_partial: bool = True) -> SystemSpec:
    _reject_gpus_per_node_config(config)
    hw_name = config.get("hw", "nvidia_h100_sxm")
    hw = load_hw(hw_name)

    gpus_per_node = _inferred_gpus_per_node(hw)
    world_size_override = None
    if "world_size" in config:
        world_size_override = int(config["world_size"])
        nodes = _ceil_nodes_for_world_size(world_size_override, gpus_per_node)
    else:
        nodes = config.get("nodes", 1)

    return _system_from_hw(
        hw,
        nodes=nodes,
        gpus_per_node=gpus_per_node,
        world_size_override=world_size_override,
        host_mem_gb=config.get("host_mem_gb", 256.0),
        warn_partial=warn_partial,
    )


def _make_strategy_from_config(config: Dict) -> Strategy:
    recompute = RecomputePolicy()
    rc_str = config.get("recompute", "none")
    if rc_str == "selective":
        recompute.per_layer = {"moe": {"attn"}, "dense": {"attn"}}
    elif rc_str == "full":
        recompute.per_layer = {"moe": {"full"}, "dense": {"full"}}
    elif rc_str == "mhc":
        recompute.per_layer = {"moe": {"hc"}}

    muon_config = None
    opt_str = config.get("optimizer", "adam")
    if opt_str == "muon":
        muon_config = MuonConfig(rotation=config.get("muon_rotation", True))

    pp_schedule = PPSched(config.get("pp_schedule", "1f1b"))
    vpp_chunks = config.get("vpp_chunks", 1)
    if pp_schedule not in (PPSched.INTERLEAVED, PPSched.DUALPIPE_V):
        vpp_chunks = 1

    return Strategy(
        tp=config.get("tp", 1),
        cp=config.get("cp", 1),
        pp=config.get("pp", 1),
        ep=config.get("ep", 1),
        dp=config.get("dp", 1),
        micro_batch=config.get("micro_batch", 1),
        global_batch=config.get("global_batch", 0),
        pp_schedule=pp_schedule,
        vpp_chunks=vpp_chunks,
        zero_stage=config.get("zero_stage", 0),
        recompute=recompute,
        optimizer=OptKind(opt_str),
        muon_config=muon_config,
        tp_overlap=TPOverlap(config.get("tp_overlap", "none")),
        ep_overlap=config.get("ep_overlap", False),
        cp_kind=CPKind(config.get("cp_kind", "none")),
        cp_ulysses=config.get("cp_ulysses"),
        cp_ring=config.get("cp_ring"),
        dualbatch=config.get("dualbatch", False),
        dp_overlap_in_bubble=config.get("dp_overlap_in_bubble", True),
        dp_grad_buckets=config.get("dp_grad_buckets", 25),
    )


def _empty_comm_domain_columns() -> dict[str, str]:
    return {f"{axis}_comm_domain": "" for axis in _COMM_DOMAIN_AXES}


def _comm_domain_columns_from_config(config: Dict[str, Any]) -> dict[str, str]:
    try:
        system = _make_system_from_config(config, warn_partial=False)
        strategy = _make_strategy_from_config(config)
        report = comm_domain_report(
            system,
            strategy,
            kinds=tuple(axis.upper() for axis in _COMM_DOMAIN_AXES),
        )
    except Exception as exc:
        logger.debug("Unable to derive communication-domain columns: %s", exc)
        return _empty_comm_domain_columns()

    return {
        f"{axis}_comm_domain": format_comm_domain_entry(report[axis.upper()])
        for axis in _COMM_DOMAIN_AXES
    }


@dataclass
class TrainingConfigManager:
    param_grid: Dict[str, List[Any]]
    output_path: str = ""

    def __post_init__(self):
        model = self.param_grid.get("model", "unknown")
        if isinstance(model, list):
            model = model[0] if model else "unknown"
        world_sizes = self.param_grid.get("world_size", [1])
        max_world = max(world_sizes) if isinstance(world_sizes, list) else world_sizes
        self.output_path = os.path.join("output", "training_search", f"{model}_ws_{max_world}")

    def _get_divisors(self, n: int) -> List[int]:
        return [i for i in range(1, n + 1) if n % i == 0]

    def _build_strategy_for_validation(
            self, tp: int, cp: int, pp: int, ep: int, dp: int,
            other_config: Dict[str, Any]
    ) -> Strategy:
        pp_schedule = PPSched(other_config.get("pp_schedule", "1f1b"))
        vpp_chunks = other_config.get("vpp_chunks", 1)
        if pp_schedule not in (PPSched.INTERLEAVED, PPSched.DUALPIPE_V):
            vpp_chunks = 1

        recompute = RecomputePolicy()
        rc_str = other_config.get("recompute", "none")
        if rc_str == "selective":
            recompute.per_layer = {"moe": {"attn"}, "dense": {"attn"}}
        elif rc_str == "full":
            recompute.per_layer = {"moe": {"full"}, "dense": {"full"}}
        elif rc_str == "mhc":
            recompute.per_layer = {"moe": {"hc"}}

        muon_config = None
        opt_str = other_config.get("optimizer", "adam")
        if opt_str == "muon":
            muon_config = MuonConfig(rotation=other_config.get("muon_rotation", True))

        return Strategy(
            tp=tp, cp=cp, pp=pp, ep=ep, dp=dp,
            micro_batch=other_config.get("micro_batch", 1),
            global_batch=other_config.get("global_batch", 0),
            pp_schedule=pp_schedule,
            vpp_chunks=vpp_chunks,
            zero_stage=other_config.get("zero_stage", 0),
            recompute=recompute,
            optimizer=OptKind(opt_str),
            muon_config=muon_config,
            tp_overlap=TPOverlap(other_config.get("tp_overlap", "none")),
            ep_overlap=other_config.get("ep_overlap", False),
            cp_kind=CPKind(other_config.get("cp_kind", "none")),
            cp_ulysses=other_config.get("cp_ulysses"),
            cp_ring=other_config.get("cp_ring"),
            dualbatch=other_config.get("dualbatch", False),
            dp_overlap_in_bubble=other_config.get("dp_overlap_in_bubble", True),
            dp_grad_buckets=other_config.get("dp_grad_buckets", 25),
        )

    def _expand_auto_values_optimized(
        self,
        grid: Dict[str, List[Any]],
        world_size: int,
        model: ModelSpec | None = None,
    ) -> None:
        """
        核心优化：基于已知维度的最小值，动态裁剪并收紧 auto 变量的选择范围。
        避免盲目扩展成全量 world_size 的约数，将 auto 并行搜索空间缩减 90% 以上。
        
        注意：EP 不占用额外 rank，world_size = TP*CP*PP*DP
        """
        rank_keys = ["tp", "cp", "pp", "dp"]
        all_keys = ["tp", "cp", "pp", "ep", "dp"]

        # 确保所有并行维度都有默认值
        for key in all_keys:
            if key not in grid:
                grid[key] = [1]

        # 1. 整理出哪些键是固定值或具体范围，哪些键被设为了 "auto"
        explicit_rank_keys = []
        auto_rank_keys = []
        auto_ep = False

        for key in all_keys:
            vals = grid[key]
            is_auto = vals == "auto" or vals == ["auto"] or "auto" in vals
            if key == "ep":
                auto_ep = is_auto
            elif is_auto:
                auto_rank_keys.append(key)
            else:
                if key in rank_keys:
                    explicit_rank_keys.append(key)

        if not auto_rank_keys and not auto_ep:
            return

        # 2. 计算已知 rank 维度的乘积最小值（不包括 EP）
        min_explicit_prod = 1
        for key in explicit_rank_keys:
            vals = [v for v in grid[key] if v != "auto"]
            if vals:
                min_explicit_prod *= min(vals)

        # 3. 动态推导 auto rank 变量所允许的最大上限边界值
        max_allowed_val = world_size // min_explicit_prod

        # 提取当前合法边界内所有符合物理整除条件的约数
        all_divisors = self._get_divisors(world_size)
        optimized_divisors = [d for d in all_divisors if d <= max_allowed_val]

        # 4. 回填 rank 参数网格（tp, cp, pp, dp）
        for key in auto_rank_keys:
            vals = grid[key]
            if isinstance(vals, list):
                clean_vals = [v for v in vals if v != "auto"]
                grid[key] = sorted(list(set(clean_vals + optimized_divisors)))
            else:
                grid[key] = optimized_divisors

        # 5. EP 特殊处理：EP 独立扩展，不参与 rank 计算
        if auto_ep:
            if model is None:
                ep_candidates = optimized_divisors
            elif model.num_experts > 0:
                ep_candidates = self._get_divisors(model.num_experts)
            else:
                ep_candidates = [1]

            vals = grid["ep"]
            if isinstance(vals, list):
                clean_vals = [v for v in vals if v != "auto"]
                grid["ep"] = sorted(list(set(clean_vals + ep_candidates)))
            else:
                grid["ep"] = ep_candidates

    def _enumerate_valid_parallel_configs(
            self, grid: Dict[str, List[Any]], target_ws: int,
            model: ModelSpec = None, system: SystemSpec = None,
            other_config: Dict[str, Any] = None
    ) -> Generator[Tuple[int, int, int, int, int], None, None]:
        tp_vals = grid.get("tp", [1])
        cp_vals = grid.get("cp", [1])
        pp_vals = grid.get("pp", [1])
        ep_vals = grid.get("ep", [1])
        global_batch = other_config.get("global_batch", 0) if other_config else 0
        micro_batch = other_config.get("micro_batch", 1) if other_config else 1
        zero_stage = other_config.get("zero_stage", 0) if other_config else 0
        cp_kind_str = other_config.get("cp_kind", "none") if other_config else "none"
        pp_schedule_str = other_config.get("pp_schedule", "1f1b") if other_config else "1f1b"

        for tp in tp_vals:
            if model is not None:
                if model.num_heads % tp != 0:
                    continue
                if model.hidden % tp != 0:
                    continue
                if model.num_kv_heads % tp != 0 and model.num_kv_heads >= tp:
                    continue
                if model.ffn % tp != 0:
                    continue

            for cp in cp_vals:
                if model is not None and cp_kind_str == "ulysses":
                    if model.num_heads % cp != 0:
                        continue

                for pp in pp_vals:
                    if model is not None and pp > len(model.layers):
                        continue

                    remaining = target_ws // (tp * cp * pp)
                    if remaining <= 0 or target_ws % (tp * cp * pp) != 0:
                        continue

                    for dp in grid.get("dp", [1]):
                        if tp * cp * pp * dp != target_ws:
                            continue

                        if not _passes_pod_packing(
                            tp=tp, cp=cp, pp=pp, dp=dp,
                            target_ws=target_ws, system=system,
                            other_config=other_config,
                        ):
                            continue

                        if global_batch > 0:
                            if global_batch % (micro_batch * dp) != 0:
                                continue

                        if zero_stage >= 1 and dp <= 1:
                            continue

                        for ep in ep_vals:
                            if ep > 1 and (dp < ep or dp % ep != 0):
                                continue
                            if model is not None and ep > 1:
                                if model.num_experts <= 0:
                                    continue
                                if model.num_experts % ep != 0:
                                    continue
                                # Must have at least ep distinct ranks outside
                                # TP to place each expert on its own rank
                                # group; ep > rank_pool means we can't even
                                # cover one EP replica.
                                rank_pool = dp * pp * cp
                                if ep > rank_pool:
                                    continue

                            yield (tp, cp, pp, ep, dp)

    def count_total_configs(self) -> int:
        grid = {k: (v if isinstance(v, list) else [v]) for k, v in self.param_grid.items()}
        _reject_gpus_per_node_config(grid)
        world_sizes = grid.get("world_size", [1])
        target_ws = world_sizes[0]
        model_name = grid.get("model", ["unknown"])[0] if grid.get("model") else "unknown"
        seq_len = grid.get("seq_len", [4096])[0] if grid.get("seq_len") else 4096
        model_for_auto = None
        try:
            model_for_auto = _load_model_spec(model_name)
            model_for_auto.seq_len = seq_len
        except Exception:
            pass
        self._expand_auto_values_optimized(grid, target_ws, model_for_auto)

        parallel_keys = ["tp", "cp", "pp", "ep", "dp"]
        
        total_token_vals = grid.get("total_token", [])
        has_total_token = total_token_vals and any(v is not None and v > 0 for v in total_token_vals)
        
        other_keys = [k for k in grid.keys() if k not in parallel_keys and k != "world_size"]
        if has_total_token and "global_batch" in other_keys:
            other_keys.remove("global_batch")

        hw_name = grid.get("hw", ["nvidia_h100_sxm"])[0] if grid.get("hw") else "nvidia_h100_sxm"

        model = None
        try:
            model = _load_model_spec(model_name)
            model.seq_len = seq_len
        except Exception:
            pass

        system = None
        try:
            hw = load_hw(hw_name)
            gpus_per_node = _inferred_gpus_per_node(hw)
            nodes = _ceil_nodes_for_world_size(target_ws, gpus_per_node)
            system = _system_from_hw(
                hw,
                nodes=nodes,
                gpus_per_node=gpus_per_node,
                world_size_override=target_ws,
                host_mem_gb=grid.get("host_mem_gb", [256.0])[0] if grid.get("host_mem_gb") else 256.0,
            )
        except Exception:
            system = None

        other_combinations = 1
        for k in other_keys:
            other_combinations *= len(grid[k])

        total = 0
        for other_vals in itertools.product(*[grid[k] for k in other_keys]):
            base_config = dict(zip(other_keys, other_vals))
            _apply_total_token_batch_rule(
                base_config, default_seq_len=seq_len, model=model,
            )
            total += sum(1 for _ in self._enumerate_valid_parallel_configs(
                grid, target_ws, model, system, base_config
            ))

        return other_combinations * total if model is None else total

    def generate_static_configs_stream(self) -> Generator[Dict[str, Any], None, None]:
        grid = {k: (v if isinstance(v, list) else [v]) for k, v in self.param_grid.items()}
        _reject_gpus_per_node_config(grid)
        world_sizes = grid.get("world_size", [1])
        if len(world_sizes) > 1:
            raise ValueError("Only single world_size is supported when using 'auto' parallel strategy")
        target_ws = world_sizes[0]

        model_name = grid.get("model", ["unknown"])[0] if grid.get("model") else "unknown"
        seq_len = grid.get("seq_len", [4096])[0] if grid.get("seq_len") else 4096
        model_for_auto = None
        try:
            model_for_auto = _load_model_spec(model_name)
            model_for_auto.seq_len = seq_len
        except Exception:
            pass

        self._expand_auto_values_optimized(grid, target_ws, model_for_auto)

        parallel_keys = ["tp", "cp", "pp", "ep", "dp"]
        
        total_token_vals = grid.get("total_token", [])
        has_total_token = total_token_vals and any(v is not None and v > 0 for v in total_token_vals)
        
        other_keys = [k for k in grid.keys() if k not in parallel_keys and k != "world_size"]
        if has_total_token and "global_batch" in other_keys:
            other_keys.remove("global_batch")

        hw_name = grid.get("hw", ["nvidia_h100_sxm"])[0] if grid.get("hw") else "nvidia_h100_sxm"

        model = None
        system = None
        try:
            model = _load_model_spec(model_name)
            model.seq_len = seq_len
        except FileNotFoundError:
            pass
        except Exception:
            pass
        try:
            hw = load_hw(hw_name)
            gpus_per_node = _inferred_gpus_per_node(hw)
            nodes = _ceil_nodes_for_world_size(target_ws, gpus_per_node)
            system = _system_from_hw(
                hw,
                nodes=nodes,
                gpus_per_node=gpus_per_node,
                world_size_override=target_ws,
                host_mem_gb=grid.get("host_mem_gb", [256.0])[0] if grid.get("host_mem_gb") else 256.0,
            )
        except Exception:
            system = None

        other_grids = [grid[k] for k in other_keys]
        for other_vals in itertools.product(*other_grids):
            base_config = dict(zip(other_keys, other_vals))
            base_config["world_size"] = target_ws

            _apply_total_token_batch_rule(
                base_config, default_seq_len=seq_len, model=model,
            )

            for p_vals in self._enumerate_valid_parallel_configs(
                    grid, target_ws, model, system, base_config
            ):
                config = base_config.copy()
                config.update(dict(zip(parallel_keys, p_vals)))
                yield config

    def generate_static_configs(self) -> List[Dict[str, Any]]:
        return list(self.generate_static_configs_stream())


def _worker_initializer(model_name: str = "deepseek_v3_2"):
    global _WORKER_MODEL_CACHE, _WORKER_HW_CACHE, _WORKER_GRAPH_CACHE
    _WORKER_MODEL_CACHE[(model_name, None)] = _load_model_spec(model_name)
    _WORKER_GRAPH_CACHE.clear()


def _graph_cache_key(config: Dict[str, Any]) -> Tuple[Any, ...]:
    """Return the fields that determine the generated training IR graph.

    The graph depends on model geometry/dtypes and sharding collectives. It
    does not depend on DP/global batch/optimizer/ZeRO, which are consumed by
    the later estimator and memory formulas.
    """
    return (
        config.get("model", "deepseek_v3_2"),
        config.get("quant_preset") or None,
        int(config.get("seq_len", 4096)),
        int(config.get("micro_batch", 1)),
        int(config.get("tp", 1)),
        int(config.get("cp", 1)),
        int(config.get("ep", 1)),
        config.get("cp_kind", "none"),
        config.get("cp_ulysses"),
        config.get("cp_ring"),
    )


def _memory_limit_gb(config: Dict[str, Any], system: SystemSpec) -> float:
    if "max_memory_gb" in config:
        return float(config["max_memory_gb"])
    ratio = float(config.get("memory_limit_ratio", 0.8))
    return float(system.gpu.hbm_gb) * ratio


def _is_memory_feasible(config: Dict[str, Any], model: ModelSpec,
                        system: SystemSpec, strategy: Strategy) -> tuple[bool, float, float]:
    mb = memory_breakdown(None, model, system, strategy)
    memory_gb = mb.total / 1e9
    limit_gb = _memory_limit_gb(config, system)
    return memory_gb <= limit_gb, memory_gb, limit_gb


def run_training_batch_wrapper(configs: List[Dict]) -> List[Optional[Dict]]:
    return [run_training_task_wrapper(config) for config in configs]


def _batched(iterable, batch_size: int):
    batch_size = max(1, int(batch_size))
    batch = []
    for item in iterable:
        batch.append(item)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch


def run_training_task_wrapper(config: Dict) -> Optional[Dict]:
    from zrt.training.ir.builders import build_graph

    model_name = config.get("model", "deepseek_v3_2")
    hw_name = config.get("hw", "nvidia_h100_sxm")
    quant_preset = config.get("quant_preset") or None

    try:
        model_key = (model_name, quant_preset)
        model = _WORKER_MODEL_CACHE.get(model_key)
        if model is None:
            model = _load_model_spec(model_name, quant_preset=quant_preset)
            _WORKER_MODEL_CACHE[model_key] = model
        # Per-config seq_len override: the cached ModelSpec carries the YAML
        # default, but the grid varies seq_len, and Strategy does not own this
        # field. Mutating the cached object is safe — each worker is a single
        # process running one task at a time.
        seq_len = config.get("seq_len")
        if seq_len is not None:
            model.seq_len = int(seq_len)

        # System cache must key on world_size as well as hardware; nodes is
        # derived from world_size and the hardware's innermost tier size.
        sys_key = (
            hw_name,
            int(config.get("world_size", 0)),
        )
        system = _WORKER_HW_CACHE.get(sys_key)
        if system is None:
            system = _make_system_from_config(config)
            _WORKER_HW_CACHE[sys_key] = system

        strategy = _make_strategy_from_config(config)
        strategy.validate(model, system)

        feasible, memory_gb, limit_gb = _is_memory_feasible(config, model, system, strategy)
        if not feasible:
            return {
                "status": "skipped",
                "config": config,
                "type": "memory",
                "memory_gb": memory_gb,
                "memory_limit_gb": limit_gb,
            }
    except Exception as e:
        return {"status": "error", "config": config, "type": "validation_error", "message": str(e)}

    try:
        graph_key = _graph_cache_key(config)
        graph = _WORKER_GRAPH_CACHE.get(graph_key)
        if graph is None:
            graph = build_graph(model, strategy)
            _WORKER_GRAPH_CACHE[graph_key] = graph
        report = estimate(model, system, strategy, graph=graph)

        return {
            "status": "success",
            "config": config,
            "report": report,
            "model_name": model_name,
            "hw_name": hw_name,
        }
    except Exception as e:
        logger.exception(f"Evaluation failed: {e}")
        return {"status": "error", "config": config, "type": "runtime_error"}


def format_results(reports: List[TrainingReport], configs: List[Dict]) -> pd.DataFrame:
    rows = []
    for cfg, report in zip(configs, reports):
        d = cfg.copy()
        if report.memory:
            memory_gb = round(report.memory.total / 1e9, 2)
            hw_name = cfg.get("hw", "nvidia_h100_sxm")
            hw = load_hw(hw_name)
            gpu_capacity_gb = hw.memory.capacity_gb
            # *0.8利用率
            if memory_gb > gpu_capacity_gb * 0.8:
                continue
        else:
            memory_gb = None

        d.update(_comm_domain_columns_from_config(cfg))
        d["compute_time_ms"] = round(report.compute_time_ms, 2)
        d["fwd_compute_ms"] = round(report.fwd_compute_ms, 2)
        d["bwd_compute_ms"] = round(report.bwd_compute_ms, 2)
        d["exposed_comm_ms"] = round(report.exposed_comm_ms, 2)
        d["tp_total_ms"] = round(report.tp_total_ms, 2)
        d["tp_exposed_ms"] = round(report.tp_exposed_ms, 2)
        d["cp_total_ms"] = round(report.cp_total_ms, 2)
        d["cp_exposed_ms"] = round(report.cp_exposed_ms, 2)
        d["ep_total_ms"] = round(report.ep_total_ms, 2)
        d["ep_exposed_ms"] = round(report.ep_exposed_ms, 2)
        d["pp_total_ms"] = round(report.pp_total_ms, 2)
        d["pp_exposed_ms"] = round(report.pp_exposed_ms, 2)
        d["pp_hidden_ms"] = round(report.pp_hidden_ms, 2)
        d["dp_total_ms"] = round(report.dp_total_ms, 2)
        d["dp_exposed_ms"] = round(report.dp_exposed_ms, 2)
        d["optimizer_compute_ms"] = round(report.optimizer_time_ms, 4)
        d["optimizer_comm_ms"] = round(report.optimizer_comm_ms + report.optimizer_comm_hidden_ms, 2)
        d["optimizer_exposed_ms"] = round(report.optimizer_comm_ms, 2)
        d["recompute_time_ms"] = round(report.recompute_time_ms, 3)
        d["recompute_time_raw_ms"] =  round(report.recompute_time_raw_ms, 3)
        d["step_time_ms"] = round(report.step_time_ms, 3)
        d["pipeline_time_ms"] = round(report.pipeline_time_ms, 3)
        d["mfu"] = round(report.mfu, 4)
        d["mfu_native"] = round(report.mfu_native, 4)
        d["hfu"] = round(report.hfu, 4)
        d["bubble_fraction"] = round(report.bubble_fraction, 4)
        d["bubble_time_ms"] = round(report.bubble_time_ms, 2)
        d["tokens_per_sec"] = round(report.tokens_per_sec, 1)
        if report.memory:
            # Per-rank weight/grad/opt-state can drop to single-digit MB after
            # ZeRO-3 + EP + TP sharding; 2-decimal GB rounds those to 0.00 and
            # hides whether sharding is sane. Use 4 decimals (≈100 KB resolution).
            d["weights_gb"] = round(report.memory.weights / 1e9, 4)
            d["grads_gb"] = round(report.memory.grads / 1e9, 4)
            d["opt_state_gb"] = round(report.memory.opt_state / 1e9, 4)
            d["activations_gb"] = round(report.memory.activations / 1e9, 2)
            d["comm_buffers_gb"] = round(report.memory.comm_buffers / 1e9, 2)
            d["memory_gb"] = memory_gb
        rows.append(d)

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("tokens_per_sec", ascending=False)

    metric_cols = ["compute_time_ms", "fwd_compute_ms", "bwd_compute_ms", "exposed_comm_ms",
                   "tp_total_ms", "tp_exposed_ms", "cp_total_ms", "cp_exposed_ms",
                   "ep_total_ms", "ep_exposed_ms", "pp_total_ms", "pp_exposed_ms",
                   "pp_hidden_ms", "dp_total_ms", "dp_exposed_ms",
                   "optimizer_compute_ms", "optimizer_comm_ms", "optimizer_exposed_ms", "recompute_time_ms",
                   "recompute_time_raw_ms", "step_time_ms", "pipeline_time_ms",
                   "mfu", "mfu_native", "hfu", "bubble_fraction", "bubble_time_ms", "tokens_per_sec",
                   "weights_gb", "grads_gb", "opt_state_gb", "activations_gb",
                   "comm_buffers_gb", "memory_gb"]
    config_cols = [k for k in rows[0].keys() if k not in metric_cols] if rows else []
    cols = config_cols + [c for c in metric_cols if c in df.columns]
    df = df[[c for c in cols if c in df.columns]]
    return df


def save_results(df: pd.DataFrame, output_path: str):
    os.makedirs(output_path, exist_ok=True)

    csv_path = os.path.join(output_path, "results_summary.csv")
    df.to_csv(csv_path, index=False)
    logger.info(f"Results saved to: {csv_path}")

    print("=" * 60)
    print("Training Search Results")
    print("=" * 60)
    print(f"Total results: {len(df)} configs")
    print()
    print("Top 5 configs by tokens_per_sec:")
    print(df.head(5).to_string())
    print("=" * 60)


def _first_seen(values: Iterable[Any]) -> List[str]:
    seen: set[str] = set()
    out: List[str] = []
    for value in values:
        text = str(value)
        if text not in seen:
            seen.add(text)
            out.append(text)
    return out


def _normalize_comparison_hw_groups(
        df: pd.DataFrame,
        comparison_hw_groups: Optional[List[List[str]]] = None,
) -> List[List[str]]:
    if comparison_hw_groups:
        return [[str(hw) for hw in group] for group in comparison_hw_groups]
    if "hw" not in df.columns:
        return [[]]
    return [_first_seen(df["hw"])]


def _sequence_order(df: pd.DataFrame, seq_lens: Optional[List[int]] = None) -> List[int]:
    if seq_lens:
        return [int(seq) for seq in seq_lens]
    if "seq_len" not in df.columns:
        return []
    return sorted(int(seq) for seq in df["seq_len"].dropna().unique())


def select_best_configs_by_tokens(
        df: pd.DataFrame,
        *,
        comparison_hw_groups: Optional[List[List[str]]] = None,
        seq_lens: Optional[List[int]] = None,
) -> pd.DataFrame:
    """Pick the highest-token-throughput row for every group + hw + seq_len.

    The same hardware may intentionally appear in multiple comparison groups
    as the baseline. In that case the row is duplicated with a different
    ``_group_idx`` so each group can normalize against its own first hardware.
    """
    required = {"hw", "seq_len", "tokens_per_sec"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Cannot select best configs; missing columns: {missing}")
    if df.empty:
        return df.copy()

    groups = _normalize_comparison_hw_groups(df, comparison_hw_groups)
    seq_order = _sequence_order(df, seq_lens)
    allowed_hw = {hw for group in groups for hw in group}

    best_input = df.copy()
    best_input["hw"] = best_input["hw"].astype(str)
    if allowed_hw:
        best_input = best_input[best_input["hw"].isin(allowed_hw)]
    if seq_order:
        best_input = best_input[best_input["seq_len"].isin(seq_order)]
    if best_input.empty:
        return best_input

    per_hw_seq = (
        best_input.sort_values("tokens_per_sec", ascending=False)
        .groupby(["hw", "seq_len"], as_index=False, sort=False)
        .first()
    )

    row_map = {
        (str(row["hw"]), int(row["seq_len"])): row
        for _, row in per_hw_seq.iterrows()
    }
    seq_rank = {seq: idx for idx, seq in enumerate(seq_order)}
    rows: List[pd.Series] = []
    for group_idx, group in enumerate(groups):
        for seq in seq_order:
            for hw_idx, hw in enumerate(group):
                row = row_map.get((str(hw), int(seq)))
                if row is None:
                    continue
                out = row.copy()
                out["_group_idx"] = group_idx
                out["_hw_idx"] = hw_idx
                out["_seq_idx"] = seq_rank.get(int(seq), 999)
                rows.append(out)

    if not rows:
        return per_hw_seq.iloc[0:0].copy()
    return pd.DataFrame(rows).sort_values(
        ["_group_idx", "_seq_idx", "_hw_idx"]
    ).reset_index(drop=True)


def _with_group_labels(best_df: pd.DataFrame) -> pd.DataFrame:
    out = best_df.copy()
    if "组号" not in out.columns:
        out.insert(0, "组号", "")
    previous_group = None
    for idx, group_idx in enumerate(out.get("_group_idx", pd.Series([0] * len(out)))):
        if group_idx != previous_group:
            out.at[idx, "组号"] = f"对比{int(group_idx) + 1}"
            previous_group = group_idx
        else:
            out.at[idx, "组号"] = ""
    return out


def save_best_configs_csv(best_df: pd.DataFrame, output_path: str) -> str:
    os.makedirs(output_path, exist_ok=True)
    out_path = os.path.join(output_path, "best_tokens_per_hw_seq_len.csv")
    drop_cols = [col for col in best_df.columns if str(col).startswith("_")]
    _with_group_labels(best_df).drop(columns=drop_cols, errors="ignore").to_csv(
        out_path,
        index=False,
    )
    logger.info("Best-token configs saved to: %s", out_path)
    return out_path


def _seq_label(seq_len: Any) -> str:
    seq = int(seq_len)
    if seq % 1_048_576 == 0:
        value = seq // 1_048_576
        return f"{value}m"
    if seq % 1024 == 0:
        return f"{seq // 1024}k"
    return str(seq)


def _safe_float(row: pd.Series, key: str, default: float = 0.0) -> float:
    value = row.get(key, default)
    if pd.isna(value):
        return default
    return float(value)


def _copy_row_style(ws: Any, source_row: int, target_row: int, max_col: int) -> None:
    if source_row > ws.max_row:
        return
    for col_idx in range(1, max_col + 1):
        src = ws.cell(source_row, col_idx)
        dst = ws.cell(target_row, col_idx)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _sheet_headers(ws: Any, fallback: List[str]) -> List[str]:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    headers = [str(header) for header in headers if header is not None]
    return headers or fallback


def _clear_sheet(ws: Any, headers: List[str]) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx).value = header


def _unmerge_group_label_cells(ws: Any, headers: List[str]) -> None:
    if "组号" not in headers:
        return
    group_col = headers.index("组号") + 1
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_col <= group_col <= merged_range.max_col
            and merged_range.max_row >= 2
        ):
            ws.unmerge_cells(str(merged_range))


def _apply_report_fill_styles(ws: Any, headers: List[str], rows: pd.DataFrame) -> None:
    from openpyxl.styles import PatternFill

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=REPORT_HEADER_FILL_COLOR,
    )
    even_group_fill = PatternFill(
        fill_type="solid",
        fgColor=REPORT_EVEN_GROUP_FILL_COLOR,
    )

    for col_idx in range(1, len(headers) + 1):
        ws.cell(1, col_idx).fill = header_fill

    if rows.empty:
        return

    for row_idx, (_, row) in enumerate(rows.iterrows(), start=2):
        group_idx = int(row.get("_group_idx", 0))
        group_number = group_idx + 1
        if group_number % 2 != 0:
            continue
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row_idx, col_idx).fill = even_group_fill


def _merge_group_label_cells(ws: Any, headers: List[str], rows: pd.DataFrame) -> None:
    if "组号" not in headers or rows.empty:
        return

    from openpyxl.styles import Alignment

    group_col = headers.index("组号") + 1

    def center_group_cell(row_idx: int) -> None:
        cell = ws.cell(row_idx, group_col)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            text_rotation=cell.alignment.text_rotation,
            wrap_text=cell.alignment.wrap_text,
            shrink_to_fit=cell.alignment.shrink_to_fit,
            indent=cell.alignment.indent,
        )

    _unmerge_group_label_cells(ws, headers)

    group_values = list(rows.get("_group_idx", pd.Series([0] * len(rows))))
    start_row = 2
    current_group = group_values[0]
    for offset, group_idx in enumerate(group_values[1:], start=1):
        if group_idx == current_group:
            continue
        end_row = 1 + offset
        if end_row > start_row:
            ws.merge_cells(
                start_row=start_row,
                start_column=group_col,
                end_row=end_row,
                end_column=group_col,
            )
            center_group_cell(start_row)
        start_row = 2 + offset
        current_group = group_idx

    end_row = 1 + len(group_values)
    if end_row > start_row:
        ws.merge_cells(
            start_row=start_row,
            start_column=group_col,
            end_row=end_row,
            end_column=group_col,
        )
    center_group_cell(start_row)


def _write_raw_data_sheet(ws: Any, best_df: pd.DataFrame) -> None:
    headers = _sheet_headers(ws, RAW_DATA_HEADERS)
    _clear_sheet(ws, headers)
    rows = _with_group_labels(best_df)
    for row_idx, (_, row) in enumerate(rows.iterrows(), start=2):
        if row_idx > 2:
            _copy_row_style(ws, 2, row_idx, len(headers))
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, None)
            if pd.isna(value):
                value = None
            ws.cell(row_idx, col_idx).value = value
    _apply_report_fill_styles(ws, headers, rows)
    _merge_group_label_cells(ws, headers, rows)


def _baseline_tokens_by_group_seq(
        best_df: pd.DataFrame,
        comparison_hw_groups: List[List[str]],
) -> Dict[Tuple[int, int], float]:
    baselines: Dict[Tuple[int, int], float] = {}
    for group_idx, group in enumerate(comparison_hw_groups):
        if not group:
            continue
        baseline_hw = str(group[0])
        rows = best_df[
            (best_df["_group_idx"] == group_idx)
            & (best_df["hw"].astype(str) == baseline_hw)
        ]
        for _, row in rows.iterrows():
            baselines[(group_idx, int(row["seq_len"]))] = _safe_float(row, "tokens_per_sec")
    return baselines


def _analysis_value(
        header: str,
        row: pd.Series,
        baseline_tokens: Optional[float],
) -> Any:
    step_time = _safe_float(row, "step_time_ms")
    if header == "组号":
        return row.get("组号", "")
    if header == "硬件+seq":
        return f"{row.get('hw')}_{_seq_label(row.get('seq_len'))}"
    if header in RAW_TO_ANALYSIS:
        value = row.get(RAW_TO_ANALYSIS[header], None)
        return None if pd.isna(value) else value
    if header == "计算时间":
        return (
            _safe_float(row, "fwd_compute_ms")
            + _safe_float(row, "bwd_compute_ms")
            + _safe_float(row, "recompute_time_ms")
        )
    if header == "集群吞吐归一化":
        current = _safe_float(row, "tokens_per_sec")
        return round(current / baseline_tokens, 3) if baseline_tokens else None
    if step_time <= 0:
        return None
    if header == "计算占比":
        return _analysis_value("计算时间", row, baseline_tokens) / step_time
    if header == "TP通信占比":
        return _safe_float(row, "tp_exposed_ms") / step_time
    if header == "EP通信占比":
        return _safe_float(row, "ep_exposed_ms") / step_time
    if header == "PP通信占比":
        return _safe_float(row, "pp_exposed_ms") / step_time
    if header == "DP通信占比":
        return _safe_float(row, "dp_exposed_ms") / step_time
    if header == "CP通信占比":
        return _safe_float(row, "cp_exposed_ms") / step_time
    if header == "优化器占比":
        return _safe_float(row, "optimizer_compute_ms") / step_time
    if header == "空泡占比":
        return _safe_float(row, "bubble_time_ms") / step_time
    return None


def _write_analysis_sheet(
        ws: Any,
        best_df: pd.DataFrame,
        comparison_hw_groups: List[List[str]],
) -> None:
    headers = _sheet_headers(ws, ANALYSIS_HEADERS)
    _clear_sheet(ws, headers)
    rows = _with_group_labels(best_df)
    baselines = _baseline_tokens_by_group_seq(best_df, comparison_hw_groups)
    for row_idx, (_, row) in enumerate(rows.iterrows(), start=2):
        if row_idx > 2:
            _copy_row_style(ws, 2, row_idx, len(headers))
        baseline = baselines.get((int(row["_group_idx"]), int(row["seq_len"])))
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = _analysis_value(header, row, baseline)
            if header in PERCENT_ANALYSIS_HEADERS:
                cell.number_format = "0.00%"
            elif header == "集群吞吐归一化":
                cell.number_format = "0.000"
            elif isinstance(cell.value, float):
                cell.number_format = "0.00"
    _apply_report_fill_styles(ws, headers, rows)
    _merge_group_label_cells(ws, headers, rows)


def export_best_analysis_excel(
        best_df: pd.DataFrame,
        output_path: str,
        *,
        comparison_hw_groups: List[List[str]],
        template_path: Optional[str] = None,
        filename: str = "best_config_analysis.xlsx",
) -> Optional[str]:
    """Export the filtered best rows into raw_data and analysis sheets."""
    if best_df.empty:
        return None

    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        logger.warning("openpyxl is unavailable; skip best analysis Excel export")
        return None

    if template_path:
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = Workbook()
        wb.active.title = "analysis"
        wb.create_sheet("raw_data")

    if "analysis" not in wb.sheetnames:
        wb.create_sheet("analysis", 0)
    if "raw_data" not in wb.sheetnames:
        wb.create_sheet("raw_data")

    _write_raw_data_sheet(wb["raw_data"], best_df)
    _write_analysis_sheet(wb["analysis"], best_df, comparison_hw_groups)

    os.makedirs(output_path, exist_ok=True)
    out_path = os.path.join(output_path, filename)
    wb.save(out_path)
    logger.info("Best analysis Excel exported: %s", out_path)
    print(f"Best analysis Excel: {out_path}")
    return out_path


def _default_analysis_excel_name(model_name: str) -> str:
    safe_model = str(model_name).replace("/", "_").replace("\\", "_").replace(":", "_")
    return f"{safe_model}_best_config_analysis.xlsx"


def export_best_configs_excel(
        all_results: List[Dict],
        output_path: str
) -> None:
    from zrt.training.ir.builders import build_graph
    from zrt.training.models.flops import op_cost as _op_cost
    from zrt.training.io.excel_exporter import export_estimate_excel

    if not all_results:
        return

    df = pd.DataFrame([
        {
            "model": r["model_name"],
            "hw": r["hw_name"],
            "seq_len": r["config"].get("seq_len", 4096),
            "world_size": r["config"].get("world_size", 1),
            "mfu": r["report"].mfu,
            "step_time_ms": r["report"].step_time_ms,
            "config": r["config"],
            "report": r["report"],
        }
        for r in all_results
    ])

    grouped = df.groupby(["model", "hw", "seq_len", "world_size"])

    for (model_name, hw_name, seq_len, world_size), group in grouped:
        best_row = group.loc[group["step_time_ms"].idxmin()]
        best_config = best_row["config"]
        best_report = best_row["report"]

        model = _load_model_spec(
            model_name,
            quant_preset=best_config.get("quant_preset") or None,
        )
        model.seq_len = seq_len

        hw = load_hw(hw_name)
        _reject_gpus_per_node_config(best_config)
        gpus_per_node = _inferred_gpus_per_node(hw)
        nodes = _ceil_nodes_for_world_size(world_size, gpus_per_node)
        system = _system_from_hw(
            hw,
            nodes=nodes,
            gpus_per_node=gpus_per_node,
            world_size_override=world_size,
            host_mem_gb=best_config.get("host_mem_gb", 256.0),
        )

        strategy = _make_strategy_from_config(best_config)

        graph = build_graph(model, strategy)
        op_costs = {}
        for op in graph.ops:
            op_costs[op.name] = _op_cost(op, model, system)

        excel_name = f"{model_name}_{hw_name}_seq{seq_len}_ws{world_size}_best.xlsx"
        excel_path = os.path.join(output_path, excel_name)

        export_estimate_excel(
            report=best_report,
            graph=graph,
            model=model,
            system=system,
            strategy=strategy,
            op_costs=op_costs,
            output_path=excel_path,
        )
        logger.info(f"Best config Excel exported: {excel_path}")
        print(
            f"Best config Excel: {excel_path} "
            f"(step_time={best_report.step_time_ms:.3f}ms, MFU={best_report.mfu:.4%})"
        )


def run_training_search_parallel(
        param_grid: Dict[str, List[Any]],
        workers: int = 8,
        mfu_threshold: float = 0.0,
        batch_size: int = 32,
        export_best_excel: bool = True,
        export_analysis_excel: bool = True,
        analysis_excel_template: Optional[str] = None,
        analysis_excel_name: Optional[str] = None,
        comparison_hw_groups: Optional[List[List[str]]] = None,
) -> pd.DataFrame:
    model_name = param_grid.get("model", ["unknown"])
    if isinstance(model_name, list):
        model_name = model_name[0] if model_name else "unknown"

    manager = TrainingConfigManager(
        param_grid=param_grid,
    )

    total_configs = manager.count_total_configs()
    adjusted_workers = min(workers, os.cpu_count() or 8)

    if os.path.exists(manager.output_path):
        for f in os.listdir(manager.output_path):
            fpath = os.path.join(manager.output_path, f)
            if os.path.isfile(fpath):
                os.remove(fpath)
            elif os.path.isdir(fpath):
                import shutil
                shutil.rmtree(fpath)
    os.makedirs(manager.output_path, exist_ok=True)

    start_time = time.time()
    logger.info(f"Starting search: total_configs={total_configs}, workers={adjusted_workers}")

    all_results: List[Dict] = []
    error_count = 0
    skipped_count = 0

    config_generator = manager.generate_static_configs_stream()
    futures_map = {}

    with ProcessPoolExecutor(
            max_workers=adjusted_workers,
            initializer=_worker_initializer,
            initargs=(model_name,),
    ) as executor:
        with tqdm(total=total_configs, desc="Evaluating configs", unit="config") as pbar:
            for batch in _batched(config_generator, batch_size):
                fut = executor.submit(run_training_batch_wrapper, batch)
                futures_map[fut] = len(batch)

                while len(futures_map) >= adjusted_workers * 2:
                    done, _ = concurrent.futures.wait(
                        futures_map.keys(), return_when=concurrent.futures.FIRST_COMPLETED
                    )
                    for fut in done:
                        n = futures_map.pop(fut)
                        pbar.update(n)
                        for res in fut.result():
                            if res and res["status"] == "success":
                                all_results.append(res)
                            elif res and res["status"] == "error":
                                error_count += 1
                            elif res and res["status"] == "skipped":
                                skipped_count += 1

            while futures_map:
                done, _ = concurrent.futures.wait(futures_map.keys())
                for fut in done:
                    n = futures_map.pop(fut)
                    pbar.update(n)
                    for res in fut.result():
                        if res and res["status"] == "success":
                            all_results.append(res)
                        elif res and res["status"] == "error":
                            error_count += 1
                        elif res and res["status"] == "skipped":
                            skipped_count += 1

    elapsed = time.time() - start_time
    logger.info(
        f"Search completed in {elapsed:.2f} seconds, success={len(all_results)}, "
        f"skipped={skipped_count}, errors={error_count}"
    )

    if not all_results:
        logger.error("No valid configurations found.")
        return pd.DataFrame()

    # Memory-feasibility filter: drop configs whose per-rank memory exceeds
    # 80% of GPU HBM capacity. Apply before BOTH the CSV table and the Excel
    # "best" export so the latter cannot pick infeasible configs.
    feasible_results: List[Dict] = []
    for r in all_results:
        rep = r["report"]
        hw_name = r.get("hw_name") or r["config"].get("hw", "nvidia_h100_sxm")
        cap_gb = load_hw(hw_name).memory.capacity_gb
        if rep.memory is None:
            feasible_results.append(r)
            continue
        if rep.memory.total / 1e9 <= cap_gb * 0.8:
            feasible_results.append(r)
    logger.info(
        f"Memory feasibility: {len(feasible_results)}/{len(all_results)} configs "
        f"fit within 0.8 × HBM"
    )
    if not feasible_results:
        logger.error("No memory-feasible configurations found.")
        return pd.DataFrame()

    all_reports = [r["report"] for r in feasible_results]
    all_configs = [r["config"] for r in feasible_results]
    all_df = format_results(all_reports, all_configs)

    filtered_df = all_df[all_df["mfu"] > mfu_threshold] if mfu_threshold > 0 else all_df
    if filtered_df.empty:
        logger.warning(f"No results with MFU > {mfu_threshold}")
        return pd.DataFrame()

    save_results(filtered_df, manager.output_path)

    hw_values = param_grid.get("hw", [])
    default_hw_groups = (
        [[str(hw) for hw in hw_values]]
        if isinstance(hw_values, list)
        else [[str(hw_values)]]
    )
    seq_values = param_grid.get("seq_len", [])
    default_seq_order = (
        [int(seq) for seq in seq_values]
        if isinstance(seq_values, list)
        else None
    )
    active_hw_groups = comparison_hw_groups or default_hw_groups
    if {"hw", "seq_len", "tokens_per_sec"}.issubset(filtered_df.columns):
        best_df = select_best_configs_by_tokens(
            filtered_df,
            comparison_hw_groups=active_hw_groups,
            seq_lens=default_seq_order,
        )
        if not best_df.empty:
            if export_analysis_excel:
                export_best_analysis_excel(
                    best_df,
                    manager.output_path,
                    comparison_hw_groups=active_hw_groups,
                    template_path=analysis_excel_template,
                    filename=analysis_excel_name or _default_analysis_excel_name(model_name),
                )

    if export_best_excel:
        filtered_indices = [int(i) for i in filtered_df.index]
        filtered_results = [feasible_results[i] for i in filtered_indices]
        export_best_configs_excel(filtered_results, manager.output_path)

    if not filtered_df.empty:
        best = filtered_df.iloc[0]
        print("\nTop1 Result:")
        print(f"  MFU: {best['mfu']:.4f}, step_time: {best['step_time_ms']:.2f}ms")

    return filtered_df


if __name__ == "__main__":
    multiprocessing.set_start_method("spawn", force=True)

    training_param_grid = {
        "model": ["deepseek_v4_pro"],
        "hw": ["nvidia_b300", "nvidia_gb300_nvl576", "ascend_910c"],
        "world_size": [8192],
        "tp": [1, 2, 4, 8, 16, 32, 64, 128],
        "cp": [1, 2, 4, 8, 16, 32, 64, 128],
        "pp": [1, 2, 4, 8, 16],
        # EP must divide DP under the current expert-DP sharding model.
        "ep": [32, 64, 128],
        "dp": "auto",
        "micro_batch": [1],
        # Derived per seq_len as exact total_token // seq_len.
        "seq_len": [262144, 524288, 1048576],
        "total_token": [536870912],
        "zero_stage": [3],
        "pp_schedule": ["dualpipev"],
        "cp_kind": ["ulysses"],
        "vpp_chunks": [1, 2, 4],
        "tp_overlap": ["coc"],
        "ep_overlap": [True],
        "dualbatch": [True],
        "dp_overlap_in_bubble": [True],
        "recompute": ["none", "mhc", "full", "selective"],
        "optimizer": ["muon"],
        "quant_preset": ["deepseek_v4_fp8_fp4"],
    }
    comparison_hw_groups = [
        ["nvidia_b300", "nvidia_gb300_nvl576"],
        ["nvidia_b300", "ascend_910c"],
    ]

    df = run_training_search_parallel(
        param_grid=training_param_grid,
        workers=32,
        mfu_threshold=0.00,
        export_best_excel=True,
        comparison_hw_groups=comparison_hw_groups,
    )
