"""Excel exporter for spec-based training estimation (--estimate-config).

Produces a single workbook with five sheets:
  1. Summary   — E2E timing, MFU, memory, pipeline, FLOPs
  2. Ops       — Per-op details (name, kind, layer, flops, bytes, bound, latency)
  3. Model     — Model geometry, MoE, HC, MLA, dtype
  4. Hardware  — GPU specs, network, memory bandwidth
  5. Strategy  — Parallel config, micro_batch, recompute, zero_stage
"""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from zrt.training.spec.dtype import Dtype

if TYPE_CHECKING:
    from zrt.training.ir.training_graph import Graph, Op
    from zrt.training.models.flops import OpCost
    from zrt.training.spec.model import ModelSpec
    from zrt.training.spec.strategy import Strategy
    from zrt.training.spec.system import SystemSpec
    from zrt.training.spec.report import TrainingReport


def export_estimate_excel(
    *,
    report: "TrainingReport",
    graph: "Graph",
    model: "ModelSpec",
    system: "SystemSpec",
    strategy: "Strategy",
    op_costs: dict[str, "OpCost"],
    output_path: str | Path,
) -> Path:
    """Write the estimation report to a multi-sheet Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Styles ───────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_sheet(ws, rows, col_widths=None):
        """Write rows to a worksheet.
        Headers (row 1) are centered. Data cells are left-aligned.
        Empty rows (all blank) get NO style, acting as visual spacers.
        """
        for r_idx, row in enumerate(rows, 1):
            is_empty = all(v in (None, "") for v in row)
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                elif is_empty:
                    # Spacer row: explicitly clear borders/fill so it acts as whitespace
                    cell.border = Border()
                    cell.fill = PatternFill()
                else:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left")
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ["Metric", "Value", "Unit"],
        # Timing
        ["Step Time", f"{report.step_time_ms:.2f}", "ms"],
        ["Per-Stage Time", f"{report.per_stage_ms:.2f}" if report.per_stage_ms > 0 else "N/A", "ms"],
        ["", "", ""],
        ["Per-Stage Details", "", ""],
    ]
    if report.per_stage:
        for i, st in enumerate(report.per_stage):
            summary_rows.append([
                f"  Stage {i}",
                f"fwd={st.fwd*1000:.2f}ms  bwd={st.bwd*1000:.2f}ms  total={((st.fwd+st.bwd)*1000):.2f}ms",
                "",
            ])
    summary_rows.extend([
        ["", "", ""],
        ["", "", ""],
        ["Step Time Breakdown", "", ""],
        ["  Warmup", f"{report.warmup_ms:.2f}" if report.warmup_ms >= 0 else "N/A", "ms"],
        ["  Steady", f"{report.steady_ms:.2f}" if report.steady_ms >= 0 else "N/A", "ms"],
        ["  Cooldown", f"{report.cooldown_ms:.2f}" if report.cooldown_ms >= 0 else "N/A", "ms"],
        ["", "", ""],
        ["Phase Breakdown", "FWD (ms)", "BWD (ms)"],
        ["  Warmup", f"{report.warmup_fwd_ms:.2f}" if report.warmup_fwd_ms >= 0 else "N/A",
         f"{report.warmup_bwd_ms:.2f}" if report.warmup_bwd_ms >= 0 else "N/A"],
        ["  Steady", f"{report.steady_fwd_ms:.2f}" if report.steady_fwd_ms >= 0 else "N/A",
         f"{report.steady_bwd_ms:.2f}" if report.steady_bwd_ms >= 0 else "N/A"],
        ["  Cooldown", f"{report.cooldown_fwd_ms:.2f}" if report.cooldown_fwd_ms >= 0 else "N/A",
         f"{report.cooldown_bwd_ms:.2f}" if report.cooldown_bwd_ms >= 0 else "N/A"],
        ["", "", ""],
        ["Per-Microbatch (steady)", "", ""],
        ["  FWD", f"{report.steady_fwd_per_mb_ms:.2f}" if report.steady_fwd_per_mb_ms >= 0 else "N/A", "ms"],
        ["  BWD", f"{report.steady_bwd_per_mb_ms:.2f}" if report.steady_bwd_per_mb_ms >= 0 else "N/A", "ms"],
        ["  Total", f"{report.steady_per_mb_ms:.2f}" if report.steady_per_mb_ms >= 0 else "N/A", "ms"],
        ["", "", ""],
        ["DP AR Exposed", f"{report.dp_exposed_ms:.2f}" if report.dp_exposed_ms >= 0 else "N/A", "ms"],
        ["Optimizer Time", f"{report.optimizer_time_ms:.2f}" if report.optimizer_time_ms >= 0 else "N/A", "ms"],
        ["Optimizer Comm", f"{report.optimizer_comm_ms:.2f}" if report.optimizer_comm_ms >= 0 else "N/A", "ms"],
        ["", "", ""],
        # Compute / comm breakdown
        ["Compute / Comm Breakdown", "", ""],
        ["  Pipeline Time", f"{report.pipeline_time_ms:.2f}" if report.pipeline_time_ms > 0 else "-", "ms"],
        ["  Compute Time", f"{report.compute_time_ms:.2f}" if report.compute_time_ms > 0 else "-", "ms"],
        ["  Exposed Comm", f"{report.exposed_comm_ms:.2f}" if report.exposed_comm_ms > 0 else "-", "ms"],
        ["    TP (RS/AG)", f"{report.tp_exposed_ms:.2f}" if report.tp_exposed_ms > 0 else "-", "ms"],
        ["    CP (A2A)", f"{report.cp_exposed_ms:.2f}" if report.cp_exposed_ms > 0 else "-", "ms"],
        ["    EP (A2A)", f"{report.ep_exposed_ms:.2f}" if report.ep_exposed_ms > 0 else "-", "ms"],
        ["    PP (P2P)", f"{report.pp_exposed_ms:.2f}" if report.pp_exposed_ms > 0 else "-", "ms"],
        ["    DP (AR/RS)", f"{report.dp_exposed_ms:.2f}" if report.dp_exposed_ms > 0 else "-", "ms"],
        ["  Hidden Comm", f"{report.hidden_comm_ms:.2f}" if report.hidden_comm_ms > 0 else "-", "ms"],
        ["    DP hidden", f"{report.dp_hidden_ms:.2f}" if report.dp_hidden_ms > 0 else "-", "ms"],
        ["    TP hidden", f"{report.tp_hidden_ms:.2f}" if report.tp_hidden_ms > 0 else "-", "ms"],
        ["    EP hidden", f"{report.ep_hidden_ms:.2f}" if report.ep_hidden_ms > 0 else "-", "ms"],
        ["  Total Comm Vol", f"{report.total_comm_volume_ms:.2f}" if report.total_comm_volume_ms > 0 else "-", "ms"],
        ["  Exposed Ratio", f"{report.exposed_comm_ms / report.pipeline_time_ms * 100:.1f}%" if report.pipeline_time_ms > 0 and report.exposed_comm_ms > 0 else "-", "% of pipeline"],
        ["", "", ""],
        # Efficiency
        ["MFU", f"{report.mfu:.4%}" if report.mfu > 0 else "0.00%", ""],
        ["HFU", f"{report.hfu:.4%}" if report.hfu > 0 else "0.00%", ""],
        ["", "", ""],
        # FLOPs
        ["Total FLOPs/Step", f"{report.total_flops/1e12:.2f}", "TFLOPs"],
        ["Forward FLOPs", f"{report.forward_flops/1e12:.2f}" if report.forward_flops > 0 else "N/A", "TFLOPs"],
        ["Backward FLOPs", f"{report.backward_flops/1e12:.2f}" if report.backward_flops > 0 else "N/A", "TFLOPs"],
        ["", "", ""],
        # Derived FLOPs metrics
        ["FLOPs/Token", f"{report.flops_per_token:.2e}" if report.flops_per_token > 0 else "N/A", ""],
        ["Tokens/Second", f"{report.tokens_per_sec:.0f}" if report.tokens_per_sec > 0 else "N/A", ""],
        ["Effective Params", f"{report.effective_params/1e9:.2f}B" if report.effective_params > 0 else "N/A", ""],
        ["Total Params", f"{model.total_params()/1e9:.2f}B", ""],
        ["", "", ""],
        # Memory
        ["Memory (per GPU)", "", ""],
    ])
    if report.memory:
        for k, v in report.memory.to_gb().items():
            summary_rows.append([f"  {k}", f"{v:.2f}", "GB"])
    elif report.memory_breakdown:
        for k, v in report.memory_breakdown.items():
            summary_rows.append([f"  {k}", f"{v/1e9:.2f}", "GB"])

    summary_rows.extend([
        ["", "", ""],
        ["Pipeline", "", ""],
        ["Schedule", report.schedule_name, ""],
        ["Bubble Fraction", f"{report.bubble_fraction:.1%}", ""],
        ["Warmup Steps", str(report.warmup_steps) if report.warmup_steps > 0 else "N/A", ""],
        ["Steady Steps", str(report.steady_steps) if report.steady_steps > 0 else "N/A", ""],
        ["Cooldown Steps", str(report.cooldown_steps) if report.cooldown_steps > 0 else "N/A", ""],
    ])
    if report.warnings:
        summary_rows.extend([
            ["", "", ""],
            ["Warnings", "; ".join(report.warnings), ""],
        ])
    _write_sheet(ws, summary_rows, col_widths=[30, 30, 15])

    # ── Sheet 2: Ops ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Ops")
    _LAYER_KIND_MAP = {
        "dense": "Dense", "moe": "MoE", "mtp": "MTP",
    }
    from zrt.training.compose.stage import has_heterogeneous_compute, _cost_phase_time
    _act_dtype = model.act_dtype if hasattr(model, "act_dtype") else Dtype.BF16

    def _fmt_num(x: float | None) -> str:
        """Format large numbers as scientific (e.g. 2.28e13) or '-' if empty."""
        if x is None or x <= 0:
            return "-"
        return f"{x:.2e}"

    op_rows = [
        ["#", "Op Name", "Kind", "Layer", "Layer Kind",
         "Fwd FLOPs", "Bwd FLOPs", "Total FLOPs",
         "Bound", "Fwd Bytes", "Bwd Bytes", "Latency (μs)"],
    ]
    for idx, op in enumerate(graph.ops, 1):
        cost = op_costs.get(op.name)
        if cost is None:
            from zrt.training.models.flops import op_cost as _op_cost
            cost = _op_cost(op, model)

        fwd_flops = cost.fwd_flops
        bwd_flops = cost.dx_flops + cost.dw_flops
        total_flops = fwd_flops + bwd_flops
        layer_k = _LAYER_KIND_MAP.get(op.layer_kind.value, op.layer_kind.value)

        # Compute latency from FLOPs/bytes using roofline model (hetero-aware)
        gpu_name = system.gpu.name
        overlap = system.gpu.overlap_ratio.get(op.kind, 0.0) if has_heterogeneous_compute(system) else 0.0
        fwd_time = _cost_phase_time(cost, "fwd", system, gpu_name, overlap)
        bwd_time = (_cost_phase_time(cost, "dx", system, gpu_name, overlap)
                    + _cost_phase_time(cost, "dw", system, gpu_name, overlap))
        latency_us = (fwd_time + bwd_time) * 1e6

        op_rows.append([
            idx,
            op.name,
            op.kind,
            op.layer_id,
            layer_k,
            _fmt_num(fwd_flops),
            _fmt_num(bwd_flops),
            _fmt_num(total_flops),
            cost.bound,
            _fmt_num(cost.fwd_bytes),
            _fmt_num(cost.dx_bytes + cost.dw_bytes),
            f"{latency_us:.2f}" if latency_us > 0 else "-",
        ])
    _write_sheet(ws2, op_rows, col_widths=[5, 35, 18, 6, 12, 18, 18, 18, 10, 16, 16, 14])

    # ── Sheet 3: Model ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Model")
    dense_count = sum(1 for l in model.layers if l.value == "dense")
    moe_count = sum(1 for l in model.layers if l.value == "moe")
    mtp_count = sum(1 for l in model.layers if l.value == "mtp")
    model_rows = [
        ["Parameter", "Value", "Unit"],
        # Core geometry
        ["Hidden Size", model.hidden, ""],
        ["FFN Intermediate", model.ffn, ""],
        ["Num Attention Heads", model.num_heads, ""],
        ["Num KV Heads", model.num_kv_heads, ""],
        ["Head Dim", model.head_dim, ""],
        ["Vocab Size", model.vocab, ""],
        ["Seq Len", model.seq_len, ""],
        # Layer composition
        ["", "", ""],
        ["Total Layers", len(model.layers), ""],
        ["  Dense Layers", dense_count, ""],
        ["  MoE Layers", moe_count, ""],
        ["  MTP Layers", mtp_count, ""],
        # MLA
        ["", "", ""],
        ["Q LoRA Rank", model.q_lora_rank if model.q_lora_rank > 0 else "N/A", ""],
        ["KV LoRA Rank", model.kv_lora_rank if model.kv_lora_rank > 0 else "N/A", ""],
        ["QK Nope Head Dim", model.qk_nope_head_dim if model.qk_nope_head_dim > 0 else "N/A", ""],
        ["QK Rope Head Dim", model.qk_rope_head_dim if model.qk_rope_head_dim > 0 else "N/A", ""],
        ["V Head Dim", model.v_head_dim if model.v_head_dim > 0 else "N/A", ""],
        # MoE
        ["", "", ""],
        ["Num Experts", model.num_experts if model.num_experts > 0 else "N/A", ""],
        ["Top K", model.top_k if model.top_k > 0 else "N/A", ""],
        ["MoE FFN", model.moe_ffn if model.moe_ffn > 0 else "N/A", ""],
        ["Shared Experts", model.n_shared_experts, ""],
        ["Capacity Factor", model.capacity_factor, ""],
        ["Expert Imbalance", model.expert_imbalance, ""],
        ["Scoring Func", model.scoring_func, ""],
        ["Routed Expert Dtype", model.routed_expert_dtype, ""],
        # HC
        ["", "", ""],
        ["HC Mult", model.hc_mult, ""],
        ["HC Sinkhorn Iters", model.hc_sinkhorn_iters, ""],
        # Attention compression
        ["", "", ""],
        ["Attn Compression Ratio", model.attn_compression_ratio, ""],
        ["CSA Layers", model.num_csa_layers if model.num_csa_layers > 0 else "N/A", ""],
        ["HCA Layers", model.num_hca_layers if model.num_hca_layers > 0 else "N/A", ""],
        ["SWA-only Layers", model.num_swa_only_layers if model.num_swa_only_layers > 0 else "N/A", ""],
        ["SWA Window", model.swa_window if model.swa_window > 0 else "N/A", ""],
        ["Compress Ratios", str(model.compress_ratios) if model.compress_ratios else "N/A", ""],
        # Indexer
        ["", "", ""],
        ["Index N Heads", model.index_n_heads if model.index_n_heads > 0 else "N/A", ""],
        ["Index Head Dim", model.index_head_dim if model.index_head_dim > 0 else "N/A", ""],
        ["Index TopK", model.index_topk if model.index_topk > 0 else "N/A", ""],
        # O-LoRA
        ["", "", ""],
        ["O LoRA Rank", model.o_lora_rank if model.o_lora_rank > 0 else "N/A", ""],
        ["O Groups", model.o_groups if model.o_groups > 0 else "N/A", ""],
        # Dtypes
        ["", "", ""],
        ["Param Dtype", model.param_dtype.value, ""],
        ["Grad Dtype", model.grad_dtype.value, ""],
        ["Master Dtype", model.master_dtype.value, ""],
        ["Act Dtype", model.act_dtype.value, ""],
        # MTP
        ["", "", ""],
        ["MTP Depth", model.mtp_depth if model.mtp_depth > 0 else "N/A", ""],
    ]
    _write_sheet(ws3, model_rows, col_widths=[30, 45, 10])

    # ── Sheet 4: Hardware ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Hardware")
    gpu = system.gpu
    hw_rows = [
        ["Parameter", "Value", "Unit"],
        ["GPU", gpu.name, ""],
        ["Compute Peak TFLOPs (BF16)", f"{gpu.flops_bf16:.1f}" if gpu.flops_bf16 else "N/A", "TFLOPs"],
        ["Compute Peak TFLOPs (FP8)", f"{gpu.flops_fp8:.1f}" if gpu.flops_fp8 else "N/A", "TFLOPs"],
        ["HBM Capacity", f"{gpu.hbm_gb:.1f}" if gpu.hbm_gb else "N/A", "GB"],
        ["HBM Bandwidth", f"{gpu.hbm_bw_gbps:.1f}" if gpu.hbm_bw_gbps else "N/A", "GB/s"],
        ["World Size", system.world_size, ""],
        ["GPUs per Node", system.gpus_per_node, ""],
        ["Num Nodes", system.nodes, ""],
        ["Host Memory", f"{system.host_mem_gb:.1f}" if system.host_mem_gb else "N/A", "GB"],
        # Network
        ["", "", ""],
        ["Network Tiers", "", ""],
    ]
    for net in system.nets:
        hw_rows.extend([
            [f"  {net.scope}", f"{net.bw_gbps:.1f}", "GB/s"],
            [f"    Latency", f"{net.latency_us:.1f}", "μs"],
            [f"    Topology", net.topology, ""],
        ])
    _write_sheet(ws4, hw_rows, col_widths=[35, 35, 10])

    # ── Sheet 5: Strategy ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Strategy")
    strat_rows = [
        ["Parameter", "Value", "Unit"],
        ["TP", strategy.tp, ""],
        ["CP", strategy.cp, ""],
        ["PP", strategy.pp, ""],
        ["EP", strategy.ep, ""],
        ["DP", strategy.dp, ""],
        ["Total Parallelism", strategy.tp * strategy.cp * strategy.pp * strategy.ep * strategy.dp, ""],
        ["", "", ""],
        ["PP Schedule", strategy.pp_schedule.value, ""],
        ["VPP Chunks", strategy.vpp_chunks, ""],
        ["CP Kind", strategy.cp_kind.value, ""],
        ["", "", ""],
        ["Micro Batch", strategy.micro_batch, ""],
        ["Global Batch", strategy.global_batch, ""],
        ["Num Microbatches", strategy.num_microbatches(), ""],
        ["", "", ""],
        ["Zero Stage", strategy.zero_stage, ""],
        ["Optimizer", strategy.optimizer.value, ""],
        ["TP Overlap", strategy.tp_overlap.value, ""],
        ["EP Overlap", str(strategy.ep_overlap), ""],
        ["Dual-Batch Overlap", str(strategy.dualbatch), ""],
        ["", "", ""],
        ["Recompute", "", ""],
    ]
    rc = strategy.recompute
    if rc.per_layer:
        for lk, cats in rc.per_layer.items():
            strat_rows.append([f"  {lk}", ", ".join(sorted(cats)), ""])
    else:
        strat_rows.append(["  None", "", ""])
    # Offload
    strat_rows.extend([
        ["", "", ""],
        ["Offload", "", ""],
    ])
    ol = strategy.offload
    if hasattr(ol, 'per_layer') and ol.per_layer:
        for lk, cats in ol.per_layer.items():
            strat_rows.append([f"  {lk}", ", ".join(sorted(cats)), ""])
    else:
        strat_rows.append(["  None", "", ""])
    _write_sheet(ws5, strat_rows, col_widths=[30, 45, 10])

    wb.save(str(output_path))
    return output_path
