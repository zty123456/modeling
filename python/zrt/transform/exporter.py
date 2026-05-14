"""Export transformed OpGraph to Excel, JSON, and ONNX with parallelism annotations."""
from __future__ import annotations

import json
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from python.zrt.ir.graph import OpGraph
from python.zrt.ir.node import OpNode
from python.zrt.ir.param_count import op_short
from python.zrt.transform.context import TransformContext, ParallelConfig

logger = logging.getLogger(__name__)


_PRE_LAYER_KEY = -1_000_000
_POST_LAYER_KEY = 1_000_000
_UNKNOWN_LAYER_BASE = 999_999


def _parse_layer(layer: str) -> int | None:
    if not layer:
        return None
    try:
        return int(layer)
    except (ValueError, TypeError):
        return None


def _effective_layer_map(graph, nodes: list) -> dict[str, float]:
    """Assign each node a sortable float representing its execution-order layer.

    Algorithm:

    1. Numeric ``layer`` → that integer (seed).
    2. Phase-aware predecessor BFS: untagged nodes inherit ``max(prev) + 0.5``
       — only same-phase edges are followed.  Stitched fwd→bwd cross-edges
       are ignored, otherwise a fwd head op would be pulled into the bwd
       subgraph's layer range.
    3. Phase-aware successor BFS: ``min(succ) - 0.5`` for the still-unseeded.
    4. Scope-based fallback for fully isolated nodes (no same-phase edges
       connect them to any tagged layer):
         * scope contains ``embed`` → pre-layer bucket.
         * top-level ``head`` / ``norm`` (not inside ``layers.N.*``) → post-layer.
         * else → pre-layer bucket (capture-order fallback).
    """
    by_id = {n.id: n for n in nodes}

    def _phase_of(node):
        return node.annotations.get("phase", "") if node.annotations else ""

    preds: dict[str, list[str]] = {n.id: [] for n in nodes}
    succs: dict[str, list[str]] = {n.id: [] for n in nodes}
    for e in graph.edges:
        if e.src not in by_id or e.dst not in by_id:
            continue
        ps = _phase_of(by_id[e.src])
        pd = _phase_of(by_id[e.dst])
        # Drop cross-phase edges (fwd ↔ bwd) so the BFS stays within one phase.
        if ps and pd and ps != pd:
            continue
        preds[e.dst].append(e.src)
        succs[e.src].append(e.dst)

    eff: dict[str, float] = {}
    unknown_label: dict[str, str] = {}

    # 1) seed with explicit numeric layer
    for n in nodes:
        v = _parse_layer(n.layer)
        if v is not None:
            eff[n.id] = float(v)
        elif n.layer:
            unknown_label[n.id] = n.layer  # non-numeric label, tail bucket

    max_numeric = max(eff.values(), default=0.0)

    # 2) backward BFS: pull from max predecessor (+0.5)
    changed = True
    while changed:
        changed = False
        for n in nodes:
            if n.id in eff:
                continue
            prev_vals = [eff[p] for p in preds[n.id] if p in eff]
            if prev_vals:
                eff[n.id] = max(prev_vals) + 0.5
                changed = True

    # 3) forward BFS: pull from min successor (-0.5) for nodes still unseeded
    changed = True
    while changed:
        changed = False
        for n in nodes:
            if n.id in eff or n.id in unknown_label:
                continue
            succ_vals = [eff[s] for s in succs[n.id] if s in eff]
            if succ_vals:
                eff[n.id] = min(succ_vals) - 0.5
                changed = True

    # 4) isolated nodes: classify by scope.
    post_layer_default = max_numeric + 1.0
    for n in nodes:
        if n.id in eff or n.id in unknown_label:
            continue
        scope = (n.scope or "").lower()
        is_per_layer = "layers." in scope
        if "embed" in scope and not is_per_layer:
            eff[n.id] = float(_PRE_LAYER_KEY)
        elif not is_per_layer and any(kw in scope for kw in (".head", "head_module", ".norm")):
            eff[n.id] = post_layer_default
        elif not is_per_layer and scope.endswith("norm"):
            eff[n.id] = post_layer_default
        else:
            eff[n.id] = float(_PRE_LAYER_KEY)
    return eff


def layer_stable_sort(nodes: list, graph: OpGraph | None = None) -> list:
    """Sort nodes for readable Excel output.

    Primary key is the *effective* layer (numeric ``layer`` for tagged ops,
    propagated from neighbors otherwise — see :func:`_effective_layer_map`).
    Secondary key is the original topo index so order within a layer is
    preserved.  When ``graph`` is omitted, falls back to a pure topo sort
    using only the ``layer`` field (legacy behaviour).
    """
    topo_index = {n.id: i for i, n in enumerate(nodes)}

    if graph is None:
        # Legacy path: numeric layer only, non-numeric → tail bucket.
        def _legacy_key(node: OpNode) -> tuple[float, int]:
            v = _parse_layer(node.layer)
            if v is not None:
                return (float(v), topo_index[node.id])
            return (float(_UNKNOWN_LAYER_BASE), topo_index[node.id])
        return sorted(nodes, key=_legacy_key)

    eff = _effective_layer_map(graph, nodes)
    tail_bucket = float(_UNKNOWN_LAYER_BASE)

    def _key(node: OpNode) -> tuple[float, int]:
        if node.id in eff:
            return (eff[node.id], topo_index[node.id])
        # non-numeric layer label → tail
        return (tail_bucket, topo_index[node.id])

    return sorted(nodes, key=_key)


def infer_pipeline_stage(node: OpNode, layer_to_stage: Optional[Dict[str, int]] = None) -> str:
    """Infer pipeline stage from layer index.

    In pipeline parallelism, consecutive layers are grouped into stages.
    For now, if layer info is available, use layer number; otherwise infer from scope.
    """
    if layer_to_stage and node.layer in layer_to_stage:
        return f"stage_{layer_to_stage[node.layer]}"

    # Try to extract layer number from scope (e.g., "layers.0.mlp" → layer "0")
    if node.layer:
        try:
            stage_num = int(node.layer) // 4  # assume 4 layers per stage as default
            return f"stage_{stage_num}"
        except (ValueError, TypeError):
            pass

    return "stage_0"


def format_stream_info(node: OpNode) -> str:
    """Format stream assignment info as readable string."""
    stream_id = node.annotations.get("stream_id")
    stream_type = node.annotations.get("stream_type")

    if stream_id is None:
        return ""

    if stream_type == "comm":
        return f"comm_stream_{stream_id}"
    else:
        return f"compute_stream_{stream_id}"


def get_parallelism_info(node: OpNode, parallel_config: ParallelConfig) -> Dict[str, str]:
    """Extract parallelism information from node annotations and attributes."""
    result = {
        "strategy": parallel_config.describe(),
        "collective": "",
        "group_size": "",
        "role": "",
    }

    # Check if this is a communication node
    if node.is_comm:
        result["collective"] = node.attrs.get("collective", "")
        result["group_size"] = str(node.attrs.get("group_size", ""))
        result["role"] = node.attrs.get("role", "")

        # Infer which parallel dimension this comm belongs to
        collective = result["collective"]
        group_size = node.attrs.get("group_size", 1)

        if collective == "all_reduce":
            result["parallel_type"] = "TP"  # Tensor Parallel
        elif collective == "all_to_all":
            result["parallel_type"] = "EP"  # Expert Parallel
        else:
            result["parallel_type"] = ""
    else:
        # Check if node has parallel annotations from Split passes
        tp_split = node.annotations.get("tp_split", {})
        ep_annot = node.annotations.get("ep_needs_a2a")

        parallel_types = []
        if tp_split:
            parallel_types.append("TP")
        if ep_annot:
            parallel_types.append("EP")

        result["parallel_type"] = "/".join(parallel_types) or ""

    return result


class TransformedGraphExcelWriter:
    """Write transformed OpGraph to Excel with parallelism, communication, and stream info."""

    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        self._header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        self._header_font = Font(bold=True, color="FFFFFF", size=11)
        self._comm_fill = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
        self._compute_fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
        self._memory_fill = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
        self._thin_border = Border(bottom=Side(style="thin", color="BDBDBD"))

    def write(self, graph: OpGraph, ctx: TransformContext, output_path: Path) -> None:
        """Write transformed graph to Excel with all annotations."""
        wb = openpyxl.Workbook()

        self._write_metadata_sheet(wb, graph, ctx)
        self._write_transformed_ops_sheet(wb, graph, ctx)
        self._write_communication_sheet(wb, graph, ctx)
        self._write_parallelism_summary_sheet(wb, graph, ctx)
        self._write_stream_assignment_sheet(wb, graph, ctx)

        wb.save(output_path)
        logger.info(f"Exported transformed graph to {output_path}")

    def _write_metadata_sheet(self, wb: openpyxl.Workbook,
                              graph: OpGraph, ctx: TransformContext) -> None:
        """Write graph metadata and configuration."""
        ws = wb.active
        ws.title = "Metadata"

        ws.append(["Graph Metadata"])
        ws["A1"].font = Font(bold=True, size=12)

        metadata = [
            ("Graph Name", graph.name),
            ("Phase", graph.phase),
            ("Total Nodes", len(graph.nodes)),
            ("Total Edges", len(graph.edges)),
            ("", ""),
            ("Parallelism Config", ""),
            ("  TP (Tensor Parallel)", ctx.parallel.tp),
            ("  EP (Expert Parallel)", ctx.parallel.ep),
            ("  PP (Pipeline Parallel)", ctx.parallel.pp),
            ("  DP (Data Parallel)", ctx.parallel.dp),
            ("  CP (Context Parallel)", ctx.parallel.cp),
            ("  Sequence Parallel", "Yes" if ctx.parallel.sp else "No"),
            ("  Strategy Description", ctx.parallel.describe()),
            ("", ""),
            ("Stream Config", ""),
            ("  Compute Streams", ctx.stream_config.num_compute_streams),
            ("  Comm Streams", ctx.stream_config.num_comm_streams),
        ]

        for key, value in metadata:
            ws.append([key, value])

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = self._thin_border

    def _write_transformed_ops_sheet(self, wb: openpyxl.Workbook,
                                     graph: OpGraph, ctx: TransformContext,
                                     sheet_name: str = "Transformed Operators") -> None:
        """Write all transformed operators with detailed annotations."""
        from python.zrt.simulator.backends.roofline import get_op_formulas
        ws = wb.create_sheet(sheet_name)

        columns = [
            ("Node ID", 12),
            ("Op Name", 18),
            ("Op Type", 25),
            ("Category", 12),
            ("Scope", 45),
            ("Layer", 7),
            ("Component", 18),
            ("Parallelism Strategy", 18),
            ("Collective Op", 15),
            ("Group Size", 10),
            ("Role", 10),
            ("Pipeline Stage", 14),
            ("Stream Type", 15),
            ("Stream ID", 10),
            ("Input Shapes", 50),
            ("Output Shapes", 50),
            ("Input Dtypes", 20),
            ("Output Dtypes", 20),
            # ── compute ──────────────────────────────────────────────────────
            ("FLOPs", 14),
            ("FLOPs Formula (sym)", 24),
            ("FLOPs Formula (num)", 32),
            # ── memory access ─────────────────────────────────────────────────
            ("Read Bytes (B)", 14),
            ("Read Formula (sym)", 24),
            ("Read Formula (num)", 32),
            ("Write Bytes (B)", 14),
            ("Write Formula (sym)", 24),
            ("Write Formula (num)", 32),
            # ── comm volume ───────────────────────────────────────────────────
            ("Comm Volume (B)", 14),
            # ── timing & bound ────────────────────────────────────────────────
            ("Compute (µs)", 12),
            ("Memory (µs)", 12),
            ("Total Latency (µs)", 14),
            ("Bound", 10),
            ("Arith Intensity", 14),
            ("Annotations", 60),
        ]

        self._write_header(ws, columns)

        layer_to_stage = {}
        if ctx.parallel.pp > 1:
            layers = set(n.layer for n in graph.nodes.values() if n.layer)
            sorted_layers = sorted(layers, key=lambda x: int(x) if x.isdigit() else 0)
            for i, layer in enumerate(sorted_layers):
                layer_to_stage[layer] = i % ctx.parallel.pp

        nodes_to_write = list(layer_stable_sort(graph.topo_sort(), graph=graph))
        if graph.phase == "train" or graph.metadata.get("fwd_bwd_stitched"):
            nodes_to_write = [n for n in nodes_to_write if n.annotations.get("phase") != "bwd"]

        for row_idx, node in enumerate(nodes_to_write, 2):
            parallelism = get_parallelism_info(node, ctx.parallel)
            formulas = get_op_formulas(node)

            input_shapes = ", ".join(str(t.shape) for t in node.inputs)
            output_shapes = ", ".join(str(t.shape) for t in node.outputs)
            input_dtypes = ", ".join(str(t.dtype) for t in node.inputs)
            output_dtypes = ", ".join(str(t.dtype) for t in node.outputs)

            # Comm volume for comm nodes
            comm_vol = sum(t.mem_bytes for t in node.outputs) if node.is_comm else ""

            # Build annotations string (exclude columns that have dedicated cells)
            annotations_list = []
            for key, val in node.annotations.items():
                if key not in ("stream_id", "stream_type", "flops", "compute_us",
                               "memory_us", "latency_us", "arithmetic_intensity", "bound",
                               "read_bytes", "write_bytes"):
                    if isinstance(val, dict):
                        annotations_list.append(f"{key}={str(val)[:30]}")
                    else:
                        annotations_list.append(f"{key}={val}")
            annotations_str = "; ".join(annotations_list) if annotations_list else ""

            values = [
                node.id,
                node.name or (node.scope.rsplit(".", 1)[-1] if node.scope else ""),
                node.op_type,
                node.category,
                node.scope,
                node.layer or "",
                node.component,
                parallelism["strategy"],
                parallelism.get("collective", ""),
                parallelism.get("group_size", ""),
                parallelism.get("role", ""),
                infer_pipeline_stage(node, layer_to_stage),
                node.annotations.get("stream_type", ""),
                node.annotations.get("stream_id", ""),
                input_shapes,
                output_shapes,
                input_dtypes,
                output_dtypes,
                # compute
                node.annotations.get("flops", ""),
                formulas["flops_sym"],
                formulas["flops_num"],
                # memory access
                node.annotations.get("read_bytes", ""),
                formulas["read_sym"],
                formulas["read_num"],
                node.annotations.get("write_bytes", ""),
                formulas["write_sym"],
                formulas["write_num"],
                # comm
                comm_vol,
                # timing & bound
                round(node.annotations.get("compute_us", 0), 3) if node.annotations.get("compute_us") else "",
                round(node.annotations.get("memory_us", 0), 3) if node.annotations.get("memory_us") else "",
                round(node.annotations.get("latency_us", 0), 3) if node.annotations.get("latency_us") else "",
                node.annotations.get("bound", ""),
                round(node.annotations.get("arithmetic_intensity", 0), 2) if node.annotations.get("arithmetic_intensity") else "",
                annotations_str,
            ]

            # Choose fill color based on category
            if node.is_comm:
                fill = self._comm_fill
            elif node.category == "memory":
                fill = self._memory_fill
            else:
                fill = self._compute_fill

            self._write_row(ws, row_idx, values, fill)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(graph.nodes) + 1}"
        ws.freeze_panes = "A2"

    def _write_communication_sheet(self, wb: openpyxl.Workbook,
                                   graph: OpGraph, ctx: TransformContext) -> None:
        """Write communication operators with collective details."""
        ws = wb.create_sheet("Communication Ops")

        comm_nodes = [n for n in graph.nodes.values() if n.is_comm]
        if not comm_nodes:
            ws.append(["No communication operators found"])
            return

        columns = [
            ("Node ID", 12),
            ("Collective Op", 15),
            ("Role", 10),
            ("Group Size", 10),
            ("Scope", 45),
            ("Layer", 7),
            ("Stream Type", 15),
            ("Stream ID", 10),
            ("Input Shapes", 50),
            ("Output Shapes", 50),
            ("Inserted By", 15),
            ("Data Volume (bytes)", 18),
        ]

        self._write_header(ws, columns)

        for row_idx, node in enumerate(comm_nodes, 2):
            collective = node.attrs.get("collective", "")
            group_size = node.attrs.get("group_size", "")
            role = node.attrs.get("role", "")

            # Estimate data volume: sum of output tensor sizes
            data_volume = sum(t.mem_bytes for t in node.outputs)

            input_shapes = ", ".join(str(t.shape) for t in node.inputs)
            output_shapes = ", ".join(str(t.shape) for t in node.outputs)

            values = [
                node.id,
                collective,
                role,
                group_size,
                node.scope,
                node.layer or "",
                node.annotations.get("stream_type", ""),
                node.annotations.get("stream_id", ""),
                input_shapes,
                output_shapes,
                node.annotations.get("inserted_by", ""),
                data_volume,
            ]

            self._write_row(ws, row_idx, values, self._comm_fill)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(comm_nodes) + 1}"
        ws.freeze_panes = "A2"
        logger.info(f"Found {len(comm_nodes)} communication operators")

    def _write_parallelism_summary_sheet(self, wb: openpyxl.Workbook,
                                        graph: OpGraph, ctx: TransformContext) -> None:
        """Write summary of parallelism distribution across layers."""
        ws = wb.create_sheet("Parallelism Summary")

        # Group by layer
        layer_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
            "compute_count": 0, "comm_count": 0, "memory_count": 0,
            "comm_ops": set(), "inserted_by": []
        })

        for node in graph.nodes.values():
            layer = node.layer or "non-layer"

            if node.is_comm:
                layer_stats[layer]["comm_count"] += 1
                collective = node.attrs.get("collective", "")
                layer_stats[layer]["comm_ops"].add(collective)
                inserted_by = node.annotations.get("inserted_by", "")
                if inserted_by:
                    layer_stats[layer]["inserted_by"].append(inserted_by)
            elif node.category == "memory":
                layer_stats[layer]["memory_count"] += 1
            else:
                layer_stats[layer]["compute_count"] += 1

        ws.append(["Layer", "Compute Ops", "Memory Ops", "Comm Ops", "Comm Types", "Inserted By", "Parallelism Strategy"])
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            ws[f"{col}1"].font = self._header_font
            ws[f"{col}1"].fill = self._header_fill

        for layer in sorted(layer_stats.keys(), key=lambda x: (x == "non-layer", x)):
            stats = layer_stats[layer]
            comm_types = ", ".join(sorted(stats["comm_ops"])) if stats["comm_ops"] else ""
            inserted_by = ", ".join(sorted(set(stats["inserted_by"]))) if stats["inserted_by"] else ""

            # Infer parallelism strategy for this layer
            parallel_strat = []
            if stats["comm_count"] > 0:
                if "all_reduce" in stats["comm_ops"]:
                    parallel_strat.append("TP")
                if "all_to_all" in stats["comm_ops"]:
                    parallel_strat.append("EP")
            parallel_strat_str = "/".join(parallel_strat) if parallel_strat else "no-parallel"

            ws.append([
                layer,
                stats["compute_count"],
                stats["memory_count"],
                stats["comm_count"],
                comm_types,
                inserted_by,
                parallel_strat_str,
            ])

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20

    def _write_stream_assignment_sheet(self, wb: openpyxl.Workbook,
                                      graph: OpGraph, ctx: TransformContext) -> None:
        """Write stream assignment details."""
        ws = wb.create_sheet("Stream Assignment")

        # Collect stream statistics
        stream_assignment: Dict[int, List[str]] = defaultdict(list)
        stream_types: Dict[int, str] = {}

        for node in graph.nodes.values():
            stream_id = node.annotations.get("stream_id")
            stream_type = node.annotations.get("stream_type")
            if stream_id is not None:
                stream_assignment[stream_id].append(node.id)
                stream_types[stream_id] = stream_type or "unknown"

        ws.append(["Stream ID", "Stream Type", "Assigned Nodes", "Node Count"])
        for col in ("A", "B", "C", "D"):
            ws[f"{col}1"].font = self._header_font
            ws[f"{col}1"].fill = self._header_fill

        for stream_id in sorted(stream_assignment.keys()):
            node_ids = stream_assignment[stream_id]
            stream_type = stream_types[stream_id]

            ws.append([
                stream_id,
                stream_type,
                "; ".join(node_ids),
                len(node_ids),
            ])

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 12

    def _write_header(self, ws, columns: List[tuple[str, int]]) -> None:
        """Write header row with styling."""
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_row(self, ws, row_idx: int, values: List[Any],
                   fill: Optional[PatternFill] = None,
                   center_cols: Optional[set] = None) -> None:
        """Write data row with optional fill and center-aligned columns."""
        center_cols = center_cols or set()
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = self._thin_border
            if fill:
                cell.fill = fill
            if col_idx in center_cols:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx > 10:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Step 1: Raw capture sheets (ported from graph/excel_writer.py) ────────

    def _write_model_config_sheet(self, wb: openpyxl.Workbook,
                                  config_summary: Dict[str, Any]) -> None:
        """Write model configuration (Step 1)."""
        ws = wb.create_sheet("Model Config")
        ws.append(["Parameter", "Value"])
        ws["A1"].font = Font(bold=True, size=12)
        ws["B1"].font = Font(bold=True, size=12)
        for key, val in config_summary.items():
            ws.append([key, str(val)])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = self._thin_border

    def _write_raw_operators_sheet(self, wb: openpyxl.Workbook,
                                   records: List[Dict[str, Any]],
                                   label: str = "") -> None:
        """Write raw aten operator sequence (Step 1)."""
        from python.zrt.graph.classifier import get_fill
        sheet_name = f"Raw Operators ({label})" if label else "Raw Operators"
        ws = wb.create_sheet(sheet_name)
        columns = [
            ("Node ID", 8), ("Op Short", 12), ("Aten Op", 35),
            ("Input Shapes", 50), ("Input Dtypes", 30),
            ("Output Shapes", 50), ("Output Dtypes", 30),
            ("Module Path", 55), ("Layer", 7), ("Component", 25),
            ("Source File", 28), ("Line", 6), ("Code", 60), ("Func", 22),
            ("Extra Args", 55),
        ]
        self._write_header(ws, columns)
        for row_idx, rec in enumerate(records, 2):
            values = [
                rec["node_id"], rec.get("op_short", ""), rec["aten_op"],
                rec["input_shapes"], rec["input_dtypes"],
                rec["output_shapes"], rec["output_dtypes"],
                rec["module_path"], rec["layer"], rec["component"],
                rec.get("src_file", ""), rec.get("src_line", ""),
                rec.get("src_code", ""), rec.get("src_func", ""),
                rec.get("extra_args", ""),
            ]
            self._write_row(ws, row_idx, values, get_fill(rec["component"]),
                            center_cols={1, 12})
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"
        ws.freeze_panes = "A2"

    def _write_fused_operators_sheet(self, wb: openpyxl.Workbook,
                                     fused: List[Dict[str, Any]],
                                     label: str = "") -> None:
        """Write fused operators (Step 1)."""
        from python.zrt.graph.classifier import get_fill
        sheet_name = f"Fused Operators ({label})" if label else "Fused Operators"
        ws = wb.create_sheet(sheet_name)
        columns = [
            ("Node ID", 8), ("Op Name", 18), ("Rule Name", 22),
            ("Fused Operator", 38),
            ("Constituent Aten Ops", 70),
            ("Sub-ops", 9), ("Raw Op IDs", 28), ("Layer", 7),
            ("Fused Input Shapes", 55), ("Fused Input Dtypes", 30), ("Input Sources", 60),
            ("Fused Output Shapes", 55), ("Fused Output Dtypes", 30), ("Output Sources", 60),
        ]
        self._write_header(ws, columns)
        for row_idx, rec in enumerate(fused, 2):
            values = [
                rec["node_id"],
                rec.get("op_name", ""),
                rec.get("rule_name", ""),
                rec["fused_op"],
                rec["aten_ops"],
                rec["num_sub_ops"],
                rec.get("raw_op_ids", ""),
                rec["layer"],
                rec.get("fused_input_shapes", rec["input_shapes"]),
                rec.get("fused_input_dtypes", rec["input_dtypes"]),
                rec.get("fused_input_sources", ""),
                rec.get("fused_output_shapes", rec["output_shapes"]),
                rec.get("fused_output_dtypes", rec["output_dtypes"]),
                rec.get("fused_output_sources", ""),
            ]
            self._write_row(ws, row_idx, values, get_fill(rec["fused_op"]),
                            center_cols={1, 6})
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(fused) + 1}"
        ws.freeze_panes = "A2"

    def _write_fusion_rules_sheet(self, wb: openpyxl.Workbook,
                                  fused: List[Dict[str, Any]],
                                  output_dir: Path,
                                  label: str = "") -> None:
        """Write fusion rules and export JSON (Step 1)."""
        ws = wb.create_sheet("Fusion Rules")
        columns = [
            ("Module Class", 30), ("Fusion Level", 12), ("Aten Op Sequence", 80),
            ("Sub-ops", 9), ("Occurrences", 12), ("Example Module Path", 55),
            ("Fused Input Shapes", 55), ("Fused Input Dtypes", 30), ("Input Sources", 65),
            ("Fused Output Shapes", 55), ("Fused Output Dtypes", 30), ("Output Sources", 65),
        ]
        self._write_header(ws, columns)

        # Extract FusionSpec-like entries from fused records
        arrow = " \u2192 "
        specs_by_key: Dict[tuple, Dict[str, Any]] = {}
        for g in fused:
            if g.get("num_sub_ops", 0) <= 1:
                continue
            key = (g.get("module_class", ""), g.get("fusion_level", ""))
            if key in specs_by_key:
                specs_by_key[key]["occurrences"] += 1
            else:
                specs_by_key[key] = {
                    "module_class": g.get("module_class", ""),
                    "aten_op_sequence": g.get("aten_ops", "").split(arrow),
                    "num_sub_ops": g.get("num_sub_ops", 0),
                    "fusion_level": g.get("fusion_level", ""),
                    "example_module_path": g.get("module_path", ""),
                    "occurrences": 1,
                    "fused_input_shapes": g.get("fused_input_shapes", ""),
                    "fused_input_dtypes": g.get("fused_input_dtypes", ""),
                    "fused_input_sources": g.get("fused_input_sources", ""),
                    "fused_output_shapes": g.get("fused_output_shapes", ""),
                    "fused_output_dtypes": g.get("fused_output_dtypes", ""),
                    "fused_output_sources": g.get("fused_output_sources", ""),
                }
        specs = sorted(specs_by_key.values(), key=lambda s: -s["occurrences"])

        for row_idx, spec in enumerate(specs, 2):
            values = [
                spec["module_class"], spec["fusion_level"],
                arrow.join(spec["aten_op_sequence"]),
                spec["num_sub_ops"], spec["occurrences"], spec["example_module_path"],
                spec["fused_input_shapes"], spec["fused_input_dtypes"], spec["fused_input_sources"],
                spec["fused_output_shapes"], spec["fused_output_dtypes"], spec["fused_output_sources"],
            ]
            self._write_row(ws, row_idx, values, None)
        ws.freeze_panes = "A2"

        # Export fusion rules JSON
        if output_dir:
            json_path = output_dir / "fusion_rules.json"
            json_path.write_text(json.dumps(specs, indent=2, default=str))
            logger.info("Exported fusion rules to %s", json_path)

    def _write_by_layer_sheet(self, wb: openpyxl.Workbook,
                              raw_records: List[Dict[str, Any]],
                              fused_records: List[Dict[str, Any]],
                              label: str = "") -> None:
        """Write by-layer statistics (Step 1)."""
        sheet_name = f"By Layer ({label})" if label else "By Layer"
        ws = wb.create_sheet(sheet_name)
        ws.append(["Layer", "Fused Op Count", "Raw Op Count", "Fused Operators"])
        for col in ("A", "B", "C", "D"):
            ws[f"{col}1"].font = self._header_font
            ws[f"{col}1"].fill = self._header_fill
        layer_raw: Dict[str, int] = defaultdict(int)
        for r in raw_records:
            layer_raw[r["layer"] or "non-layer"] += 1
        layer_fused_info: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"count": 0, "ops": []})
        for r in fused_records:
            key = r["layer"] or "non-layer"
            layer_fused_info[key]["count"] += 1
            layer_fused_info[key]["ops"].append(r["fused_op"])
        for layer in sorted(layer_fused_info.keys(), key=lambda x: (x == "non-layer", x)):
            info = layer_fused_info[layer]
            seen: set = set()
            unique_ops = [op for op in info["ops"] if not (op in seen or seen.add(op))]
            ws.append([layer, info["count"], layer_raw.get(layer, 0), ", ".join(unique_ops)])
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 100

    def _write_summary_sheet(self, wb: openpyxl.Workbook,
                             fused: List[Dict[str, Any]]) -> None:
        """Write fused operator summary statistics (Step 1)."""
        ws = wb.create_sheet("Summary")
        ws.append(["Fused Operator", "Count", "Avg Sub-ops"])
        for col in ("A", "B", "C"):
            ws[f"{col}1"].font = Font(bold=True, size=12)
        fused_counts: Dict[str, List[int]] = defaultdict(list)
        for r in fused:
            fused_counts[r["fused_op"]].append(r["num_sub_ops"])
        for comp in sorted(fused_counts.keys()):
            sub_ops = fused_counts[comp]
            ws.append([comp, len(sub_ops), round(sum(sub_ops) / len(sub_ops), 1)])
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12

    # ── Step 3: Performance report sheets ─────────────────────────────────────

    def _write_perf_summary_sheet(self, wb: openpyxl.Workbook,
                                  summary) -> None:
        """Write performance summary (Step 3). Supports E2ESummary and TrainingSummary."""
        from python.zrt.report.summary import E2ESummary, TrainingSummary
        is_training = isinstance(summary, TrainingSummary)
        ws = wb.create_sheet("Training Summary" if is_training else "Performance Summary")

        section_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        section_font = Font(bold=True, size=10, color="1B5E20")
        header_font = Font(bold=True, size=11)

        def _section(label: str) -> int:
            row = ws.max_row + 1
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = section_font
            cell.fill = section_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            return row

        def _row(key: str, value: Any) -> int:
            ws.append([key, value])
            return ws.max_row

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 28
        ws.cell(row=1, column=1, value="Metric").font = header_font
        ws.cell(row=1, column=2, value="Value").font = header_font

        if is_training:
            _section("── Metadata ──")
            _row("Model", summary.model)
            _row("Hardware", summary.hardware)
            _row("Parallel", summary.parallel_desc)
            _row("Batch size", summary.batch_size)
            _row("Seq len", summary.seq_len)
            _section("── Step Timing ──")
            fwd_pct = summary.forward_ms / summary.step_ms * 100 if summary.step_ms > 0 else 0
            bwd_pct = summary.backward_ms / summary.step_ms * 100 if summary.step_ms > 0 else 0
            _row("Step latency (ms)", round(summary.step_ms, 3))
            _row("Forward (ms)", f"{round(summary.forward_ms, 3)}  ({fwd_pct:.1f}%)")
            _row("Backward (ms)", f"{round(summary.backward_ms, 3)}  ({bwd_pct:.1f}%)")
            _section("── Throughput ──")
            _row("Samples / sec", round(summary.samples_per_sec, 2))
            _row("Tokens / sec", round(summary.tokens_per_sec, 1))
            _section("── HW Efficiency ──")
            _row("MFU", f"{summary.mfu:.2%}")
            _row("HBM BW util", f"{summary.hbm_bw_util:.2%}")
            _row("Arith intensity (ops/byte)", round(summary.arithmetic_intensity, 2))
            _row("Total FLOPs (T)", round(summary.total_flops / 1e12, 3))
            _row("  Forward FLOPs (T)", round(summary.fwd_flops / 1e12, 3))
            _row("  Backward FLOPs (T)", round(summary.bwd_flops / 1e12, 3))
            _section("── Memory Access ──")
            _row("Fwd Read (GB)", round(summary.fwd_read_bytes / 1e9, 3))
            _row("Fwd Write (GB)", round(summary.fwd_write_bytes / 1e9, 3))
            _row("Bwd Read (GB)", round(summary.bwd_read_bytes / 1e9, 3))
            _row("Bwd Write (GB)", round(summary.bwd_write_bytes / 1e9, 3))
            _section("── Compute / Comm ──")
            _row("Fwd compute (ms)", round(summary.fwd_compute_ms, 3))
            _row("Fwd comm (ms)", round(summary.fwd_comm_ms, 3))
            _row("Fwd exposed comm (ms)", round(summary.fwd_exposed_comm_ms, 3))
            _row("Fwd overlap ratio", f"{summary.fwd_overlap_ratio:.1%}")
            _row("Bwd compute (ms)", round(summary.bwd_compute_ms, 3))
            _row("Bwd comm (ms)", round(summary.bwd_comm_ms, 3))
            _row("Bwd exposed comm (ms)", round(summary.bwd_exposed_comm_ms, 3))
            _row("Bwd overlap ratio", f"{summary.bwd_overlap_ratio:.1%}")
            if summary.recompute_op_count > 0:
                _section("── Activation Checkpointing ──")
                _row("Recompute op count", summary.recompute_op_count)
                _row("Recompute overhead", f"{summary.recompute_ratio:.1%}")
            if summary.memory_breakdown is not None:
                mb = summary.memory_breakdown
                _section("── Memory (per GPU) ──")
                _row("Weights (GB)", round(mb.weights / 1e9, 3))
                _row("Gradients (GB)", round(mb.grads / 1e9, 3))
                _row("Opt states (GB)", round(mb.opt_state / 1e9, 3))
                _row("Activations (GB)", round(mb.activations / 1e9, 3))
                _row("Comm buffers (GB)", round(mb.comm_buffers / 1e9, 3))
                _row("Total (GB)", round(mb.total / 1e9, 3))
        else:
            _section("── Metadata ──")
            _row("Model", summary.model)
            _row("Hardware", summary.hardware)
            _row("Phase", summary.phase.upper())
            _row("Parallel", summary.parallel_desc)
            _row("Batch size", summary.batch_size)
            _row("Seq len", summary.seq_len)
            _section("── Latency ──")
            _row("Total latency (ms)", round(summary.latency_ms, 3))
            if summary.ttft_ms is not None:
                _row("TTFT (ms)", round(summary.ttft_ms, 3))
            if summary.tpot_ms is not None:
                _row("TPOT (ms/token)", round(summary.tpot_ms, 3))
            _row("Throughput (tok/s)", round(summary.tokens_per_sec, 1))
            _section("── Compute / Comm ──")
            _row("Compute (ms)", round(summary.compute_ms, 3))
            _row("Comm (ms)", round(summary.comm_ms, 3))
            _row("Exposed comm (ms)", round(summary.exposed_comm_ms, 3))
            _row("Overlap ratio", f"{summary.overlap_ratio:.1%}")
            _section("── HW Efficiency ──")
            _row("MFU", f"{summary.mfu:.2%}")
            _row("HBM BW util", f"{summary.hbm_bandwidth_util:.2%}")
            _row("Arith intensity (ops/byte)", round(summary.arithmetic_intensity, 2))
            _row("Total FLOPs (T)", round(summary.total_flops / 1e12, 4))
            _row("Total bytes (GB)", round(summary.total_bytes / 1e9, 4))
            _row("  Read (GB)", round(summary.read_bytes / 1e9, 4))
            _row("  Write (GB)", round(summary.write_bytes / 1e9, 4))

        if summary.by_component:
            _section("── By Component (% of serial latency) ──")
            for comp, pct in sorted(summary.by_component.items(), key=lambda x: -x[1]):
                _row(comp, f"{pct:.1f}%")
        if summary.by_layer:
            _section("── By Layer (ms) ──")
            n = len(summary.by_layer)
            avg = sum(summary.by_layer) / n if n else 0
            _row(f"Layers ({n}, avg {avg:.3f} ms)", "")
            for i, lat in enumerate(summary.by_layer):
                _row(f"  Layer {i}", round(lat, 4))
        if summary.memory_budget is not None:
            mb = summary.memory_budget
            _section("── Memory Budget (per GPU) ──")
            _row("Weights (GB)", round(mb.weights / 1e9, 3))
            _row("KV cache (GB)", round(mb.kv_cache_mb * 1e-3, 3))
            _row("Activations (GB)", round(mb.activation_peak_mb * 1e-3, 3))
            _row("Total (GB)", round(mb.total_mb * 1e-3, 3))
            _row("Feasible", "Yes" if mb.is_feasible else "No (OOM)")

    def _write_bottlenecks_sheet(self, wb: openpyxl.Workbook, summary) -> None:
        """Write top bottleneck operators (Step 3)."""
        ws = wb.create_sheet("Top Bottlenecks")
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12
        self._write_header(ws, [
            ("#", 8), ("Operator", 50), ("Latency (us)", 16),
            ("% of Total", 14), ("Bound", 12),
        ])
        total_us = summary.latency_ms * 1000
        for idx, (op_desc, lat_us) in enumerate(summary.top_bottleneck_ops, 1):
            pct = (lat_us / total_us * 100) if total_us > 0 else 0
            bound = ""
            if "[" in op_desc:
                parts = op_desc.split("[")
                bound = parts[1].rstrip("]") if len(parts) > 1 else ""
            self._write_row(ws, idx + 1, [idx, op_desc, round(lat_us, 2),
                                           f"{pct:.1f}%", bound])
        ws.freeze_panes = "A2"

    # ── Unified full-report writers ───────────────────────────────────────────

    def write_full_inference(
        self,
        config_summary: Dict[str, Any],
        raw_records: List[Dict[str, Any]],
        fused_records: List[Dict[str, Any]],
        transformed_graph: OpGraph,
        ctx: TransformContext,
        summary,
        output_path: Path,
    ) -> None:
        """Write complete inference report: Step 1 (capture) + Step 2 (transform) + Step 3 (perf).

        Sheet order:
          1. Model Config
          2. Raw Operators
          3. Fused Operators
          4. Summary
          5. By Layer
          6. Fusion Rules
          7. Transformed Operators
          8. Communication Ops
          9. Parallelism Summary
          10. Stream Assignment
          11. Performance Summary
          12. Top Bottlenecks
        """
        wb = openpyxl.Workbook()

        # Step 1: Raw capture
        self._write_model_config_sheet(wb, config_summary)
        self._write_raw_operators_sheet(wb, raw_records)
        self._write_fused_operators_sheet(wb, fused_records)
        self._write_summary_sheet(wb, fused_records)
        self._write_by_layer_sheet(wb, raw_records, fused_records)
        self._write_fusion_rules_sheet(wb, fused_records, output_path.parent)

        # Step 2: Transformed graph
        self._write_transformed_ops_sheet(wb, transformed_graph, ctx)
        self._write_communication_sheet(wb, transformed_graph, ctx)
        self._write_parallelism_summary_sheet(wb, transformed_graph, ctx)
        self._write_stream_assignment_sheet(wb, transformed_graph, ctx)

        # Step 3: Performance
        self._write_perf_summary_sheet(wb, summary)
        self._write_bottlenecks_sheet(wb, summary)

        wb.save(output_path)
        logger.info("Exported full inference report to %s", output_path)

    def write_full_training(
        self,
        config_summary: Dict[str, Any],
        fwd_raw: List[Dict[str, Any]],
        fwd_fused: List[Dict[str, Any]],
        bwd_raw: List[Dict[str, Any]],
        bwd_fused: List[Dict[str, Any]],
        fwd_graph: OpGraph,
        bwd_graph: OpGraph,
        ctx: TransformContext,
        training_summary,
        output_path: Path,
    ) -> None:
        """Write complete training report: Step 1 (fwd+bwd) + Step 2 (fwd+bwd) + Step 3 (perf).

        Sheet order:
          1. Model Config
          2. Raw Operators (fwd)
          3. Fused Operators (fwd)
          4. Raw Operators (bwd)
          5. Fused Operators (bwd)
          6. Summary
          7. By Layer
          8. Fusion Rules
          9. Transformed Operators (fwd)
          10. Communication Ops (fwd)
          11. Parallelism Summary
          12. Stream Assignment
          13. Backward Operators
          14. Recompute Ops
          15. Training Summary
          16. Top Bottlenecks
        """
        wb = openpyxl.Workbook()

        # Step 1: Raw capture (fwd + bwd)
        self._write_model_config_sheet(wb, config_summary)
        self._write_raw_operators_sheet(wb, fwd_raw, "fwd")
        self._write_fused_operators_sheet(wb, fwd_fused, "fwd")
        self._write_raw_operators_sheet(wb, bwd_raw, "bwd")
        self._write_fused_operators_sheet(wb, bwd_fused, "bwd")
        # Merge raw/fused for summary and by-layer
        all_raw = fwd_raw + bwd_raw
        all_fused = fwd_fused + bwd_fused
        self._write_summary_sheet(wb, all_fused)
        self._write_by_layer_sheet(wb, all_raw, all_fused)
        self._write_fusion_rules_sheet(wb, all_fused, output_path.parent)

        # Step 2: Transformed graph (fwd + bwd)
        self._write_transformed_ops_sheet(wb, fwd_graph, ctx)
        self._write_communication_sheet(wb, fwd_graph, ctx)
        self._write_parallelism_summary_sheet(wb, fwd_graph, ctx)
        self._write_stream_assignment_sheet(wb, fwd_graph, ctx)
        self._write_backward_ops_sheet(wb, bwd_graph, ctx)
        self._write_recompute_sheet(wb, bwd_graph)

        # Step 3: Performance
        self._write_perf_summary_sheet(wb, training_summary)
        self._write_bottlenecks_sheet(wb, training_summary)

        wb.save(output_path)
        logger.info("Exported full training report to %s", output_path)


def export_transformed_graph(graph: OpGraph, ctx: TransformContext,
                            output_dir: Path) -> Dict[str, Path]:
    """Export transformed graph to Excel, JSON, and optionally ONNX.

    Parameters
    ----------
    graph : OpGraph
        The transformed computation graph
    ctx : TransformContext
        Transformation context with parallel config and stream config
    output_dir : Path
        Output directory for exported files

    Returns
    -------
    dict[str, Path]
        Paths to generated files: {format: path}
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Determine output filename
    base_name = graph.name.replace("/", "_").replace(":", "_")

    # Excel export
    excel_path = output_dir / f"{base_name}_transformed_ops.xlsx"
    writer = TransformedGraphExcelWriter()
    writer.write(graph, ctx, excel_path)

    # JSON export (simplified)
    json_path = output_dir / f"{base_name}_transformed_graph.json"
    _export_json(graph, ctx, json_path)

    return {
        "excel": excel_path,
        "json": json_path,
    }


def _export_json(graph: OpGraph, ctx: TransformContext, output_path: Path) -> None:
    """Export transformed graph to JSON format."""
    data = {
        "graph": {
            "name": graph.name,
            "phase": graph.phase,
            "num_nodes": len(graph.nodes),
            "num_edges": len(graph.edges),
        },
        "parallelism": {
            "strategy": ctx.parallel.describe(),
            "tp": ctx.parallel.tp,
            "ep": ctx.parallel.ep,
            "pp": ctx.parallel.pp,
            "dp": ctx.parallel.dp,
            "sp": ctx.parallel.sp,
        },
        "stream_config": {
            "compute_streams": ctx.stream_config.num_compute_streams,
            "comm_streams": ctx.stream_config.num_comm_streams,
        },
        "nodes": [],
        "edges": [],
    }

    # Convert nodes
    for node in graph.topo_sort():
        node_data = {
            "id": node.id,
            "op_type": node.op_type,
            "name": node.name,
            "module_class": node.module_class,
            "category": node.category,
            "scope": node.scope,
            "layer": node.layer,
            "fused_from": list(node.fused_from),
            "num_sub_ops": node.num_sub_ops,
            "fusion_level": node.fusion_level,
            "attrs": node.attrs,
            "annotations": node.annotations,
            "input_shapes": [list(t.shape) for t in node.inputs],
            "output_shapes": [list(t.shape) for t in node.outputs],
            "input_dtypes": [t.dtype.value for t in node.inputs],
            "output_dtypes": [t.dtype.value for t in node.outputs],
        }
        data["nodes"].append(node_data)

    # Convert edges
    for edge in graph.edges:
        edge_data = {
            "src": edge.src,
            "dst": edge.dst,
            "src_idx": edge.src_idx,
            "dst_idx": edge.dst_idx,
        }
        data["edges"].append(edge_data)

    output_path.write_text(json.dumps(data, indent=2, default=str))
    logger.info(f"Exported transformed graph JSON to {output_path}")


# ── Training export ───────────────────────────────────────────────────────────

class TrainingGraphExcelWriter(TransformedGraphExcelWriter):
    """Write forward + backward OpGraphs to a single Excel workbook.

    Adds two training-specific sheets on top of the standard 5:
      - Training Summary: step-level metrics from TrainingSummary
      - Recompute Ops:    backward-graph ops flagged as activation-checkpoint recompute
    """

    def write_training(
        self,
        fwd_graph: OpGraph,
        bwd_graph: OpGraph | None,
        ctx: TransformContext,
        output_path: Path,
        training_summary=None,
        fwd_records: List[Dict[str, Any]] | None = None,
        bwd_records: List[Dict[str, Any]] | None = None,
    ) -> None:
        """Write training workbook (fwd + bwd graphs + summary sheet)."""
        wb = openpyxl.Workbook()

        self._write_metadata_sheet(wb, fwd_graph, ctx)

        # Pre-fusion vs post-fusion comparison sheets (when raw records provided).
        if fwd_records:
            self._write_raw_operators_sheet(wb, fwd_records, "fwd")
            self._write_fused_operators_sheet(
                wb, _graph_to_fused_records(fwd_graph, phase_filter="fwd"), "fwd")
        if bwd_records:
            self._write_raw_operators_sheet(wb, bwd_records, "bwd")
            if bwd_graph is not None:
                self._write_fused_operators_sheet(
                    wb, _graph_to_fused_records(bwd_graph, phase_filter="bwd"), "bwd")

        self._write_transformed_ops_sheet(
            wb, fwd_graph, ctx, sheet_name="Forward Operators")
        self._write_communication_sheet(wb, fwd_graph, ctx)
        self._write_parallelism_summary_sheet(wb, fwd_graph, ctx)
        self._write_stream_assignment_sheet(wb, fwd_graph, ctx)

        if bwd_graph is not None:
            self._write_backward_ops_sheet(wb, bwd_graph, ctx)
            self._write_recompute_sheet(wb, bwd_graph)
        if training_summary is not None:
            self._write_training_summary_sheet(wb, training_summary)

        wb.save(output_path)
        logger.info(f"Exported training graphs to {output_path}")

    def _write_backward_ops_sheet(
        self, wb: openpyxl.Workbook, graph: OpGraph, ctx: TransformContext
    ) -> None:
        """Backward graph operators (mirrors Transformed Operators sheet)."""
        from python.zrt.simulator.backends.roofline import get_op_formulas
        ws = wb.create_sheet("Backward Operators")

        columns = [
            ("Node ID", 12),
            ("Op Name", 18),
            ("Op Type", 25),
            ("Category", 12),
            ("Scope", 45),
            ("Layer", 7),
            ("Component", 18),
            ("Recompute", 10),
            ("Input Shapes", 50),
            ("Output Shapes", 50),
            # ── compute ──────────────────────────────────────────────────────
            ("FLOPs", 14),
            ("FLOPs Formula (sym)", 24),
            ("FLOPs Formula (num)", 32),
            # ── memory access ─────────────────────────────────────────────────
            ("Read Bytes (B)", 14),
            ("Read Formula (sym)", 24),
            ("Read Formula (num)", 32),
            ("Write Bytes (B)", 14),
            ("Write Formula (sym)", 24),
            ("Write Formula (num)", 32),
            # ── comm & timing ─────────────────────────────────────────────────
            ("Comm Volume (B)", 14),
            ("Compute (µs)", 12),
            ("Memory (µs)", 12),
            ("Total Latency (µs)", 14),
            ("Bound", 10),
            ("Arith Intensity", 14),
        ]
        self._write_header(ws, columns)

        _recompute_fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")

        nodes_to_write = list(layer_stable_sort(graph.topo_sort(), graph=graph))
        if graph.phase == "train" or graph.metadata.get("fwd_bwd_stitched"):
            nodes_to_write = [n for n in nodes_to_write if n.annotations.get("phase") == "bwd"]

        for row_idx, node in enumerate(nodes_to_write, 2):
            is_recompute = (
                node.annotations.get("recompute", False)
                or node.attrs.get("recompute", False)
            )
            fill = _recompute_fill if is_recompute else (
                self._comm_fill if node.is_comm else self._compute_fill
            )
            formulas = get_op_formulas(node)
            comm_vol = sum(t.mem_bytes for t in node.outputs) if node.is_comm else ""
            values = [
                node.id,
                node.name or (node.scope.rsplit(".", 1)[-1] if node.scope else ""),
                node.op_type,
                node.category,
                node.scope,
                node.layer or "",
                node.component,
                "YES" if is_recompute else "",
                ", ".join(str(t.shape) for t in node.inputs),
                ", ".join(str(t.shape) for t in node.outputs),
                # compute
                node.annotations.get("flops", ""),
                formulas["flops_sym"],
                formulas["flops_num"],
                # memory access
                node.annotations.get("read_bytes", ""),
                formulas["read_sym"],
                formulas["read_num"],
                node.annotations.get("write_bytes", ""),
                formulas["write_sym"],
                formulas["write_num"],
                # comm & timing
                comm_vol,
                round(node.annotations.get("compute_us", 0), 3) if node.annotations.get("compute_us") else "",
                round(node.annotations.get("memory_us", 0), 3) if node.annotations.get("memory_us") else "",
                round(node.annotations.get("latency_us", 0), 3) if node.annotations.get("latency_us") else "",
                node.annotations.get("bound", ""),
                round(node.annotations.get("arithmetic_intensity", 0), 2) if node.annotations.get("arithmetic_intensity") else "",
            ]
            self._write_row(ws, row_idx, values, fill)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{graph.num_nodes() + 1}"
        ws.freeze_panes = "A2"

    def _write_recompute_sheet(self, wb: openpyxl.Workbook, bwd_graph: OpGraph) -> None:
        """Summarize ops flagged as activation-checkpoint recompute."""
        ws = wb.create_sheet("Recompute Ops")

        recompute_nodes = [
            n for n in bwd_graph.nodes.values()
            if n.annotations.get("recompute", False) or n.attrs.get("recompute", False)
        ]

        if not recompute_nodes:
            ws.append(["No recompute ops detected (activation checkpointing not active)"])
            return

        columns = [
            ("Node ID", 14),
            ("Op Type", 25),
            ("Scope", 45),
            ("Layer", 7),
            ("FLOPs", 12),
            ("Latency (µs)", 14),
            ("Bound", 10),
            ("Input Shapes", 50),
        ]
        self._write_header(ws, columns)

        _fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
        total_lat = 0.0
        total_flops = 0

        for row_idx, node in enumerate(recompute_nodes, 2):
            lat = node.annotations.get("latency_us", 0) or 0
            flops = node.annotations.get("flops", 0) or 0
            total_lat   += lat
            total_flops += flops
            values = [
                node.id,
                node.op_type,
                node.scope,
                node.layer or "",
                flops,
                round(lat, 3),
                node.annotations.get("bound", ""),
                ", ".join(str(t.shape) for t in node.inputs),
            ]
            self._write_row(ws, row_idx, values, _fill)

        # Summary footer
        footer_row = len(recompute_nodes) + 2
        ws.cell(row=footer_row, column=1, value="TOTAL")
        ws.cell(row=footer_row, column=5, value=total_flops)
        ws.cell(row=footer_row, column=6, value=round(total_lat, 3))
        for col in range(1, 9):
            ws.cell(row=footer_row, column=col).font = Font(bold=True)

        ws.freeze_panes = "A2"

    def _write_training_summary_sheet(self, wb: openpyxl.Workbook, ts) -> None:
        """Write TrainingSummary metrics as a key-value sheet."""
        ws = wb.create_sheet("Training Summary")

        ws.append(["Training Step Summary"])
        ws["A1"].font = Font(bold=True, size=13)

        fwd_pct = ts.forward_ms / ts.step_ms * 100 if ts.step_ms > 0 else 0.0
        bwd_pct = ts.backward_ms / ts.step_ms * 100 if ts.step_ms > 0 else 0.0

        rows: list[tuple[str, Any]] = [
            ("Model",          ts.model),
            ("Hardware",       ts.hardware),
            ("Parallelism",    ts.parallel_desc),
            ("Batch size",     ts.batch_size),
            ("Sequence length", ts.seq_len),
            ("", ""),
            ("=== Step Timing ===", ""),
            ("Step latency (ms)",   round(ts.step_ms, 3)),
            ("Forward (ms)",        f"{round(ts.forward_ms, 3)}  ({fwd_pct:.1f}%)"),
            ("Backward (ms)",       f"{round(ts.backward_ms, 3)}  ({bwd_pct:.1f}%)"),
            ("", ""),
            ("=== Throughput ===", ""),
            ("Samples / sec",   round(ts.samples_per_sec, 2)),
            ("Tokens / sec",    round(ts.tokens_per_sec, 1)),
            ("", ""),
            ("=== HW Efficiency ===", ""),
            ("MFU",                   f"{ts.mfu:.2%}"),
            ("HBM BW util",           f"{ts.hbm_bw_util:.2%}"),
            ("Arithmetic Intensity",  f"{ts.arithmetic_intensity:.2f} ops/byte"),
            ("Total FLOPs (T)",       round(ts.total_flops / 1e12, 3)),
            ("  Forward FLOPs (T)",   round(ts.fwd_flops   / 1e12, 3)),
            ("  Backward FLOPs (T)",  round(ts.bwd_flops   / 1e12, 3)),
            ("", ""),
            ("=== Memory Access (Read+Write) ===", ""),
            ("Fwd Read (GB)",    round(ts.fwd_read_bytes  / 1e9, 3)),
            ("Fwd Write (GB)",   round(ts.fwd_write_bytes / 1e9, 3)),
            ("Fwd Total (GB)",   round(ts.fwd_bytes       / 1e9, 3)),
            ("Bwd Read (GB)",    round(ts.bwd_read_bytes  / 1e9, 3)),
            ("Bwd Write (GB)",   round(ts.bwd_write_bytes / 1e9, 3)),
            ("Bwd Total (GB)",   round(ts.bwd_bytes       / 1e9, 3)),
            ("", ""),
            ("=== Compute / Comm Breakdown ===", ""),
            ("Fwd compute (ms)",      round(ts.fwd_compute_ms, 3)),
            ("Fwd comm (ms)",         round(ts.fwd_comm_ms, 3)),
            ("Fwd exposed comm (ms)", round(ts.fwd_exposed_comm_ms, 3)),
            ("Fwd overlap ratio",     f"{ts.fwd_overlap_ratio:.1%}"),
            ("Bwd compute (ms)",      round(ts.bwd_compute_ms, 3)),
            ("Bwd comm (ms)",         round(ts.bwd_comm_ms, 3)),
            ("Bwd exposed comm (ms)", round(ts.bwd_exposed_comm_ms, 3)),
            ("Bwd overlap ratio",     f"{ts.bwd_overlap_ratio:.1%}"),
            ("", ""),
            ("=== Activation Checkpointing ===", ""),
            ("Recompute op count", ts.recompute_op_count),
            ("Recompute overhead", f"{ts.recompute_ratio:.1%}"),
        ]

        if ts.memory_breakdown is not None:
            mb = ts.memory_breakdown
            rows += [
                ("", ""),
                ("=== Memory (per GPU) ===", ""),
                ("Weights (GB)",     round(mb.weights      / 1e9, 3)),
                ("Gradients (GB)",   round(mb.grads        / 1e9, 3)),
                ("Opt states (GB)",  round(mb.opt_state    / 1e9, 3)),
                ("Activations (GB)", round(mb.activations  / 1e9, 3)),
                ("Comm buffers (GB)", round(mb.comm_buffers / 1e9, 3)),
                ("Total (GB)",       round(mb.total        / 1e9, 3)),
            ]

        if ts.top_bottleneck_ops:
            rows += [("", ""), ("=== Top Bottleneck Ops ===", "")]
            for op_desc, lat_us in ts.top_bottleneck_ops:
                rows.append((op_desc, f"{lat_us:.1f} µs"))

        for key, val in rows:
            ws.append([key, val])
            if str(key).startswith("==="):
                row_num = ws.max_row
                ws.cell(row=row_num, column=1).font = Font(bold=True)

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 28


def _get_fused_shapes_from_sem_io(node: OpNode, direction: str) -> str:
    """Get fused shapes from sem_io annotation (includes CP split).
    
    Args:
        node: OpNode to extract shapes from
        direction: "input" or "output"
        
    Returns:
        Comma-separated string of shapes
    """
    sem_io = node.annotations.get("sem_io", {})
    
    if direction == "input":
        # Priority: activation, then all non-output roles
        if "activation" in sem_io:
            return str(sem_io["activation"].get("shape", ""))
        shapes = [str(v.get("shape", "")) for k, v in sem_io.items() 
                  if k not in ("output",)]
        return ", ".join(shapes) if shapes else ", ".join(str(t.shape) for t in node.inputs)
    else:
        # Output direction
        if "output" in sem_io:
            return str(sem_io["output"].get("shape", ""))
        shapes = [str(v.get("shape", "")) for k, v in sem_io.items() 
                  if k == "output"]
        return ", ".join(shapes) if shapes else ", ".join(str(t.shape) for t in node.outputs)


def _get_fused_dtypes_from_sem_io(node: OpNode, direction: str) -> str:
    """Get fused dtypes from sem_io annotation (includes CP split).
    
    Args:
        node: OpNode to extract dtypes from
        direction: "input" or "output"
        
    Returns:
        Comma-separated string of dtypes
    """
    sem_io = node.annotations.get("sem_io", {})
    sem_dtype = node.annotations.get("sem_dtype", "")
    
    if direction == "input":
        if "activation" in sem_io:
            dtype = sem_io["activation"].get("dtype", sem_dtype)
            return str(dtype) if dtype else ""
        dtypes = [str(v.get("dtype", "")) for k, v in sem_io.items() 
                  if k not in ("output",)]
        return ", ".join(dtypes) if dtypes else ", ".join(str(t.dtype) for t in node.inputs)
    else:
        if "output" in sem_io:
            dtype = sem_io["output"].get("dtype", sem_dtype)
            return str(dtype) if dtype else ""
        dtypes = [str(v.get("dtype", "")) for k, v in sem_io.items() 
                  if k == "output"]
        return ", ".join(dtypes) if dtypes else ", ".join(str(t.dtype) for t in node.outputs)


def _graph_to_fused_records(graph: OpGraph,
                            phase_filter: str | None = None) -> List[Dict[str, Any]]:
    """Build fused-record dicts from an OpGraph for the Fused Operators sheet.

    When the graph is a stitched fwd+bwd unified graph, ``phase_filter``
    keeps only nodes whose ``annotations["phase"]`` matches.
    """
    arrow = " → "
    records: List[Dict[str, Any]] = []
    nodes = layer_stable_sort(graph.topo_sort(), graph=graph)
    if phase_filter is not None and (graph.phase == "train"
                                     or graph.metadata.get("fwd_bwd_stitched")):
        nodes = [n for n in nodes if n.annotations.get("phase") == phase_filter]
    for idx, node in enumerate(nodes):
        constituents = node.fused_from or [node.op_type]
        # Recover the "raw Node IDs that fused into this node".  Strip the
        # ``op_`` / ``bwd_op_`` prefix so the value lines up with the Node ID
        # column in the Raw Operators sheet.
        src_ids_raw = node.annotations.get("source_op_ids") or [node.id]
        def _strip(idstr: str) -> str:
            s = str(idstr)
            for pre in ("bwd_op_", "op_"):
                if s.startswith(pre):
                    return s[len(pre):]
            return s
        raw_ids_str = ",".join(_strip(s) for s in src_ids_raw)
        # When a node was never fused (single raw op carried over), the scope
        # tail (e.g. "norm", "attn") is meaningless as a per-op identifier.
        # Surface the operator's own short name (e.g. "add", "rsqrt") so the
        # row is self-describing.  Fused nodes keep their semantic op_name
        # (the leaf attr / wrapper module name).
        is_unfused_single = (
            (node.num_sub_ops or 1) <= 1
            and not node.annotations.get("fused_by_rule")
            and (node.op_type.startswith("aten.") or node.op_type.startswith("comm."))
        )
        if is_unfused_single:
            op_name_value = op_short(node.op_type)
        else:
            op_name_value = (
                node.name or (node.scope.rsplit(".", 1)[-1] if node.scope else "")
            )
        records.append({
            "node_id": idx,
            "op_name": op_name_value,
            "rule_name": node.annotations.get("fused_by_rule", "") or "",
            "fused_op": node.op_type,
            "aten_ops": arrow.join(constituents),
            "num_sub_ops": node.num_sub_ops or 1,
            "raw_op_ids": raw_ids_str,
            "layer": node.layer or "",
            "module_class": node.module_class,
            "module_path": node.scope,
            "fusion_level": node.fusion_level or ("leaf" if (node.num_sub_ops or 0) <= 1 else "parent"),
            "input_shapes": ", ".join(str(t.shape) for t in node.inputs),
            "input_dtypes": ", ".join(str(t.dtype) for t in node.inputs),
            "output_shapes": ", ".join(str(t.shape) for t in node.outputs),
            "output_dtypes": ", ".join(str(t.dtype) for t in node.outputs),
            # Use sem_io shapes (includes CP split) if available
            "fused_input_shapes": _get_fused_shapes_from_sem_io(node, "input"),
            "fused_input_dtypes": _get_fused_dtypes_from_sem_io(node, "input"),
            "fused_output_shapes": _get_fused_shapes_from_sem_io(node, "output"),
            "fused_output_dtypes": _get_fused_dtypes_from_sem_io(node, "output"),
            "fused_input_sources": "",
            "fused_output_sources": "",
        })
    return records


def export_training_graphs(
    fwd_graph: OpGraph,
    bwd_graph: OpGraph | None,
    ctx: TransformContext,
    output_dir: Path,
    training_summary=None,
    fwd_records: List[Dict[str, Any]] | None = None,
    bwd_records: List[Dict[str, Any]] | None = None,
) -> Dict[str, Path]:
    """Export forward + backward training graphs to Excel and JSON.

    Parameters
    ----------
    fwd_graph
        Transformed train_forward graph, or unified graph (phase="train").
    bwd_graph
        Transformed train_backward graph, or None if fwd_graph is unified.
    ctx
        Transform context (shared between phases).
    output_dir
        Output directory.
    training_summary
        Optional ``TrainingSummary`` — written as a dedicated Excel sheet.

    Returns
    -------
    dict[str, Path]
        {"excel": ..., "json_fwd": ..., "json_bwd": ...}
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    base = fwd_graph.name.replace("/", "_").replace(":", "_")
    for suffix in ("_train_forward", "_train_backward", "_train"):
        base = base.replace(suffix, "")

    excel_path = output_dir / f"{base}_training.xlsx"
    writer = TrainingGraphExcelWriter()
    writer.write_training(
        fwd_graph, bwd_graph, ctx, excel_path, training_summary,
        fwd_records=fwd_records, bwd_records=bwd_records,
    )

    json_fwd = output_dir / f"{base}_train_forward.json"
    _export_json(fwd_graph, ctx, json_fwd)

    if bwd_graph is not None and bwd_graph != fwd_graph:
        json_bwd = output_dir / f"{base}_train_backward.json"
        _export_json(bwd_graph, ctx, json_bwd)
        return {"excel": excel_path, "json_fwd": json_fwd, "json_bwd": json_bwd}

    return {"excel": excel_path, "json_fwd": json_fwd}


# ── Unified full-report exports ───────────────────────────────────────────────

def export_full_report(
    config_summary: Dict[str, Any],
    raw_records: List[Dict[str, Any]],
    fused_records: List[Dict[str, Any]],
    transformed_graph: OpGraph,
    ctx: TransformContext,
    summary,
    output_dir: Path,
) -> Dict[str, Path]:
    """Export complete inference report: capture + transform + performance.

    Single Excel file with all sheets:
      Model Config → Raw Operators → Fused Operators → Summary → By Layer →
      Fusion Rules → Transformed Operators → Communication Ops →
      Parallelism Summary → Stream Assignment → Performance Summary →
      Top Bottlenecks

    Parameters
    ----------
    config_summary : dict
        Model configuration from ``build_config_summary()``.
    raw_records : list[dict]
        Raw aten op records from ``run_trace_phases``.
    fused_records : list[dict]
        Fused op records from ``FusionEngine.fuse_keep_children()``.
    transformed_graph : OpGraph
        Transformed graph from ``TransformPipeline.run()``.
    ctx : TransformContext
        Transform context with parallel/stream config.
    summary : E2ESummary
        Performance summary from ``build_summary()``.
    output_dir : Path
        Output directory.

    Returns
    -------
    dict[str, Path]
        {"excel": ..., "json": ...}
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    base = summary.model.replace("/", "_").replace(":", "_")
    phase = summary.phase

    excel_path = output_dir / f"{base}_{phase}_full_report.xlsx"
    writer = TransformedGraphExcelWriter()
    writer.write_full_inference(
        config_summary, raw_records, fused_records,
        transformed_graph, ctx, summary, excel_path,
    )

    json_path = output_dir / f"{base}_{phase}_full_report.json"
    _export_json(transformed_graph, ctx, json_path)

    return {"excel": excel_path, "json": json_path}


def export_full_training_report(
    config_summary: Dict[str, Any],
    fwd_raw: List[Dict[str, Any]],
    fwd_fused: List[Dict[str, Any]],
    bwd_raw: List[Dict[str, Any]],
    bwd_fused: List[Dict[str, Any]],
    fwd_graph: OpGraph,
    bwd_graph: OpGraph,
    ctx: TransformContext,
    training_summary,
    output_dir: Path,
) -> Dict[str, Path]:
    """Export complete training report: capture(fwd+bwd) + transform(fwd+bwd) + performance.

    Single Excel file with all sheets:
      Model Config → Raw Ops (fwd) → Fused Ops (fwd) → Raw Ops (bwd) →
      Fused Ops (bwd) → Summary → By Layer → Fusion Rules →
      Transformed Ops (fwd) → Communication Ops (fwd) →
      Parallelism Summary → Stream Assignment → Backward Operators →
      Recompute Ops → Training Summary → Top Bottlenecks

    Parameters
    ----------
    config_summary : dict
        Model configuration.
    fwd_raw / fwd_fused : list[dict]
        Forward raw and fused op records.
    bwd_raw / bwd_fused : list[dict]
        Backward raw and fused op records.
    fwd_graph / bwd_graph : OpGraph
        Transformed forward/backward graphs.
    ctx : TransformContext
        Transform context (with training config).
    training_summary : TrainingSummary
        Training performance summary.
    output_dir : Path
        Output directory.

    Returns
    -------
    dict[str, Path]
        {"excel": ..., "json_fwd": ..., "json_bwd": ...}
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    base = fwd_graph.name.replace("/", "_").replace(":", "_").replace("_train_forward", "")

    excel_path = output_dir / f"{base}_full_training.xlsx"
    writer = TransformedGraphExcelWriter()
    writer.write_full_training(
        config_summary, fwd_raw, fwd_fused, bwd_raw, bwd_fused,
        fwd_graph, bwd_graph, ctx, training_summary, excel_path,
    )

    json_fwd = output_dir / f"{base}_train_forward.json"
    _export_json(fwd_graph, ctx, json_fwd)
    json_bwd = output_dir / f"{base}_train_backward.json"
    _export_json(bwd_graph, ctx, json_bwd)

    return {"excel": excel_path, "json_fwd": json_fwd, "json_bwd": json_bwd}
