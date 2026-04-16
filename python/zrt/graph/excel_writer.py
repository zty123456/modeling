"""Write operator records to a formatted Excel workbook + fusion rules JSON."""
from __future__ import annotations

import json
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from python.zrt.graph.classifier import get_fill
from python.zrt.graph.fusion import FusionEngine, FusionSpec
from python.zrt.graph.tracker import ModuleTracker

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Write operator records to a formatted Excel workbook + fusion rules JSON."""

    def __init__(self, tracker: ModuleTracker):
        self._tracker = tracker
        self._fusion_engine = FusionEngine(tracker)
        self._header_fill = PatternFill(start_color="263238", end_color="263238", fill_type="solid")
        self._header_font_white = Font(bold=True, color="FFFFFF", size=11)
        self._header_font = Font(bold=True, size=12)
        self._thin_border = Border(bottom=Side(style="thin", color="BDBDBD"))

    def write(self, records: List[Dict[str, Any]], output_path: Path,
              config_summary: Dict[str, Any]):
        wb = openpyxl.Workbook()

        self._write_config_sheet(wb, config_summary)
        fused = self._fusion_engine.fuse(records)
        self._write_fused_sheet(wb, fused, records)
        self._write_raw_sheet(wb, records)
        self._write_summary_sheet(wb, fused)
        self._write_by_layer_sheet(wb, records, fused)
        self._write_fusion_rules_sheet(wb, fused, output_path)

        wb.save(output_path)

    def _write_config_sheet(self, wb, config_summary):
        ws = wb.active
        ws.title = "Model Config"
        ws.append(["Parameter", "Value"])
        ws["A1"].font = self._header_font
        ws["B1"].font = self._header_font
        for key, val in config_summary.items():
            ws.append([key, str(val)])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

    def _write_fused_sheet(self, wb, fused, records):
        ws = wb.create_sheet("Fused Operators")
        columns = [
            ("Node ID", 8), ("Fused Operator", 38), ("Constituent Aten Ops", 70),
            ("Sub-ops", 9), ("Layer", 7),
            ("Fused Input Shapes", 55), ("Fused Input Dtypes", 30), ("Input Sources", 60),
            ("Fused Output Shapes", 55), ("Fused Output Dtypes", 30), ("Output Sources", 60),
        ]
        self._write_header(ws, columns)

        for row_idx, rec in enumerate(fused, 2):
            values = [
                rec["node_id"], rec["fused_op"], rec["aten_ops"], rec["num_sub_ops"], rec["layer"],
                rec.get("fused_input_shapes", rec["input_shapes"]),
                rec.get("fused_input_dtypes", rec["input_dtypes"]),
                rec.get("fused_input_sources", ""),
                rec.get("fused_output_shapes", rec["output_shapes"]),
                rec.get("fused_output_dtypes", rec["output_dtypes"]),
                rec.get("fused_output_sources", ""),
            ]
            self._write_row(ws, row_idx, values, get_fill(rec["fused_op"]), center_cols={1, 4})

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(fused) + 1}"
        ws.freeze_panes = "A2"
        logger.info("Fused %d raw ops -> %d fused operators", len(records), len(fused))

    def _write_raw_sheet(self, wb, records):
        ws = wb.create_sheet("Raw Operator Sequence")
        columns = [
            ("Node ID", 8), ("Op Short", 12), ("Aten Op", 35),
            ("Input Shapes", 50), ("Input Dtypes", 30),
            ("Output Shapes", 50), ("Output Dtypes", 30),
            ("Module Path", 55), ("Layer", 7), ("Component", 25),
            ("Source File", 28), ("Line", 6), ("Code", 60), ("Func", 22),
            ("Extra Args", 55),
        ]
        self._write_header(ws, columns)

        for row_idx, rec in enumerate(records, 2):
            values = [
                rec["node_id"], rec.get("op_short", ""), rec["aten_op"],
                rec["input_shapes"], rec["input_dtypes"],
                rec["output_shapes"], rec["output_dtypes"],
                rec["module_path"], rec["layer"], rec["component"],
                rec.get("src_file", ""), rec.get("src_line", ""),
                rec.get("src_code", ""), rec.get("src_func", ""),
                rec.get("extra_args", ""),
            ]
            self._write_row(ws, row_idx, values, get_fill(rec["component"]), center_cols={1, 12})

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"
        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, wb, fused):
        ws = wb.create_sheet("Summary")
        ws.append(["Fused Operator", "Count", "Avg Sub-ops"])
        for col in ("A", "B", "C"):
            ws[f"{col}1"].font = self._header_font

        fused_counts: Dict[str, List[int]] = defaultdict(list)
        for r in fused:
            fused_counts[r["fused_op"]].append(r["num_sub_ops"])
        for comp in sorted(fused_counts.keys()):
            sub_ops = fused_counts[comp]
            ws.append([comp, len(sub_ops), round(sum(sub_ops) / len(sub_ops), 1)])
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12

    def _write_by_layer_sheet(self, wb, records, fused):
        ws = wb.create_sheet("By Layer")
        ws.append(["Layer", "Fused Op Count", "Raw Op Count", "Fused Operators"])
        for col in ("A", "B", "C", "D"):
            ws[f"{col}1"].font = self._header_font

        layer_raw = defaultdict(int)
        for r in records:
            layer_raw[r["layer"] or "non-layer"] += 1
        layer_fused_info = defaultdict(lambda: {"count": 0, "ops": []})
        for r in fused:
            key = r["layer"] or "non-layer"
            layer_fused_info[key]["count"] += 1
            layer_fused_info[key]["ops"].append(r["fused_op"])

        for layer in sorted(layer_fused_info.keys(), key=lambda x: (x == "non-layer", x)):
            info = layer_fused_info[layer]
            seen = set()
            unique_ops = [op for op in info["ops"] if not (op in seen or seen.add(op))]
            ws.append([layer, info["count"], layer_raw.get(layer, 0), ", ".join(unique_ops)])
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 100

    def _write_fusion_rules_sheet(self, wb, fused, output_path):
        fusion_specs = self._fusion_engine.extract_specs(fused)
        ws = wb.create_sheet("Fusion Rules")
        columns = [
            ("Module Class", 30), ("Fusion Level", 12), ("Aten Op Sequence", 80),
            ("Sub-ops", 9), ("Occurrences", 12), ("Example Module Path", 55),
            ("Fused Input Shapes", 55), ("Fused Input Dtypes", 30), ("Input Sources", 65),
            ("Fused Output Shapes", 55), ("Fused Output Dtypes", 30), ("Output Sources", 65),
        ]
        self._write_header(ws, columns)

        for row_idx, spec in enumerate(fusion_specs, 2):
            values = [
                spec.module_class, spec.fusion_level,
                " \u2192 ".join(spec.aten_op_sequence),
                spec.num_sub_ops, spec.occurrences, spec.example_module_path,
                spec.fused_input_shapes, spec.fused_input_dtypes, spec.fused_input_sources,
                spec.fused_output_shapes, spec.fused_output_dtypes, spec.fused_output_sources,
            ]
            self._write_row(ws, row_idx, values, None)

        ws.freeze_panes = "A2"
        logger.info("Discovered %d unique fusion patterns", len(fusion_specs))

        self._export_fusion_json(fusion_specs, output_path)

    def _export_fusion_json(self, specs: List[FusionSpec], output_path: Path):
        json_path = output_path.with_name(output_path.stem + "_fusion_rules.json")
        json_data = [
            {
                "module_class": s.module_class,
                "aten_op_sequence": s.aten_op_sequence,
                "num_sub_ops": s.num_sub_ops,
                "fusion_level": s.fusion_level,
                "example_module_path": s.example_module_path,
                "occurrences": s.occurrences,
                "fused_input_shapes": s.fused_input_shapes,
                "fused_input_dtypes": s.fused_input_dtypes,
                "fused_input_sources": s.fused_input_sources,
                "fused_output_shapes": s.fused_output_shapes,
                "fused_output_dtypes": s.fused_output_dtypes,
                "fused_output_sources": s.fused_output_sources,
            }
            for s in specs
        ]
        json_path.write_text(json.dumps(json_data, indent=2))
        logger.info("Exported fusion rules to %s", json_path)

    def _write_header(self, ws, columns):
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = self._header_font_white
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_row(self, ws, row_idx, values, fill, center_cols=None):
        center_cols = center_cols or set()
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = self._thin_border
            if fill:
                cell.fill = fill
            if col_idx in center_cols:
                cell.alignment = Alignment(horizontal="center")
