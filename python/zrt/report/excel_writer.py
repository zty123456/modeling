"""Write operator records to a formatted Excel workbook + fusion rules JSON."""
from __future__ import annotations

import json
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from python.zrt.graph.classifier import get_fill
from python.zrt.graph.tracker import ModuleTracker

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Write operator records to a formatted Excel workbook + fusion rules JSON."""

    def __init__(self, tracker: ModuleTracker, platform: str = "generic"):
        self._tracker = tracker
        self._platform = platform
        self._header_fill = PatternFill(start_color="263238", end_color="263238", fill_type="solid")
        self._header_font_white = Font(bold=True, color="FFFFFF", size=11)
        self._header_font = Font(bold=True, size=12)
        self._thin_border = Border(bottom=Side(style="thin", color="BDBDBD"))

    def write(self, records: List[Dict[str, Any]], output_path: Path,
              config_summary: Dict[str, Any]):
        wb = openpyxl.Workbook()

        from python.zrt.transform.fusion._dict_bridge import fuse_records
        self._write_config_sheet(wb, config_summary)
        fused = fuse_records(records, self._tracker, platform=self._platform, keep_children=False)
        self._write_fused_sheet(wb, fused, records)
        self._write_raw_sheet(wb, records)
        self._write_summary_sheet(wb, fused)
        self._write_by_layer_sheet(wb, records, fused)
        self._write_fusion_rules_sheet(wb, fused, output_path)

        wb.save(output_path)

    def _write_optimizer_sheet(self, wb, summary: "TrainingSummary") -> None:
        """Write Optimizer sheet (per §6.3 of muon_optimizer_design.md)."""
        ws = wb.create_sheet("Optimizer")
        header_font = Font(bold=True, size=11)

        ws.cell(row=1, column=1, value="Metric").font = header_font
        ws.cell(row=1, column=2, value="Value").font = header_font

        rows = [
            ("optimizer_type", summary.optimizer_type),
            ("muon_param_fraction", f"{summary.muon_param_fraction:.2%}"),
            ("opt_state_gb", round(summary.opt_state_gb, 3)),
            ("opt_state_savings_gb", round(summary.opt_state_savings_gb, 3)),
            ("optimizer_step_ms", round(summary.optimizer_step_ms, 3)),
            ("muon_ag_rs_ms", round(summary.muon_ag_rs_ms, 3)),
            ("muon_ns_tflops", round(summary.muon_ns_tflops, 3)),
            ("optimizer_time_fraction", f"{summary.optimizer_time_fraction:.2%}"),
        ]

        for row_idx, (metric, value) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _write_config_sheet(self, wb, config_summary):
        ws = wb.active
        ws.title = "Model Config"
        ws.append(["Parameter", "Value"])
        ws["A1"].font = self._header_font
        ws["B1"].font = self._header_font
        for key, val in config_summary.items():
            ws.append([key, str(val)])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

    def _write_fused_sheet(self, wb, fused, records):
        ws = wb.create_sheet("Fused Operators")
        columns = [
            ("Node ID", 8), ("Fused Operator", 38), ("Constituent Aten Ops", 70),
            ("Sub-ops", 9), ("Layer", 7),
            ("Fused Input Shapes", 55), ("Fused Input Dtypes", 30), ("Input Sources", 60),
            ("Fused Output Shapes", 55), ("Fused Output Dtypes", 30), ("Output Sources", 60),
        ]
        self._write_header(ws, columns)

        for row_idx, rec in enumerate(fused, 2):
            values = [
                rec["node_id"], rec["fused_op"], rec["aten_ops"], rec["num_sub_ops"], rec["layer"],
                rec.get("fused_input_shapes", rec["input_shapes"]),
                rec.get("fused_input_dtypes", rec["input_dtypes"]),
                rec.get("fused_input_sources", ""),
                rec.get("fused_output_shapes", rec["output_shapes"]),
                rec.get("fused_output_dtypes", rec["output_dtypes"]),
                rec.get("fused_output_sources", ""),
            ]
            self._write_row(ws, row_idx, values, get_fill(rec["fused_op"]), center_cols={1, 4})

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(fused) + 1}"
        ws.freeze_panes = "A2"
        logger.info("Fused %d raw ops -> %d fused operators", len(records), len(fused))

    def _write_raw_sheet(self, wb, records):
        ws = wb.create_sheet("Raw Operator Sequence")
        columns = [
            ("Node ID", 8), ("Op Short", 12), ("Aten Op", 35),
            ("Input Shapes", 50), ("Input Dtypes", 30),
            ("Output Shapes", 50), ("Output Dtypes", 30),
            ("Module Path", 55), ("Layer", 7), ("Component", 25),
            ("Source File", 28), ("Line", 6), ("Code", 60), ("Func", 22),
            ("Extra Args", 55),
        ]
        self._write_header(ws, columns)

        for row_idx, rec in enumerate(records, 2):
            values = [
                rec["node_id"], rec.get("op_short", ""), rec["aten_op"],
                rec["input_shapes"], rec["input_dtypes"],
                rec["output_shapes"], rec["output_dtypes"],
                rec["module_path"], rec["layer"], rec["component"],
                rec.get("src_file", ""), rec.get("src_line", ""),
                rec.get("src_code", ""), rec.get("src_func", ""),
                rec.get("extra_args", ""),
            ]
            self._write_row(ws, row_idx, values, get_fill(rec["component"]), center_cols={1, 12})

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"
        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, wb, fused):
        ws = wb.create_sheet("Summary")
        ws.append(["Fused Operator", "Count", "Avg Sub-ops"])
        for col in ("A", "B", "C"):
            ws[f"{col}1"].font = self._header_font

        fused_counts: Dict[str, List[int]] = defaultdict(list)
        for r in fused:
            fused_counts[r["fused_op"]].append(r["num_sub_ops"])
        for comp in sorted(fused_counts.keys()):
            sub_ops = fused_counts[comp]
            ws.append([comp, len(sub_ops), round(sum(sub_ops) / len(sub_ops), 1)])
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12

    def _write_by_layer_sheet(self, wb, records, fused):
        ws = wb.create_sheet("By Layer")
        ws.append(["Layer", "Fused Op Count", "Raw Op Count", "Fused Operators"])
        for col in ("A", "B", "C", "D"):
            ws[f"{col}1"].font = self._header_font

        layer_raw = defaultdict(int)
        for r in records:
            layer_raw[r["layer"] or "non-layer"] += 1
        layer_fused_info = defaultdict(lambda: {"count": 0, "ops": []})
        for r in fused:
            key = r["layer"] or "non-layer"
            layer_fused_info[key]["count"] += 1
            layer_fused_info[key]["ops"].append(r["fused_op"])

        for layer in sorted(layer_fused_info.keys(), key=lambda x: (x == "non-layer", x)):
            info = layer_fused_info[layer]
            seen = set()
            unique_ops = [op for op in info["ops"] if not (op in seen or seen.add(op))]
            ws.append([layer, info["count"], layer_raw.get(layer, 0), ", ".join(unique_ops)])
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 100

    def _write_fusion_rules_sheet(self, wb, fused, output_path):
        from python.zrt.transform.fusion._dict_bridge import extract_fusion_specs
        fusion_specs = extract_fusion_specs(fused)
        ws = wb.create_sheet("Fusion Rules")
        columns = [
            ("Module Class", 30), ("Fusion Level", 12), ("Aten Op Sequence", 80),
            ("Sub-ops", 9), ("Occurrences", 12), ("Example Module Path", 55),
            ("Fused Input Shapes", 55), ("Fused Input Dtypes", 30), ("Input Sources", 65),
            ("Fused Output Shapes", 55), ("Fused Output Dtypes", 30), ("Output Sources", 65),
        ]
        self._write_header(ws, columns)

        for row_idx, spec in enumerate(fusion_specs, 2):
            values = [
                spec.module_class, spec.fusion_level,
                " \u2192 ".join(spec.aten_op_sequence),
                spec.num_sub_ops, spec.occurrences, spec.example_module_path,
                spec.fused_input_shapes, spec.fused_input_dtypes, spec.fused_input_sources,
                spec.fused_output_shapes, spec.fused_output_dtypes, spec.fused_output_sources,
            ]
            self._write_row(ws, row_idx, values, None)

        ws.freeze_panes = "A2"
        logger.info("Discovered %d unique fusion patterns", len(fusion_specs))

        self._export_fusion_json(fusion_specs, output_path)

    def _export_fusion_json(self, specs: List["FusionSpec"], output_path: Path):
        json_path = output_path.with_name(output_path.stem + "_fusion_rules.json")
        json_data = [
            {
                "module_class": s.module_class,
                "aten_op_sequence": s.aten_op_sequence,
                "num_sub_ops": s.num_sub_ops,
                "fusion_level": s.fusion_level,
                "example_module_path": s.example_module_path,
                "occurrences": s.occurrences,
                "fused_input_shapes": s.fused_input_shapes,
                "fused_input_dtypes": s.fused_input_dtypes,
                "fused_input_sources": s.fused_input_sources,
                "fused_output_shapes": s.fused_output_shapes,
                "fused_output_dtypes": s.fused_output_dtypes,
                "fused_output_sources": s.fused_output_sources,
            }
            for s in specs
        ]
        json_path.write_text(json.dumps(json_data, indent=2))
        logger.info("Exported fusion rules to %s", json_path)

    def _write_summary_perf_sheet(self, ws, summary: "E2ESummary") -> None:
        """Write one E2ESummary into a pre-created worksheet."""
        header_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        section_font = Font(bold=True, size=10, color="1B5E20")

        def _section(label: str) -> None:
            row = ws.max_row + 1
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = section_font
            cell.fill = section_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        def _row(key: str, value: Any) -> None:
            ws.append([key, value])

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 24
        ws.cell(row=1, column=1, value="Metric").font = header_font
        ws.cell(row=1, column=2, value="Value").font = header_font

        _section("── Metadata ──")
        _row("Model",         summary.model)
        _row("Hardware",      summary.hardware)
        _row("Phase",         summary.phase)
        _row("Parallel",      summary.parallel_desc)
        _row("Batch size",    summary.batch_size)
        _row("Seq len",       summary.seq_len)

        _section("── Latency ──")
        _row("Total latency (ms)",  round(summary.latency_ms,     3))
        if summary.ttft_ms is not None:
            _row("TTFT (ms)",       round(summary.ttft_ms,        3))
        if summary.tpot_ms is not None:
            _row("TPOT (ms/token)", round(summary.tpot_ms,        3))
        _row("Throughput (tok/s)",  round(summary.tokens_per_sec, 1))

        _section("── Compute / Comm ──")
        _row("Compute (ms)",        round(summary.compute_ms,      3))
        _row("Comm (ms)",           round(summary.comm_ms,         3))
        _row("Exposed comm (ms)",   round(summary.exposed_comm_ms, 3))
        _row("Overlap ratio",       f"{summary.overlap_ratio:.1%}")

        _section("── HW Efficiency ──")
        _row("MFU",                 f"{summary.mfu:.2%}")
        _row("HBM BW util",         f"{summary.hbm_bandwidth_util:.2%}")
        _row("Total FLOPs (T)",     round(summary.total_flops / 1e12, 4))
        _row("Total bytes (GB)",    round(summary.total_bytes / 1e9,  4))

        if summary.by_component:
            _section("── By Component (% of serial latency) ──")
            for comp, pct in sorted(summary.by_component.items(), key=lambda x: -x[1]):
                _row(comp, f"{pct:.1f}%")

        if summary.by_layer:
            _section("── By Layer (ms) ──")
            for i, lat in enumerate(summary.by_layer):
                _row(f"Layer {i}", round(lat, 4))

        if summary.top_bottleneck_ops:
            _section("── Top Bottleneck Ops (µs) ──")
            for op_desc, lat_us in summary.top_bottleneck_ops:
                _row(op_desc, round(lat_us, 2))

    def _write_header(self, ws, columns):
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = self._header_font_white
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_row(self, ws, row_idx, values, fill, center_cols=None):
        center_cols = center_cols or set()
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = self._thin_border
            if fill:
                cell.fill = fill
            if col_idx in center_cols:
                cell.alignment = Alignment(horizontal="center")


def append_perf_summary(xlsx_path: Path, summary: "E2ESummary") -> None:
    """Open *xlsx_path* and append a performance-report sheet for *summary*.

    Sheet name: ``Perf Report (<phase>)``.  Safe to call multiple times
    (once per phase); each call adds one sheet.
    """
    from python.zrt.report.summary import E2ESummary  # local import avoids circular deps

    wb = openpyxl.load_workbook(xlsx_path)
    sheet_name = f"Perf Report ({summary.phase})"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    writer = ExcelWriter.__new__(ExcelWriter)   # create without __init__ (no tracker needed)
    writer._header_font = Font(bold=True, size=12)
    writer._write_summary_perf_sheet(ws, summary)

    wb.save(xlsx_path)
    logger.info("Appended '%s' sheet to %s", sheet_name, xlsx_path)


def append_optimizer_sheet(xlsx_path: Path, summary: "TrainingSummary") -> None:
    """Open *xlsx_path* and append Optimizer sheet for *summary*.

    Per §6.3 of muon_optimizer_design.md, includes optimizer_type,
    muon_param_fraction, opt_state_gb, optimizer_step_ms, etc.
    """
    wb = openpyxl.load_workbook(xlsx_path)

    if "Optimizer" in wb.sheetnames:
        del wb["Optimizer"]

    writer = ExcelWriter.__new__(ExcelWriter)
    writer._write_optimizer_sheet(wb, summary)

    wb.save(xlsx_path)
    logger.info("Appended 'Optimizer' sheet to %s", xlsx_path)
