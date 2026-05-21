"""Tests for python.zrt.transform pipeline."""
import pytest
from types import SimpleNamespace

from python.zrt.ir.node import OpNode
from python.zrt.ir.edge import Edge
from python.zrt.ir.graph import OpGraph
from python.zrt.ir.types import TensorMeta, DType
import python.zrt.hardware.registry as hw_registry
from python.zrt.transform import (
    ParallelConfig, StreamConfig, QuantConfig, TransformContext,
    TensorParallelPass, ExpertParallelPass, CommInserterPass,
    build_default_pipeline,
)
from python.zrt.transform.analysis.passes import FlopsPass, RooflinePass, StreamAssignPass
from python.zrt.transform.parallel.expert_grouped_mm import (
    ExpertGroupedMMPass,
    _expert_weight_name,
)


# ── helpers ───────────────────────────────────────────────────────────────────

def _t(tid, shape, dtype=DType.BF16):
    return TensorMeta.from_shape_dtype(tid, shape, dtype)


def _linear_node(nid, scope, in_shape, out_shape):
    return OpNode(
        id=nid,
        op_type="aten.mm.default",
        inputs=[_t(f"{nid}_in", in_shape)],
        outputs=[_t(f"{nid}_out", out_shape)],
        scope=scope,
        category="compute",
    )


def simple_linear_graph(tp=1):
    """q_proj (column) → o_proj (row): minimal two-node graph."""
    q = _linear_node("q", "model.layers.0.self_attn.q_proj", (128, 4096), (128, 4096))
    o = _linear_node("o", "model.layers.0.self_attn.o_proj", (128, 4096), (128, 4096))
    edge = Edge(src="q", src_idx=0, dst="o", dst_idx=0,
                tensor=_t("e0", (128, 4096)))
    return OpGraph(name="test", phase="prefill",
                   nodes={"q": q, "o": o},
                   edges=[edge])


def _ctx(tp=1, ep=1, hw_name="nvidia_h100_sxm",
         compute_streams=1, comm_streams=1, quant=None, flags=None):
    hw = hw_registry.load(hw_name)
    return TransformContext(
        hw_spec=hw,
        parallel=ParallelConfig(tp=tp, ep=ep),
        stream_config=StreamConfig(num_compute_streams=compute_streams,
                                   num_comm_streams=comm_streams),
        quant=quant,
        optim_flags=flags or set(),
    )


# ── TensorParallelPass ────────────────────────────────────────────────────────

def test_tp_pass_no_change_when_tp1():
    g = simple_linear_graph()
    ctx = _ctx(tp=1)
    out = TensorParallelPass().run(g, ctx)
    assert out is g   # returned original (tp=1 early return)


def test_tp_pass_column_parallel_shape():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    out = TensorParallelPass().run(g, ctx)
    q_out_shape = out.nodes["q"].outputs[0].shape
    assert q_out_shape[-1] == 4096 // 4, "column parallel: output last dim / tp"


def test_tp_pass_row_parallel_annotation():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    out = TensorParallelPass().run(g, ctx)
    ann = out.nodes["o"].annotations.get("tp_split", {})
    assert ann.get("comm_after") == "all_reduce"


def test_tp_pass_column_parallel_annotation():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    out = TensorParallelPass().run(g, ctx)
    ann = out.nodes["q"].annotations.get("tp_split", {})
    assert ann.get("comm_after") is None
    assert ann.get("tp") == 4


def test_tp_pass_does_not_mutate_original():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    original_shape = g.nodes["q"].outputs[0].shape
    TensorParallelPass().run(g, ctx)
    assert g.nodes["q"].outputs[0].shape == original_shape


def test_tp_pass_skips_routed_experts_but_keeps_shared_experts():
    routed_gate = _linear_node(
        "routed_gate",
        "transformer.layers.0.ffn.experts.0.w1",
        (128, 7168),
        (128, 3072),
    )
    routed_down = _linear_node(
        "routed_down",
        "transformer.layers.0.ffn.experts.0.w2",
        (128, 3072),
        (128, 7168),
    )
    shared_gate = _linear_node(
        "shared_gate",
        "transformer.layers.0.ffn.shared_experts.w1",
        (128, 7168),
        (128, 3072),
    )
    g = OpGraph(
        name="tp_expert_scope",
        phase="prefill",
        nodes={
            "routed_gate": routed_gate,
            "routed_down": routed_down,
            "shared_gate": shared_gate,
        },
        edges=[],
    )
    out = TensorParallelPass().run(g, _ctx(tp=8))

    assert "tp_split" not in out.nodes["routed_gate"].annotations
    assert out.nodes["routed_gate"].outputs[0].shape == (128, 3072)
    assert "tp_split" not in out.nodes["routed_down"].annotations
    assert out.nodes["routed_down"].inputs[0].shape == (128, 3072)
    assert out.nodes["shared_gate"].annotations["tp_split"]["tp"] == 8
    assert out.nodes["shared_gate"].outputs[0].shape == (128, 384)


# ── CommInserterPass ──────────────────────────────────────────────────────────

def test_comm_inserter_adds_allreduce():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    g2 = TensorParallelPass().run(g, ctx)
    g3 = CommInserterPass().run(g2, ctx)

    comm_nodes = [n for n in g3.nodes.values() if n.category == "communication"]
    assert len(comm_nodes) >= 1

    comm_ops = {n.op_type for n in comm_nodes}
    assert "comm.all_reduce" in comm_ops


def test_comm_inserter_all_reduce_after_o_proj():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    g2 = TensorParallelPass().run(g, ctx)
    g3 = CommInserterPass().run(g2, ctx)

    # The all_reduce node should have "o" as its predecessor
    all_reduce = next(
        n for n in g3.nodes.values() if n.op_type == "comm.all_reduce"
    )
    preds = g3.predecessors(all_reduce.id)
    assert "o" in preds


def test_comm_inserter_rewires_successors():
    """Nodes downstream of all_reduce should have the comm node as predecessor."""
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    g2 = TensorParallelPass().run(g, ctx)
    g3 = CommInserterPass().run(g2, ctx)

    # "o" original successors (none in this simple graph, but structure is valid)
    assert g3.num_nodes() == g2.num_nodes() + 1   # one comm node added


def test_comm_inserter_inserts_ep_a2a_per_phase():
    fwd_gate = _linear_node("layer0_grouped_gate_up", "layer.0.ffn.moe", (2, 8), (2, 16))
    fwd_down = _linear_node("layer0_grouped_down", "layer.0.ffn.moe", (2, 16), (2, 8))
    bwd_down = _linear_node("layer0_grouped_down_bwd", "layer.0.ffn.moe", (2, 8), (2, 16))
    bwd_gate = _linear_node("layer0_grouped_gate_up_bwd", "layer.0.ffn.moe", (2, 16), (2, 32))
    fwd_gate.annotations.update({
        "phase": "fwd",
        "ep_needs_a2a": True,
        "ep_block_down_id": "layer0_grouped_down",
    })
    fwd_down.annotations["phase"] = "fwd"
    bwd_down.annotations.update({
        "phase": "bwd",
        "ep_needs_a2a": True,
        "ep_block_down_id": "layer0_grouped_gate_up_bwd",
    })
    bwd_gate.annotations["phase"] = "bwd"
    g = OpGraph(
        name="ep_phase_comm",
        phase="train",
        nodes={n.id: n for n in (fwd_gate, fwd_down, bwd_down, bwd_gate)},
        edges=[
            Edge("layer0_grouped_gate_up", 0, "layer0_grouped_down", 0, fwd_gate.outputs[0]),
            Edge("layer0_grouped_down_bwd", 0, "layer0_grouped_gate_up_bwd", 0, bwd_down.outputs[0]),
        ],
        metadata={"seq_len": 2, "hidden": 8},
    )

    out = CommInserterPass().run(g, _ctx(ep=8))

    expected = {
        "comm_a2a_dispatch_layer0_grouped_gate_up": "fwd",
        "comm_a2a_combine_layer0_grouped_down": "fwd",
        "comm_a2a_dispatch_layer0_grouped_down_bwd": "bwd",
        "comm_a2a_combine_layer0_grouped_gate_up_bwd": "bwd",
    }
    for node_id, phase in expected.items():
        assert node_id in out.nodes
        assert out.nodes[node_id].annotations["phase"] == phase
        assert out.nodes[node_id].attrs["msg_bytes_semantics"] == "per_a2a_direction"
        assert out.nodes[node_id].attrs["dtype_bytes"] == 2


def test_comm_inserter_ep_msg_bytes_uses_ceil_routed_tokens():
    node = _linear_node("layer0_grouped_gate_up", "layer.0.ffn.moe", (2, 8), (2, 16))
    node.annotations.update({
        "phase": "fwd",
        "ep_needs_a2a": True,
        "ep_block_down_id": "layer0_grouped_gate_up",
    })
    graph = OpGraph(
        name="ep_msg_bytes_ceil",
        phase="train",
        nodes={node.id: node},
        metadata={"seq_len": 5, "hidden": 8},
    )
    ctx = _ctx(ep=4)
    ctx.profile = SimpleNamespace(moe_active=3)

    out = CommInserterPass().run(graph, ctx)

    dispatch = out.nodes["comm_a2a_dispatch_layer0_grouped_gate_up"]
    assert dispatch.attrs["msg_bytes"] == 4 * 8 * 2


def test_expert_grouped_mm_backward_preserves_external_outputs_and_gate_up_width():
    src = _linear_node("src", "input", (2, 8), (2, 8))
    down = _linear_node(
        "down_bwd",
        "transformer.layers.0.ffn.experts.0.w2",
        (2, 8),
        (2, 4),
    )
    gate = _linear_node(
        "gate_bwd",
        "transformer.layers.0.ffn.experts.0.w1",
        (2, 4),
        (2, 5),
    )
    up = _linear_node(
        "up_bwd",
        "transformer.layers.0.ffn.experts.0.w3",
        (2, 4),
        (2, 7),
    )
    up.outputs.append(_t("up_bwd_aux_out", (2, 7)))
    gate_sink = _linear_node("gate_sink", "post.gate", (2, 5), (2, 5))
    up_sink = _linear_node("up_sink", "post.up", (2, 7), (2, 7))
    for n in (down, gate, up):
        n.annotations.update({"phase": "bwd", "ep_needs_a2a": True})

    graph = OpGraph(
        name="bwd_grouped",
        phase="train",
        nodes={n.id: n for n in (src, down, gate, up, gate_sink, up_sink)},
        edges=[
            Edge("src", 0, "down_bwd", 0, src.outputs[0]),
            Edge("down_bwd", 0, "gate_bwd", 0, down.outputs[0]),
            Edge("down_bwd", 0, "up_bwd", 0, down.outputs[0]),
            Edge("gate_bwd", 0, "gate_sink", 0, gate.outputs[0]),
            Edge("up_bwd", 1, "up_sink", 0, up.outputs[1]),
        ],
        metadata={"seq_len": 4, "hidden": 8},
    )
    ctx = _ctx(ep=2)
    ctx.profile = SimpleNamespace(num_experts=4, moe_active=2)

    out = ExpertGroupedMMPass().run(graph, ctx)

    gate_up_id = "transformer_layers_0_ffn_grouped_gate_up_bwd"
    assert gate_up_id in out.nodes
    gate_up = out.nodes[gate_up_id]
    assert gate_up.inputs[1].shape == (2, 4, 12)
    assert gate_up.outputs[0].shape == (2, 2, 12)
    grouped_down = out.nodes["transformer_layers_0_ffn_grouped_down_bwd"]
    assert grouped_down.inputs[0].shape == (2, 2, 8)
    assert grouped_down.inputs[1].shape == (2, 8, 4)
    assert grouped_down.outputs[0].shape == (2, 2, 4)
    assert "src" in out.predecessors(grouped_down.id)
    assert "gate_sink" in out.successors(gate_up_id)
    assert "up_sink" in out.successors(gate_up_id)
    outbound = [e for e in out.edges if e.src == gate_up_id and e.dst in {"gate_sink", "up_sink"}]
    assert len(outbound) == 2
    for edge in outbound:
        assert edge.src_idx == 0
        assert edge.tensor.shape == gate_up.outputs[0].shape


def test_expert_grouped_mm_backward_preserves_reachable_external_inputs():
    src = _linear_node("src", "input", (2, 8), (2, 8))
    down = _linear_node(
        "down_bwd",
        "transformer.layers.0.ffn.experts.0.w2",
        (2, 8),
        (2, 4),
    )
    gate = _linear_node(
        "gate_bwd",
        "transformer.layers.0.ffn.experts.0.w1",
        (2, 4),
        (2, 4),
    )
    up = _linear_node(
        "up_bwd",
        "transformer.layers.0.ffn.experts.0.w3",
        (2, 4),
        (2, 4),
    )
    bridge = _linear_node("bridge", "activation.backward", (2, 4), (2, 8))
    sink = _linear_node("sink", "post.gate", (2, 8), (2, 8))
    for n in (down, gate, up):
        n.annotations.update({"phase": "bwd", "ep_needs_a2a": True})

    graph = OpGraph(
        name="bwd_grouped_bridge",
        phase="train",
        nodes={n.id: n for n in (src, down, gate, up, bridge, sink)},
        edges=[
            Edge("src", 0, "down_bwd", 0, src.outputs[0]),
            Edge("gate_bwd", 0, "bridge", 0, gate.outputs[0]),
            Edge("bridge", 0, "down_bwd", 0, bridge.outputs[0]),
            Edge("up_bwd", 0, "sink", 0, up.outputs[0]),
        ],
        metadata={"seq_len": 4, "hidden": 8},
    )
    ctx = _ctx(ep=2)
    ctx.profile = SimpleNamespace(num_experts=4, moe_active=2)

    out = ExpertGroupedMMPass().run(graph, ctx)

    gate_up_id = "transformer_layers_0_ffn_grouped_gate_up_bwd"
    out.topo_sort()
    assert "sink" in out.successors(gate_up_id)
    assert gate_up_id not in out.predecessors("bridge")
    assert "bridge" in out.predecessors("transformer_layers_0_ffn_grouped_down_bwd")


def test_expert_weight_name_matches_path_segments_only():
    assert _expert_weight_name("transformer.layers.0.ffn.experts.0.w1") == "w1"
    assert _expert_weight_name("transformer.layers.0.ffn.experts.0.gate_proj.weight") == "gate_proj"
    assert _expert_weight_name("transformer.layers.0.ffn.experts.0.up_proj.weight") == "up_proj"
    assert _expert_weight_name("transformer.layers.0.ffn.experts.0.down_proj.weight") == "down_proj"
    assert _expert_weight_name("transformer.layers.0.ffn.experts.0.w12_proj") == ""


def test_flops_pass_does_not_ep_scale_expert_grouped_mm_again():
    grouped = OpNode(
        id="grouped",
        op_type="GroupedMatMul",
        inputs=[_t("a", (2, 3, 4)), _t("b", (2, 4, 5))],
        outputs=[_t("out", (2, 3, 5))],
        scope="model.layers.0.ffn.experts.grouped",
        category="compute",
        annotations={
            "fused_by": "expert_grouped_mm",
            "ep_experts_local": 4,
        },
    )
    graph = OpGraph(
        name="flops_grouped",
        phase="train",
        nodes={"grouped": grouped},
        metadata={"moe_active_experts": 6, "moe_total_experts": 8},
    )

    out = FlopsPass().run(graph, _ctx(ep=2))

    assert out.nodes["grouped"].annotations["flops"] == 2 * 2 * 3 * 4 * 5


def test_flops_pass_does_not_active_scale_ep_unfused_expert_helper_without_local_count():
    helper = OpNode(
        id="helper",
        op_type="aten.silu.default",
        inputs=[_t("x", (2, 3))],
        outputs=[_t("y", (2, 3))],
        scope="model.layers.0.ffn.experts.0.activation",
        category="compute",
    )
    graph = OpGraph(
        name="flops_ep_helper",
        phase="train",
        nodes={"helper": helper},
        metadata={"moe_active_experts": 6, "moe_total_experts": 8},
    )

    out = FlopsPass().run(graph, _ctx(ep=2))

    assert out.nodes["helper"].annotations["flops"] == 4 * 2 * 3


def test_excel_export_separates_base_and_effective_flops(tmp_path):
    from openpyxl import load_workbook
    from python.zrt.transform.exporter import TransformedGraphExcelWriter

    node = _linear_node("mm", "model.layers.0.mlp", (2, 4), (2, 8))
    node.annotations.update({"flops": 100, "flops_fwd": 200})
    graph = OpGraph(name="export_flops", phase="train", nodes={"mm": node})
    path = tmp_path / "report.xlsx"

    TransformedGraphExcelWriter().write(graph, _ctx(), path)

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Transformed Operators"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    flops_idx = header.index("FLOPs")
    effective_idx = header.index("Effective FLOPs (with recompute)")
    assert rows[1][flops_idx] == 100
    assert rows[1][effective_idx] == 200


# ── StreamAssignPass ──────────────────────────────────────────────────────────

def test_stream_assign_all_nodes_get_stream():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    g2 = TensorParallelPass().run(g, ctx)
    g3 = CommInserterPass().run(g2, ctx)
    g4 = StreamAssignPass().run(g3, ctx)

    for node in g4.nodes.values():
        assert "stream_id" in node.annotations
        assert "stream_type" in node.annotations


def test_stream_assign_comm_nodes_on_comm_stream():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    g2 = TensorParallelPass().run(g, ctx)
    g3 = CommInserterPass().run(g2, ctx)
    g4 = StreamAssignPass().run(g3, ctx)

    for node in g4.nodes.values():
        if node.category == "communication":
            # comm stream ids start at num_compute_streams
            assert node.annotations["stream_id"] >= ctx.stream_config.num_compute_streams
            assert node.annotations["stream_type"] == "comm"
        else:
            assert node.annotations["stream_id"] < ctx.stream_config.num_compute_streams
            assert node.annotations["stream_type"] == "compute"


def test_stream_assign_multi_comm_streams():
    g = simple_linear_graph()
    ctx = _ctx(tp=4, compute_streams=2, comm_streams=2)
    g2 = TensorParallelPass().run(g, ctx)
    g3 = CommInserterPass().run(g2, ctx)
    g4 = StreamAssignPass().run(g3, ctx)

    comm_sids = {
        n.annotations["stream_id"] for n in g4.nodes.values()
        if n.category == "communication"
    }
    # All comm stream IDs must be in the valid range [2, 3]
    for sid in comm_sids:
        assert 2 <= sid <= 3


# ── FlopsPass ─────────────────────────────────────────────────────────────────

def test_flops_pass_annotates_all_nodes():
    g = simple_linear_graph()
    ctx = _ctx()
    g2 = FlopsPass().run(g, ctx)
    for node in g2.nodes.values():
        assert "flops" in node.annotations
        assert node.annotations["flops"] >= 0


def test_flops_pass_mm_formula():
    """mm (128, 4096) @ (4096, 4096) → flops = 2*128*4096*4096."""
    g = simple_linear_graph()
    ctx = _ctx()
    g2 = FlopsPass().run(g, ctx)
    # q_proj: (128, 4096) inputs[0]; but inputs[1] weight shape may not be stored
    # Just check flops > 0 for a matmul node
    assert g2.nodes["q"].annotations["flops"] > 0


# ── RooflinePass ──────────────────────────────────────────────────────────────

def test_roofline_pass_bound_annotation():
    g = simple_linear_graph()
    ctx = _ctx()
    g2 = FlopsPass().run(g, ctx)
    g3 = RooflinePass().run(g2, ctx)
    for node in g3.nodes.values():
        assert node.annotations["bound"] in ("compute", "memory", "latency")
        assert node.annotations["latency_us"] >= 0


# ── Full pipeline ─────────────────────────────────────────────────────────────

def test_default_pipeline_tp4():
    g = simple_linear_graph()
    ctx = _ctx(tp=4)
    pipe = build_default_pipeline()
    result = pipe.run(g, ctx)

    # comm nodes inserted
    comm = [n for n in result.nodes.values() if n.category == "communication"]
    assert len(comm) >= 1

    # all nodes have stream assignments
    for node in result.nodes.values():
        assert "stream_id" in node.annotations

    # all nodes have flops annotations
    for node in result.nodes.values():
        assert "flops" in node.annotations


def test_default_pipeline_single_device():
    g = simple_linear_graph()
    ctx = _ctx(tp=1)
    pipe = build_default_pipeline()
    result = pipe.run(g, ctx)

    # no comm nodes added
    comm = [n for n in result.nodes.values() if n.category == "communication"]
    assert len(comm) == 0

    # stream assignments still present
    for node in result.nodes.values():
        assert "stream_id" in node.annotations


def test_pipeline_does_not_mutate_input():
    g = simple_linear_graph()
    original_node_count = g.num_nodes()
    ctx = _ctx(tp=4)
    pipe = build_default_pipeline()
    pipe.run(g, ctx)
    assert g.num_nodes() == original_node_count


def test_quant_pass_annotations():
    from python.zrt.transform import QuantizationPass
    g = simple_linear_graph()
    ctx = _ctx(quant=QuantConfig(weight="int8", activation="int8"))
    g2 = QuantizationPass().run(g, ctx)
    for node in g2.nodes.values():
        if node.category == "compute":
            assert node.annotations.get("quant_weight") == "int8"
