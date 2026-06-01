"""Tests for spec-based training Excel export."""
from __future__ import annotations

import shutil
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from zrt.training.io.excel_exporter import export_estimate_excel
from zrt.training.ir.builders import build_graph
from zrt.training.ir.training_graph import Collective
from zrt.training.models.flops import op_cost
from zrt.training.search.training_search_util import _make_system_from_config
from zrt.training.spec.model import LayerKind, ModelSpec
from zrt.training.spec.report import TrainingReport
from zrt.training.spec.strategy import CPKind, Strategy


def test_estimate_excel_includes_comm_domain_sheet():
    model = ModelSpec(
        hidden=128,
        ffn=256,
        num_heads=8,
        num_kv_heads=8,
        head_dim=16,
        vocab=32000,
        seq_len=1024,
        layers=[LayerKind.DENSE, LayerKind.DENSE],
    )
    system = _make_system_from_config({
        "hw": "nvidia_gb300_nvl576",
        "world_size": 128,
    })
    strategy = Strategy(
        tp=4,
        cp=2,
        pp=2,
        ep=2,
        dp=8,
        micro_batch=1,
        global_batch=8,
    )

    output_dir = Path("output") / "test_estimate_excel_comm_domain"
    if output_dir.exists():
        shutil.rmtree(output_dir)

    try:
        output_path = export_estimate_excel(
            report=TrainingReport(step_time_ms=100.0, mfu=0.1),
            graph=SimpleNamespace(ops=[]),
            model=model,
            system=system,
            strategy=strategy,
            op_costs={},
            output_path=output_dir / "estimate.xlsx",
        )

        wb = load_workbook(output_path, data_only=True)
        assert "Comm Domains" in wb.sheetnames
        ws = wb["Comm Domains"]

        headers = [ws.cell(row=1, column=i).value for i in range(1, 7)]
        assert headers == ["Metric", "EP", "PP", "DP", "TP", "CP"]

        rows = {
            ws.cell(row=i, column=1).value: [
                ws.cell(row=i, column=j).value for j in range(2, 7)
            ]
            for i in range(2, ws.max_row + 1)
        }
        assert rows["Group Size"] == [2, 2, 8, 4, 2]
        assert all(rows["Tier"])
        assert all("[" in str(v) for v in rows["Rank Sample"])
    finally:
        if output_dir.exists():
            shutil.rmtree(output_dir)


def test_estimate_excel_summary_includes_operator_time_share(tmp_path):
    model = ModelSpec(
        hidden=128,
        ffn=256,
        num_heads=4,
        num_kv_heads=1,
        head_dim=32,
        vocab=1000,
        seq_len=64,
        layers=[LayerKind.MOE, LayerKind.MOE, LayerKind.MOE],
        model_type="deepseek_v4",
        q_lora_rank=16,
        qk_rope_head_dim=8,
        o_lora_rank=16,
        o_groups=2,
        compress_ratios=[4, 128, 0],
        swa_window=16,
        index_topk=16,
        num_experts=8,
        moe_ffn=256,
        top_k=2,
    )
    system = _make_system_from_config({
        "hw": "nvidia_h100_sxm",
        "world_size": 1,
    })
    strategy = Strategy()
    graph = build_graph(model, strategy)
    op_costs = {op.name: op_cost(op, model, system) for op in graph.ops}

    output_path = export_estimate_excel(
        report=TrainingReport(step_time_ms=100.0, compute_time_ms=50.0),
        graph=graph,
        model=model,
        system=system,
        strategy=strategy,
        op_costs=op_costs,
        output_path=tmp_path / "estimate.xlsx",
    )

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Summary"]
    summary_values = [
        ws.cell(row=i, column=1).value
        for i in range(1, ws.max_row + 1)
    ]

    assert "Operator Time Share" in summary_values
    assert "  Matmul family total" in summary_values
    assert "  Attention matmul family" in summary_values
    assert "  MoE/FFN matmul family" in summary_values
    assert "  LM head matmul" in summary_values
    assert "  CSA attention block" in summary_values
    assert "  HCA attention block" in summary_values
    assert "  CSA/HCA/SWA composite attention core" in summary_values

    header_row = summary_values.index("Operator Time Share") + 1
    assert ws.cell(row=header_row, column=3).value == "% of Step"
    assert ws.cell(row=header_row, column=4).value == "% of Useful Compute"

    for label in ("  Matmul family total", "  Attention matmul family", "  MoE/FFN matmul family", "  LM head matmul"):
        matmul_row = summary_values.index(label) + 1
        step_share = ws.cell(row=matmul_row, column=3).value
        useful_compute_share = ws.cell(row=matmul_row, column=4).value
        assert "ops" in step_share
        assert "ops" in useful_compute_share

        step_pct = float(step_share.split("%", 1)[0])
        useful_compute_pct = float(useful_compute_share.split("%", 1)[0])
        assert abs(useful_compute_pct - step_pct * 2) <= 0.1 + 1e-12


def test_estimate_excel_summary_labels_cp_by_actual_collective_kind(tmp_path):
    model = ModelSpec(
        hidden=128,
        ffn=256,
        num_heads=8,
        num_kv_heads=8,
        head_dim=16,
        vocab=32000,
        seq_len=1024,
        layers=[LayerKind.DENSE],
    )
    system = _make_system_from_config({
        "hw": "nvidia_h100_sxm",
        "world_size": 4,
    })
    strategy = Strategy(cp=2, dp=2, cp_kind=CPKind.RING)
    graph = SimpleNamespace(
        ops=[],
        collectives=[
            Collective(name="cp_ring_fwd", kind="P2P", group="CP", bytes_=1024),
        ],
    )

    output_path = export_estimate_excel(
        report=TrainingReport(step_time_ms=100.0, cp_exposed_ms=5.0),
        graph=graph,
        model=model,
        system=system,
        strategy=strategy,
        op_costs={},
        output_path=tmp_path / "estimate.xlsx",
    )

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Summary"]
    summary_labels = [
        ws.cell(row=i, column=1).value
        for i in range(1, ws.max_row + 1)
    ]

    assert "    CP (P2P)" in summary_labels
    assert "    CP (A2A)" not in summary_labels


def test_estimate_excel_strategy_total_parallelism_excludes_ep(tmp_path):
    model = ModelSpec(
        hidden=128,
        ffn=256,
        num_heads=8,
        num_kv_heads=8,
        head_dim=16,
        vocab=32000,
        seq_len=1024,
        layers=[LayerKind.MOE],
        num_experts=8,
        moe_ffn=256,
        top_k=2,
    )
    system = _make_system_from_config({
        "hw": "nvidia_h100_sxm",
        "world_size": 64,
    })
    strategy = Strategy(tp=2, cp=2, pp=2, ep=4, dp=8)

    output_path = export_estimate_excel(
        report=TrainingReport(step_time_ms=100.0),
        graph=SimpleNamespace(ops=[]),
        model=model,
        system=system,
        strategy=strategy,
        op_costs={},
        output_path=tmp_path / "estimate.xlsx",
    )

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Strategy"]
    rows = {
        ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value
        for i in range(1, ws.max_row + 1)
    }

    assert rows["Total Parallelism"] == 64
