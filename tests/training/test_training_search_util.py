"""Tests for training search utility."""
from __future__ import annotations

import shutil
from pathlib import Path
from types import SimpleNamespace

import pytest
import pandas as pd

from zrt.training.search.training_search_util import (
    TrainingConfigManager,
    _load_model_spec,
    _make_strategy_from_config,
    _make_system_from_config,
    _analysis_value,
    _passes_pod_packing,
    export_best_configs_excel,
    export_best_analysis_excel,
    format_results,
    run_training_search_parallel,
    run_training_task_wrapper,
    save_results,
    select_best_configs_by_tokens,
)
from zrt.training.spec.report import TrainingReport
from zrt.training.spec.model import LayerKind, ModelSpec
from zrt.training.models.memory import MemBreakdown
from zrt.training.spec.strategy import (
    CPKind,
    OptKind,
    PPSched,
    RecomputePolicy,
    TPOverlap,
)


class TestLoadModelSpec:
    """Test ModelSpec loading from YAML."""

    def test_load_deepseek_v3_2(self):
        model = _load_model_spec("deepseek_v3_2")

        assert model.hidden == 7168
        assert model.num_heads == 128
        assert model.num_experts == 256
        assert model.seq_len == 4096
        assert len(model.layers) == 61

    def test_load_invalid_model_raises(self):
        with pytest.raises(FileNotFoundError) as exc_info:
            _load_model_spec("nonexistent_model")

        assert "Model config not found" in str(exc_info.value)


class TestMakeSystemFromConfig:
    """Test SystemSpec construction from config dict."""

    def test_default_values(self):
        config = {}
        system = _make_system_from_config(config)

        assert system.nodes == 1
        assert system.gpus_per_node == 8
        assert system.world_size == 8
        assert "H100" in system.gpu.name

    def test_custom_values(self):
        config = {
            "hw": "nvidia_h100_sxm",
            "nodes": 8,
            "host_mem_gb": 2048.0,
        }
        system = _make_system_from_config(config)

        assert system.nodes == 8
        assert system.gpus_per_node == 8
        assert system.world_size == 64
        assert system.host_mem_gb == 2048.0

    def test_system_gpus_per_node_inferred_from_ascend_superpod_tier(self):
        system = _make_system_from_config({
            "hw": "ascend_910c_superpod",
            "world_size": 384,
        })

        assert system.nodes == 1
        assert system.gpus_per_node == 384
        assert system.world_size == 384
        assert system.allocated_gpus == 384
        assert system.idle_gpus == 0

    def test_system_gpus_per_node_inferred_from_gb300_first_tier(self):
        system = _make_system_from_config({
            "hw": "nvidia_gb300_nvl576",
            "world_size": 72,
        })

        assert system.nodes == 18
        assert system.gpus_per_node == 4
        assert system.world_size == 72
        assert system.allocated_gpus == 72
        assert system.idle_gpus == 0

    def test_search_world_size_preserves_exact_pod_allocation(self):
        system = _make_system_from_config({
            "hw": "nvidia_h100_sxm",
            "world_size": 128,
        })

        assert system.nodes == 16
        assert system.gpus_per_node == 8
        assert system.world_size == 128
        assert system.allocated_gpus == 128
        assert system.idle_gpus == 0

    def test_search_world_size_warns_for_partial_allocation(self, caplog):
        caplog.set_level("WARNING", logger="zrt.training.search.training_search_util")

        system = _make_system_from_config({
            "hw": "ascend_910c_superpod",
            "world_size": 1024,
        })

        assert system.nodes == 3
        assert system.gpus_per_node == 384
        assert system.world_size == 1024
        assert system.allocated_gpus == 1152
        assert system.idle_gpus == 128
        assert any("idle_gpus=128" in r.message for r in caplog.records)

    def test_gpus_per_node_config_is_rejected(self):
        with pytest.raises(ValueError, match="gpus_per_node.*removed"):
            _make_system_from_config({
                "hw": "nvidia_h100_sxm",
                "world_size": 128,
                "gpus_per_node": 8,
            })


class TestMakeStrategyFromConfig:
    """Test Strategy construction from config dict."""

    def test_default_values(self):
        config = {}
        strategy = _make_strategy_from_config(config)

        assert strategy.tp == 1
        assert strategy.cp == 1
        assert strategy.pp == 1
        assert strategy.ep == 1
        assert strategy.dp == 1
        assert strategy.micro_batch == 1
        assert strategy.global_batch == 0
        assert strategy.pp_schedule == PPSched.ONE_F_ONE_B
        assert strategy.vpp_chunks == 1
        assert strategy.zero_stage == 0
        assert strategy.optimizer == OptKind.ADAM
        assert strategy.tp_overlap == TPOverlap.NONE
        assert strategy.cp_kind == CPKind.NONE
        assert strategy.ep_overlap is False
        assert strategy.dualbatch is False
        assert strategy.recompute.per_layer == {}

    def test_custom_parallel_config(self):
        config = {
            "tp": 8,
            "cp": 2,
            "pp": 4,
            "ep": 64,
            "dp": 2,
            "micro_batch": 4,
            "global_batch": 512,
        }
        strategy = _make_strategy_from_config(config)

        assert strategy.tp == 8
        assert strategy.cp == 2
        assert strategy.pp == 4
        assert strategy.ep == 64
        assert strategy.dp == 2
        assert strategy.micro_batch == 4
        assert strategy.global_batch == 512

    def test_recompute_selective(self):
        config = {"recompute": "selective"}
        strategy = _make_strategy_from_config(config)

        assert strategy.recompute.per_layer == {"moe": {"attn"}, "dense": {"attn"}}

    def test_recompute_full(self):
        config = {"recompute": "full"}
        strategy = _make_strategy_from_config(config)

        assert strategy.recompute.per_layer == {"moe": {"full"}, "dense": {"full"}}

    def test_optimizer_muon(self):
        config = {"optimizer": "muon", "muon_rotation": False}
        strategy = _make_strategy_from_config(config)

        assert strategy.optimizer == OptKind.MUON
        assert strategy.muon_config is not None
        assert strategy.muon_config.rotation is False

    def test_pp_schedule_interleaved_with_vpp(self):
        config = {"pp_schedule": "i1f1b", "vpp_chunks": 4}
        strategy = _make_strategy_from_config(config)

        assert strategy.pp_schedule == PPSched.INTERLEAVED
        assert strategy.vpp_chunks == 4

    def test_pp_schedule_dualpipev_with_vpp(self):
        config = {"pp_schedule": "dualpipev", "vpp_chunks": 4}
        strategy = _make_strategy_from_config(config)

        assert strategy.pp_schedule == PPSched.DUALPIPE_V
        assert strategy.vpp_chunks == 4

    def test_pp_schedule_non_interleaved_ignores_vpp(self):
        config = {"pp_schedule": "1f1b", "vpp_chunks": 4}
        strategy = _make_strategy_from_config(config)

        assert strategy.pp_schedule == PPSched.ONE_F_ONE_B
        assert strategy.vpp_chunks == 1

    def test_tp_overlap(self):
        config = {"tp_overlap": "coc"}
        strategy = _make_strategy_from_config(config)

        assert strategy.tp_overlap == TPOverlap.COC

    def test_cp_kind(self):
        config = {"cp_kind": "ulysses"}
        strategy = _make_strategy_from_config(config)

        assert strategy.cp_kind == CPKind.ULYSSES


class TestTrainingConfigManager:
    """Test TrainingConfigManager functionality."""

    def test_generate_static_configs_basic(self):
        manager = TrainingConfigManager(
            param_grid={"world_size": [1], "tp": [1], "cp": [1], "pp": [1], "dp": [1]},
        )
        configs = manager.generate_static_configs()

        assert len(configs) == 1
        assert configs[0]["tp"] == 1

    def test_generate_static_configs_grid_expansion(self):
        manager = TrainingConfigManager(
            param_grid={
                "world_size": [4],
                "tp": [1, 2, 4],
                "cp": [1],
                "pp": [1, 2],
                "dp": [1],
            }
        )
        configs = manager.generate_static_configs()

        valid = [(c["tp"], c["pp"]) for c in configs]
        assert (4, 1) in valid
        assert (2, 2) in valid

    def test_generate_static_configs_filters_invalid_world_size(self):
        manager = TrainingConfigManager(
            param_grid={
                "world_size": [8],
                "tp": [1, 2],
                "cp": [1],
                "pp": [1],
                "dp": [1, 2, 4, 8],
            }
        )
        configs = manager.generate_static_configs()

        for cfg in configs:
            product = cfg["tp"] * cfg["cp"] * cfg["pp"] * cfg["dp"]
            assert product == 8

    def test_generate_static_configs_valid_world_size(self):
        manager = TrainingConfigManager(
            param_grid={
                "world_size": [16],
                "tp": [1, 2, 4],
                "cp": [1, 2],
                "pp": [1, 2],
                "dp": [1, 2],
            }
        )
        configs = manager.generate_static_configs()

        for cfg in configs:
            ws = cfg["world_size"]
            product = cfg["tp"] * cfg["cp"] * cfg["pp"] * cfg["dp"]
            assert ws == product, f"world_size={ws} != tp*cp*pp*dp={product}"

    def test_gpus_per_node_param_grid_is_rejected(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [128],
                "gpus_per_node": [8],
                "tp": [8],
                "cp": [1],
                "pp": [1],
                "ep": [1],
                "dp": [16],
            }
        )

        with pytest.raises(ValueError, match="gpus_per_node.*removed"):
            manager.generate_static_configs()

    def test_tier_aware_pod_packing_uses_ascend_supernode_not_hardcoded_8(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["ascend_910c_superpod"],
                "world_size": [1024],
                "seq_len": [8192],
                "tp": [32],
                "cp": [4],
                "pp": [1],
                "ep": [1],
                "dp": [8],
                "micro_batch": [1],
                "global_batch": [1024],
                "zero_stage": [0],
                "pp_schedule": ["1f1b"],
                "recompute": ["none"],
                "optimizer": ["adam"],
            }
        )

        configs = manager.generate_static_configs()
        combos = {(c["tp"], c["cp"], c["pp"], c["dp"]) for c in configs}

        assert combos == {(32, 4, 1, 8)}
        assert all(c["tp"] * c["cp"] * c["pp"] * c["dp"] == 1024 for c in configs)

    def test_pod_packing_requires_system(self):
        with pytest.raises(ValueError, match="requires system"):
            _passes_pod_packing(
                tp=2, cp=1, pp=1, dp=5,
                target_ws=10, system=None, other_config=None,
            )

    def test_pod_packing_rejects_2tier_partial_tp_crossing_innermost_tier(self):
        system = _make_system_from_config({
            "hw": "nvidia_h100_sxm",
            "world_size": 10,
        })

        assert not _passes_pod_packing(
            tp=10, cp=1, pp=1, dp=1,
            target_ws=10, system=system, other_config=None,
        )

    def test_ep_auto_uses_num_experts_divisors_not_rank_divisors(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [384],
                "seq_len": [4096],
                "tp": [1],
                "cp": [1],
                "pp": [1],
                "ep": ["auto"],
                "dp": [384],
                "micro_batch": [1],
                "global_batch": [384],
                "zero_stage": [0],
                "pp_schedule": ["1f1b"],
                "recompute": ["none"],
                "optimizer": ["adam"],
            }
        )

        configs = manager.generate_static_configs()
        ep_values = {cfg["ep"] for cfg in configs}

        assert 128 in ep_values
        assert 256 not in ep_values
        assert all(384 % ep == 0 for ep in ep_values)
        assert manager.count_total_configs() == len(configs)

    def test_ep_auto_falls_back_to_rank_divisors_without_model_context(self):
        manager = TrainingConfigManager(param_grid={})
        grid = {"tp": [1], "cp": [1], "pp": [1], "ep": ["auto"], "dp": [1]}

        manager._expand_auto_values_optimized(grid, world_size=12, model=None)

        assert grid["ep"] == [1, 2, 3, 4, 6, 12]

    def test_rank_auto_expands_world_size_divisors_for_list_and_scalar_inputs(self):
        manager = TrainingConfigManager(param_grid={})
        grid = {"tp": [2, "auto"], "cp": "auto", "pp": [1], "ep": [1], "dp": [1]}

        manager._expand_auto_values_optimized(grid, world_size=12, model=None)

        assert grid["tp"] == [1, 2, 3, 4, 6, 12]
        assert grid["cp"] == [1, 2, 3, 4, 6, 12]

    def test_ep_auto_collapses_to_one_for_dense_model(self):
        manager = TrainingConfigManager(param_grid={})
        model = ModelSpec(
            hidden=128,
            ffn=256,
            num_heads=8,
            num_kv_heads=8,
            head_dim=16,
            vocab=32000,
            seq_len=1024,
            layers=[LayerKind.DENSE],
            num_experts=0,
        )
        grid = {"tp": [1], "cp": [1], "pp": [1], "ep": "auto", "dp": [1]}

        manager._expand_auto_values_optimized(grid, world_size=12, model=model)

        assert grid["ep"] == [1]

    def test_count_total_configs_matches_generated_configs_without_gpus_per_node_grid(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [128],
                "seq_len": [4096],
                "tp": [4, 8],
                "cp": [1],
                "pp": [1, 2],
                "ep": [1],
                "dp": [8, 16, 32],
                "micro_batch": [1],
                "global_batch": [128],
                "zero_stage": [1],
                "pp_schedule": ["1f1b"],
                "recompute": ["none"],
                "optimizer": ["adam"],
            }
        )

        configs = manager.generate_static_configs()
        assert manager.count_total_configs() == len(configs)

    def test_total_token_derives_exact_global_batch_from_each_seq_len(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [1],
                "seq_len": [512, 256],
                "total_token": [1024],
                "tp": [1],
                "cp": [1],
                "pp": [1],
                "ep": [1],
                "dp": [1],
                "micro_batch": [1],
                "global_batch": [128, 256],
                "zero_stage": [0],
                "pp_schedule": ["1f1b"],
                "recompute": ["none"],
                "optimizer": ["adam"],
            }
        )

        configs = manager.generate_static_configs()

        assert [(c["seq_len"], c["global_batch"]) for c in configs] == [
            (512, 2),
            (256, 4),
        ]
        assert manager.count_total_configs() == len(configs)

    def test_total_token_rejects_seq_len_with_remainder(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [1],
                "seq_len": [512],
                "total_token": [1025],
                "tp": [1],
                "cp": [1],
                "pp": [1],
                "ep": [1],
                "dp": [1],
                "micro_batch": [1],
                "zero_stage": [0],
                "pp_schedule": ["1f1b"],
                "recompute": ["none"],
                "optimizer": ["adam"],
            }
        )

        with pytest.raises(ValueError, match="total_token.*divisible.*seq_len"):
            manager.generate_static_configs()

    def test_search_filters_non_divisible_ep_dp_combos(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [768],
                "seq_len": [4096],
                "tp": [1],
                "cp": [1],
                "pp": [8],
                "ep": [64],
                "dp": [96],
                "micro_batch": [1],
                "global_batch": [768],
                "zero_stage": [1],
                "pp_schedule": ["1f1b"],
                "recompute": ["none"],
                "optimizer": ["muon"],
            }
        )

        configs = manager.generate_static_configs()

        assert configs == []
        assert manager.count_total_configs() == 0

    def test_search_keeps_valid_ep_dp_combos(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["deepseek_v3_2"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [512],
                "seq_len": [4096],
                "tp": [1],
                "cp": [1],
                "pp": [8],
                "ep": [64],
                "dp": [64],
                "micro_batch": [1],
                "global_batch": [512],
                "zero_stage": [1],
                "pp_schedule": ["1f1b"],
                "recompute": ["none"],
                "optimizer": ["muon"],
            }
        )

        configs = manager.generate_static_configs()

        assert [(c["ep"], c["dp"]) for c in configs] == [(64, 64)]
        assert manager.count_total_configs() == len(configs)

    def test_count_total_configs_builds_system_when_model_is_unknown(self):
        manager = TrainingConfigManager(
            param_grid={
                "model": ["unknown"],
                "hw": ["nvidia_h100_sxm"],
                "world_size": [4],
                "tp": [1, 2, 4],
                "cp": [1],
                "pp": [1],
                "ep": [1],
                "dp": [1, 2, 4],
            }
        )

        assert manager.count_total_configs() == len(manager.generate_static_configs())

    def test_worker_uses_search_world_size_not_floor_pod_allocation(self):
        result = run_training_task_wrapper({
            "model": "deepseek_v3_2",
            "hw": "nvidia_h100_sxm",
            "world_size": 8192,
            "seq_len": 8192,
            "tp": 8,
            "cp": 1,
            "pp": 16,
            "ep": 1,
            "dp": 64,
            "micro_batch": 1,
            "global_batch": 8192,
            "zero_stage": 1,
            "pp_schedule": "1f1b",
            "recompute": "none",
            "optimizer": "adam",
            "max_memory_gb": 1024,
        })

        assert result["status"] == "success"
        assert result["report"].config_summary["parallelism"].startswith(
            "TP*CP*PP*DP = 8192"
        )

    def test_muon_rotation_hidden_comm_is_preserved_in_search_summary(self):
        cfg = {
            "model": "deepseek_v3_2",
            "hw": "nvidia_h100_sxm",
            "world_size": 128,
            "seq_len": 4096,
            "tp": 4,
            "cp": 1,
            "pp": 4,
            "ep": 8,
            "dp": 8,
            "micro_batch": 1,
            "global_batch": 512,
            "zero_stage": 1,
            "pp_schedule": "1f1b",
            "recompute": "none",
            "optimizer": "muon",
            "muon_rotation": True,
            "max_memory_gb": 1024,
        }

        result = run_training_task_wrapper(cfg)

        assert result["status"] == "success"
        report = result["report"]
        assert report.optimizer_comm_ms == pytest.approx(0.0)
        assert report.optimizer_comm_hidden_ms > 0.0

        report.memory = None
        df = format_results([report], [cfg])
        row = df.iloc[0]
        assert row["optimizer_exposed_ms"] == pytest.approx(0.0)
        assert row["optimizer_comm_ms"] == pytest.approx(
            round(report.optimizer_comm_hidden_ms, 2)
        )

    def test_worker_skips_memory_infeasible_config_before_graph_build(self, monkeypatch):
        import zrt.training.ir.builders as builders
        import zrt.training.search.training_search_util as search_util

        class FakeStrategy:
            def validate(self, model, system):
                return None

        search_util._WORKER_MODEL_CACHE.clear()
        search_util._WORKER_HW_CACHE.clear()
        search_util._WORKER_GRAPH_CACHE.clear()

        monkeypatch.setattr(
            search_util,
            "_load_model_spec",
            lambda *a, **k: SimpleNamespace(seq_len=4096),
        )
        monkeypatch.setattr(search_util, "_make_strategy_from_config", lambda cfg: FakeStrategy())
        monkeypatch.setattr(
            search_util,
            "_make_system_from_config",
            lambda cfg: SimpleNamespace(gpu=SimpleNamespace(hbm_gb=10), world_size=1),
        )
        monkeypatch.setattr(
            search_util,
            "memory_breakdown",
            lambda graph, model, system, strategy: MemBreakdown(weights=9e9),
        )

        def fail_build_graph(*args, **kwargs):
            raise AssertionError("build_graph should not run for memory-skipped configs")

        monkeypatch.setattr(builders, "build_graph", fail_build_graph)

        result = run_training_task_wrapper({"model": "fake", "hw": "fake", "world_size": 1})

        assert result["status"] == "skipped"
        assert result["type"] == "memory"

    def test_worker_reuses_graph_cache_for_equivalent_graph_configs(self, monkeypatch):
        import zrt.training.ir.builders as builders
        import zrt.training.search.training_search_util as search_util

        class FakeStrategy:
            def __init__(self, dp, global_batch):
                self.dp = dp
                self.global_batch = global_batch

            def validate(self, model, system):
                return None

        search_util._WORKER_MODEL_CACHE.clear()
        search_util._WORKER_HW_CACHE.clear()
        search_util._WORKER_GRAPH_CACHE.clear()

        build_calls = []
        monkeypatch.setattr(
            search_util,
            "_load_model_spec",
            lambda *a, **k: SimpleNamespace(seq_len=4096),
        )
        monkeypatch.setattr(
            search_util,
            "_make_system_from_config",
            lambda cfg: SimpleNamespace(
                gpu=SimpleNamespace(hbm_gb=80),
                world_size=cfg.get("world_size", 1),
            ),
        )
        monkeypatch.setattr(
            search_util,
            "_make_strategy_from_config",
            lambda cfg: FakeStrategy(dp=cfg.get("dp", 1), global_batch=cfg.get("global_batch", 0)),
        )
        monkeypatch.setattr(
            search_util,
            "memory_breakdown",
            lambda graph, model, system, strategy: MemBreakdown(weights=1e9),
        )

        def fake_build_graph(model, strategy):
            build_calls.append((strategy.dp, strategy.global_batch))
            return SimpleNamespace(ops=[], collectives=[])

        monkeypatch.setattr(builders, "build_graph", fake_build_graph)
        monkeypatch.setattr(
            search_util,
            "estimate",
            lambda model, system, strategy, graph=None: TrainingReport(mfu=0.1),
        )

        base = {
            "model": "fake",
            "hw": "fake",
            "world_size": 8,
            "seq_len": 4096,
            "micro_batch": 1,
            "tp": 1,
            "cp": 1,
            "pp": 1,
            "ep": 1,
            "cp_kind": "none",
        }

        first = run_training_task_wrapper({**base, "dp": 8, "global_batch": 8})
        second = run_training_task_wrapper({**base, "dp": 4, "global_batch": 16})

        assert first["status"] == "success"
        assert second["status"] == "success"
        assert len(build_calls) == 1

    def test_output_path_generation(self):
        manager = TrainingConfigManager(
            param_grid={"model": ["test_model"]},
        )

        assert "test_model" in manager.output_path

    def test_output_path_with_world_size(self):
        manager = TrainingConfigManager(
            param_grid={"model": ["test_model"], "world_size": [64]},
        )

        assert "ws_64" in manager.output_path


class TestFormatResults:
    """Test format_results function."""

    def test_format_results_basic(self):
        report = TrainingReport(
            step_time_ms=100.0,
            pipeline_time_ms=95.0,
            mfu=0.45,
            hfu=0.50,
            bubble_fraction=0.1,
            tokens_per_sec=1000.0,
        )
        config = {
            "model": "test_model",
            "world_size": 8,
            "tp": 2,
            "cp": 1,
            "pp": 2,
            "ep": 1,
            "dp": 2,
            "zero_stage": 1,
            "pp_schedule": "1f1b",
            "recompute": "selective",
            "optimizer": "adam",
        }

        df = format_results([report], [config])

        assert len(df) == 1
        assert df.iloc[0]["model"] == "test_model"
        assert df.iloc[0]["world_size"] == 8
        assert df.iloc[0]["tp"] == 2
        assert df.iloc[0]["step_time_ms"] == 100.0
        assert df.iloc[0]["mfu"] == 0.45

    def test_format_results_sorted_by_mfu(self):
        reports = [
            TrainingReport(mfu=0.3, step_time_ms=200),
            TrainingReport(mfu=0.5, step_time_ms=100),
            TrainingReport(mfu=0.4, step_time_ms=150),
        ]
        configs = [{"model": f"model_{i}"} for i in range(3)]

        df = format_results(reports, configs)

        assert df.iloc[0]["mfu"] == 0.5
        assert df.iloc[1]["mfu"] == 0.4
        assert df.iloc[2]["mfu"] == 0.3

    def test_format_results_preserves_small_nonzero_optimizer_compute(self):
        report = TrainingReport(
            step_time_ms=100.0,
            mfu=0.45,
            optimizer_time_ms=0.000330720499,
        )
        config = {"model": "tiny", "hw": "nvidia_h100_sxm"}

        df = format_results([report], [config])

        assert df.iloc[0]["optimizer_compute_ms"] > 0.0

    def test_format_results_includes_comm_totals(self):
        """验证 format_results 包含各策略的通信总时间字段，且列按策略分组."""
        report = TrainingReport(
            step_time_ms=100.0,
            mfu=0.45,
            tp_exposed_ms=5.0,
            tp_hidden_ms=2.0,
            tp_total_ms=7.0,
            cp_exposed_ms=3.0,
            cp_total_ms=3.0,
            ep_exposed_ms=4.0,
            ep_hidden_ms=1.0,
            ep_total_ms=5.0,
            pp_exposed_ms=2.0,
            pp_hidden_ms=0.5,
            pp_total_ms=2.5,
            dp_exposed_ms=6.0,
            dp_hidden_ms=3.0,
            dp_total_ms=9.0,
        )
        config = {"model": "test", "tp": 8, "cp": 2, "ep": 64, "pp": 4}

        df = format_results([report], [config])

        assert "tp_total_ms" in df.columns
        assert "tp_exposed_ms" in df.columns
        assert "cp_total_ms" in df.columns
        assert "cp_exposed_ms" in df.columns
        assert "ep_total_ms" in df.columns
        assert "ep_exposed_ms" in df.columns
        assert "pp_total_ms" in df.columns
        assert "pp_exposed_ms" in df.columns
        assert "pp_hidden_ms" in df.columns
        assert "dp_total_ms" in df.columns
        assert "dp_exposed_ms" in df.columns
        assert df.iloc[0]["tp_total_ms"] == 7.0
        assert df.iloc[0]["tp_exposed_ms"] == 5.0
        assert df.iloc[0]["cp_total_ms"] == 3.0
        assert df.iloc[0]["cp_exposed_ms"] == 3.0
        assert df.iloc[0]["ep_total_ms"] == 5.0
        assert df.iloc[0]["ep_exposed_ms"] == 4.0
        assert df.iloc[0]["pp_total_ms"] == 2.5
        assert df.iloc[0]["pp_exposed_ms"] == 2.0
        assert df.iloc[0]["pp_hidden_ms"] == 0.5
        assert df.iloc[0]["dp_total_ms"] == 9.0
        assert df.iloc[0]["dp_exposed_ms"] == 6.0

    def test_format_results_includes_comm_domain_columns(self):
        report = TrainingReport(step_time_ms=100.0, mfu=0.45)
        config = {
            "model": "test",
            "hw": "nvidia_gb300_nvl576",
            "world_size": 128,
            "tp": 4,
            "cp": 2,
            "pp": 2,
            "ep": 2,
            "dp": 8,
        }

        df = format_results([report], [config])
        row = df.iloc[0]

        for col in [
            "ep_comm_domain",
            "pp_comm_domain",
            "dp_comm_domain",
            "tp_comm_domain",
            "cp_comm_domain",
        ]:
            assert col in df.columns
            assert "size=" in row[col]
            assert "tier=" in row[col]

        assert "size=4" in row["tp_comm_domain"]
        assert "size=8" in row["dp_comm_domain"]

    def test_format_results_includes_memory_and_filters_over_hbm_budget(self):
        reports = [
            TrainingReport(
                mfu=0.5,
                memory=MemBreakdown(
                    weights=1.25e9,
                    grads=0.5e9,
                    opt_state=0.25e9,
                    activations=2.0e9,
                    comm_buffers=0.125e9,
                ),
            ),
            TrainingReport(
                mfu=0.9,
                memory=MemBreakdown(weights=70e9, activations=10e9),
            ),
        ]
        configs = [
            {"model": "fits", "hw": "nvidia_h100_sxm"},
            {"model": "too_large", "hw": "nvidia_h100_sxm"},
        ]

        df = format_results(reports, configs)

        assert list(df["model"]) == ["fits"]
        row = df.iloc[0]
        assert row["weights_gb"] == 1.25
        assert row["grads_gb"] == 0.5
        assert row["opt_state_gb"] == 0.25
        assert row["activations_gb"] == 2.0
        assert row["comm_buffers_gb"] == 0.12
        assert row["memory_gb"] == 4.12


class TestBestAnalysisReport:
    def test_select_best_configs_uses_tokens_per_sec_per_hw_seq(self):
        df = pd.DataFrame([
            {"hw": "hw_a", "seq_len": 8192, "tokens_per_sec": 100.0, "step_time_ms": 10.0},
            {"hw": "hw_a", "seq_len": 8192, "tokens_per_sec": 200.0, "step_time_ms": 20.0},
            {"hw": "hw_b", "seq_len": 8192, "tokens_per_sec": 150.0, "step_time_ms": 15.0},
        ])

        best = select_best_configs_by_tokens(
            df,
            comparison_hw_groups=[["hw_a", "hw_b"]],
            seq_lens=[8192],
        )

        assert list(best["hw"]) == ["hw_a", "hw_b"]
        assert list(best["tokens_per_sec"]) == [200.0, 150.0]

    def test_select_best_configs_allows_same_baseline_in_multiple_groups(self):
        df = pd.DataFrame([
            {"hw": "base", "seq_len": 8192, "tokens_per_sec": 100.0},
            {"hw": "hw_b", "seq_len": 8192, "tokens_per_sec": 90.0},
            {"hw": "hw_c", "seq_len": 8192, "tokens_per_sec": 80.0},
        ])

        best = select_best_configs_by_tokens(
            df,
            comparison_hw_groups=[["base", "hw_b"], ["base", "hw_c"]],
            seq_lens=[8192],
        )

        assert list(best["hw"]) == ["base", "hw_b", "base", "hw_c"]
        assert list(best["_group_idx"]) == [0, 0, 1, 1]

    def test_analysis_values_use_compute_sum_and_group_seq_baseline(self):
        row = pd.Series({
            "组号": "对比1",
            "hw": "hw_b",
            "seq_len": 8192,
            "step_time_ms": 100.0,
            "tokens_per_sec": 150.0,
            "fwd_compute_ms": 20.0,
            "bwd_compute_ms": 30.0,
            "recompute_time_ms": 10.0,
            "tp_exposed_ms": 2.0,
            "ep_exposed_ms": 3.0,
            "pp_exposed_ms": 4.0,
            "dp_exposed_ms": 5.0,
            "cp_exposed_ms": 6.0,
            "optimizer_compute_ms": 7.0,
            "optimizer_exposed_ms": 8.0,
            "bubble_time_ms": 9.0,
        })

        assert _analysis_value("硬件+seq", row, 100.0) == "hw_b_8k"
        assert _analysis_value("计算时间", row, 100.0) == 60.0
        assert _analysis_value("计算占比", row, 100.0) == 0.6
        assert _analysis_value("优化器占比", row, 100.0) == 0.07
        assert _analysis_value("单卡吞吐归一化", row, 100.0) == 1.5

    def test_export_best_analysis_excel_merges_group_label_cells(self):
        output_dir = Path("output") / "test_best_analysis_group_merge"
        if output_dir.exists():
            shutil.rmtree(output_dir)

        df = pd.DataFrame([
            {
                "hw": "base",
                "seq_len": 8192,
                "tokens_per_sec": 100.0,
                "step_time_ms": 10.0,
                "fwd_compute_ms": 1.0,
                "bwd_compute_ms": 2.0,
                "recompute_time_ms": 0.0,
            },
            {
                "hw": "hw_b",
                "seq_len": 8192,
                "tokens_per_sec": 90.0,
                "step_time_ms": 11.0,
                "fwd_compute_ms": 1.0,
                "bwd_compute_ms": 2.0,
                "recompute_time_ms": 0.0,
            },
            {
                "hw": "hw_c",
                "seq_len": 8192,
                "tokens_per_sec": 80.0,
                "step_time_ms": 12.0,
                "fwd_compute_ms": 1.0,
                "bwd_compute_ms": 2.0,
                "recompute_time_ms": 0.0,
            },
        ])
        best = select_best_configs_by_tokens(
            df,
            comparison_hw_groups=[["base", "hw_b"], ["base", "hw_c"]],
            seq_lens=[8192],
        )

        try:
            excel_path = export_best_analysis_excel(
                best,
                str(output_dir),
                comparison_hw_groups=[["base", "hw_b"], ["base", "hw_c"]],
            )

            import openpyxl

            wb = openpyxl.load_workbook(excel_path)
            assert "A2:A3" in {str(rng) for rng in wb["raw_data"].merged_cells.ranges}
            assert "A4:A5" in {str(rng) for rng in wb["raw_data"].merged_cells.ranges}
            assert "A2:A3" in {str(rng) for rng in wb["analysis"].merged_cells.ranges}
            assert "A4:A5" in {str(rng) for rng in wb["analysis"].merged_cells.ranges}
            assert wb["analysis"]["A2"].value == "对比1"
            assert wb["analysis"]["A2"].alignment.horizontal == "center"
            assert wb["analysis"]["A2"].alignment.vertical == "center"
            for sheet_name in ("raw_data", "analysis"):
                ws = wb[sheet_name]
                assert ws["A1"].fill.fgColor.rgb == "004F81BD"
                assert ws["B1"].fill.fgColor.rgb == "004F81BD"
                assert ws["A4"].fill.fgColor.rgb == "00C6EFCE"
                assert ws["B4"].fill.fgColor.rgb == "00C6EFCE"
                assert ws["B5"].fill.fgColor.rgb == "00C6EFCE"
                assert ws["B2"].fill.fgColor.rgb != "00C6EFCE"
        finally:
            if output_dir.exists():
                shutil.rmtree(output_dir)


class TestSearchOutputs:
    """Test search output helper behavior."""

    def test_save_results_writes_summary_csv_and_prints_top_rows(self, capsys):
        output_dir = Path("output") / "test_search_outputs_save"
        if output_dir.exists():
            shutil.rmtree(output_dir)
        df = format_results(
            [TrainingReport(step_time_ms=10.0, mfu=0.4)],
            [{"model": "deepseek_v3_2", "world_size": 8}],
        )

        try:
            save_results(df, str(output_dir))

            csv_path = output_dir / "results_summary.csv"
            assert csv_path.exists()
            assert "deepseek_v3_2" in csv_path.read_text(encoding="utf-8")
            out = capsys.readouterr().out
            assert "Training Search Results" in out
            assert "Total results: 1 configs" in out
        finally:
            if output_dir.exists():
                shutil.rmtree(output_dir)

    def test_export_best_configs_excel_ignores_empty_results(self):
        output_dir = Path("output") / "test_search_outputs_empty"
        if output_dir.exists():
            shutil.rmtree(output_dir)

        export_best_configs_excel([], str(output_dir))

        assert not output_dir.exists()

    def test_export_best_configs_excel_exports_lowest_step_time_per_group(self, monkeypatch):
        import zrt.training.io.excel_exporter as excel_exporter
        import zrt.training.ir.builders as builders
        import zrt.training.models.flops as flops

        output_dir = Path("output") / "test_search_outputs_best"
        if output_dir.exists():
            shutil.rmtree(output_dir)
        exported = []

        monkeypatch.setattr(
            builders,
            "build_graph",
            lambda model, strategy: SimpleNamespace(ops=[SimpleNamespace(name="fake_op")]),
        )
        monkeypatch.setattr(flops, "op_cost", lambda op, model, system: 1.0)

        def fake_export_estimate_excel(**kwargs):
            exported.append(kwargs)
            Path(kwargs["output_path"]).parent.mkdir(parents=True, exist_ok=True)
            Path(kwargs["output_path"]).write_text("excel", encoding="utf-8")

        monkeypatch.setattr(excel_exporter, "export_estimate_excel", fake_export_estimate_excel)

        try:
            export_best_configs_excel(
                [
                    {
                        "model_name": "deepseek_v3_2",
                        "hw_name": "nvidia_h100_sxm",
                        "config": {
                            "seq_len": 4096,
                            "world_size": 8,
                            "tp": 1,
                            "cp": 1,
                            "pp": 1,
                            "ep": 1,
                            "dp": 8,
                        },
                        "report": TrainingReport(mfu=0.2, step_time_ms=10.0),
                    },
                    {
                        "model_name": "deepseek_v3_2",
                        "hw_name": "nvidia_h100_sxm",
                        "config": {
                            "seq_len": 4096,
                            "world_size": 8,
                            "tp": 2,
                            "cp": 1,
                            "pp": 1,
                            "ep": 1,
                            "dp": 4,
                        },
                        "report": TrainingReport(mfu=0.6, step_time_ms=20.0),
                    },
                    {
                        "model_name": "deepseek_v3_2",
                        "hw_name": "nvidia_h100_sxm",
                        "config": {
                            "seq_len": 8192,
                            "world_size": 8,
                            "tp": 4,
                            "cp": 1,
                            "pp": 1,
                            "ep": 1,
                            "dp": 2,
                        },
                        "report": TrainingReport(mfu=0.4, step_time_ms=5.0),
                    },
                ],
                str(output_dir),
            )

            assert len(exported) == 2
            assert exported[0]["strategy"].tp == 1
            assert exported[0]["system"].world_size == 8
            assert exported[0]["op_costs"] == {"fake_op": 1.0}
            assert (output_dir / "deepseek_v3_2_nvidia_h100_sxm_seq4096_ws8_best.xlsx").exists()
            assert (output_dir / "deepseek_v3_2_nvidia_h100_sxm_seq8192_ws8_best.xlsx").exists()
        finally:
            if output_dir.exists():
                shutil.rmtree(output_dir)

    def test_run_training_search_parallel_exports_best_after_mfu_threshold(self, monkeypatch):
        import concurrent.futures
        import zrt.training.search.training_search_util as search_util

        class FakeExecutor:
            def __init__(self, *args, **kwargs):
                pass

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def submit(self, fn, *args, **kwargs):
                fut = concurrent.futures.Future()
                fut.set_result(fn(*args, **kwargs))
                return fut

        configs = [
            {
                "id": "below_threshold_fast",
                "model": "threshold_model",
                "hw": "nvidia_h100_sxm",
                "world_size": 1,
            },
            {
                "id": "above_threshold_slow",
                "model": "threshold_model",
                "hw": "nvidia_h100_sxm",
                "world_size": 1,
            },
        ]

        def fake_task(config):
            if config["id"] == "below_threshold_fast":
                report = TrainingReport(mfu=0.1, step_time_ms=1.0)
            else:
                report = TrainingReport(mfu=0.5, step_time_ms=10.0)
            return {
                "status": "success",
                "config": config,
                "report": report,
                "model_name": config["model"],
                "hw_name": config["hw"],
            }

        exported = []

        monkeypatch.setattr(search_util, "ProcessPoolExecutor", FakeExecutor)
        monkeypatch.setattr(search_util, "_worker_initializer", lambda model_name: None)
        monkeypatch.setattr(search_util, "run_training_task_wrapper", fake_task)
        monkeypatch.setattr(TrainingConfigManager, "count_total_configs", lambda self: len(configs))
        monkeypatch.setattr(
            TrainingConfigManager,
            "generate_static_configs_stream",
            lambda self: iter(configs),
        )
        monkeypatch.setattr(
            search_util,
            "export_best_configs_excel",
            lambda results, output_path: exported.append((results, output_path)),
        )

        output_dir = Path("output") / "training_search" / "threshold_model_ws_1"
        if output_dir.exists():
            shutil.rmtree(output_dir)

        try:
            df = run_training_search_parallel(
                {"model": ["threshold_model"], "world_size": [1]},
                workers=1,
                mfu_threshold=0.2,
                batch_size=1,
                export_best_excel=True,
            )

            assert list(df["id"]) == ["above_threshold_slow"]
            assert len(exported) == 1
            exported_results, _ = exported[0]
            assert [r["config"]["id"] for r in exported_results] == ["above_threshold_slow"]
        finally:
            if output_dir.exists():
                shutil.rmtree(output_dir)

    def test_run_training_search_parallel_exports_model_named_analysis_without_best_csv(self, monkeypatch):
        import concurrent.futures
        import zrt.training.search.training_search_util as search_util

        class FakeExecutor:
            def __init__(self, *args, **kwargs):
                pass

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def submit(self, fn, *args, **kwargs):
                fut = concurrent.futures.Future()
                fut.set_result(fn(*args, **kwargs))
                return fut

        config = {
            "model": "deepseek_v4_pro",
            "hw": "nvidia_h100_sxm",
            "world_size": 1,
            "seq_len": 8192,
        }

        monkeypatch.setattr(search_util, "ProcessPoolExecutor", FakeExecutor)
        monkeypatch.setattr(search_util, "_worker_initializer", lambda model_name: None)
        monkeypatch.setattr(
            search_util,
            "run_training_task_wrapper",
            lambda cfg: {
                "status": "success",
                "config": cfg,
                "report": TrainingReport(
                    mfu=0.5,
                    step_time_ms=10.0,
                    tokens_per_sec=100.0,
                    fwd_compute_ms=1.0,
                    bwd_compute_ms=2.0,
                ),
                "model_name": cfg["model"],
                "hw_name": cfg["hw"],
            },
        )
        monkeypatch.setattr(TrainingConfigManager, "count_total_configs", lambda self: 1)
        monkeypatch.setattr(
            TrainingConfigManager,
            "generate_static_configs_stream",
            lambda self: iter([config]),
        )

        output_dir = Path("output") / "training_search" / "deepseek_v4_pro_ws_1"
        if output_dir.exists():
            shutil.rmtree(output_dir)

        try:
            run_training_search_parallel(
                {
                    "model": ["deepseek_v4_pro"],
                    "hw": ["nvidia_h100_sxm"],
                    "world_size": [1],
                    "seq_len": [8192],
                },
                workers=1,
                batch_size=1,
                export_best_excel=False,
            )

            assert (output_dir / "deepseek_v4_pro_best_config_analysis.xlsx").exists()
            assert not (output_dir / "best_config_analysis.xlsx").exists()
            assert not (output_dir / "best_tokens_per_hw_seq_len.csv").exists()
        finally:
            if output_dir.exists():
                shutil.rmtree(output_dir)

    def test_run_training_search_parallel_returns_empty_when_no_configs(self, monkeypatch):
        import zrt.training.search.training_search_util as search_util

        class FakeExecutor:
            def __init__(self, *args, **kwargs):
                pass

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        monkeypatch.setattr(search_util, "ProcessPoolExecutor", FakeExecutor)
        monkeypatch.setattr(TrainingConfigManager, "count_total_configs", lambda self: 0)
        monkeypatch.setattr(
            TrainingConfigManager,
            "generate_static_configs_stream",
            lambda self: iter(()),
        )

        output_dir = Path("output") / "training_search" / "coverage_noop_ws_1"
        if output_dir.exists():
            shutil.rmtree(output_dir)

        try:
            df = run_training_search_parallel(
                {"model": ["coverage_noop"], "world_size": [1]},
                workers=1,
            )

            assert df.empty
        finally:
            if output_dir.exists():
                shutil.rmtree(output_dir)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
