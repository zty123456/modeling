from __future__ import annotations

import openpyxl
import pytest

from zrt.hardware.spec import InterconnectSpec, LinkSpec
from zrt.training.compose.stage import _cost_phase_time
from zrt.training.io.excel_exporter import export_estimate_excel
from zrt.training.ir.training_graph import Graph, Op, Tensor
from zrt.training.models.flops import op_cost
from zrt.training.spec.dtype import Dtype
from zrt.training.spec.model import LayerKind, ModelSpec
from zrt.training.spec.report import TrainingReport
from zrt.training.spec.strategy import Strategy
from zrt.training.spec.system import GPU, SystemSpec


def _system() -> SystemSpec:
    link = LinkSpec(type="test", bandwidth_gbps=1000, latency_us=1)
    return SystemSpec(
        gpu=GPU(
            name="test-gpu",
            flops_bf16=1,
            flops_fp8=4,
            flops_fp4=8,
            hbm_gb=80,
            hbm_bw_gbps=1_000_000,
        ),
        host_mem_gb=1024,
        interconnect=InterconnectSpec(intra_node=link, inter_node=link),
        nodes=1,
        gpus_per_node=1,
    )


def _model() -> ModelSpec:
    return ModelSpec(
        hidden=128,
        ffn=256,
        num_heads=4,
        num_kv_heads=4,
        head_dim=32,
        vocab=1000,
        seq_len=64,
        layers=[LayerKind.MOE],
        num_experts=8,
        moe_ffn=256,
        top_k=2,
        moe_act_dtype=Dtype.FP8_E4M3,
        routed_expert_compute_dtype=Dtype.FP8_E4M3,
        routed_expert_weight_dtype=Dtype.FP4,
    )


def _routed_op() -> Op:
    x = Tensor("x", (1024, 1024), (1024, 1024), Dtype.FP8_E4M3, True)
    y = Tensor("y", (1024, 1024), (1024, 1024), Dtype.FP8_E4M3, True)
    return Op(
        name="L0.routed_expert_ffn",
        kind="matmul",
        inputs=[x],
        outputs=[y],
        meta={"m": 1024, "n": 1024, "k": 1024, "fwd_multiplier": 6},
        layer_id=0,
        layer_kind=LayerKind.MOE,
        component="routed_expert",
    )


def _ops_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Ops"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [
        dict(zip(header, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row and row[1]
    ]


def _summary_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Summary"]
    return {
        row[0]: row
        for row in ws.iter_rows(values_only=True)
        if row and row[0]
    }


def test_excel_ops_latency_uses_routed_expert_compute_dtype(tmp_path):
    model = _model()
    system = _system()
    strategy = Strategy()
    op = _routed_op()
    graph = Graph(ops=[op], layer_index={0: (0, 1)})
    cost = op_cost(op, model, system)

    path = tmp_path / "report.xlsx"
    export_estimate_excel(
        report=TrainingReport(step_time_ms=1.0),
        graph=graph,
        model=model,
        system=system,
        strategy=strategy,
        op_costs={op.name: cost},
        output_path=path,
    )

    row = _ops_rows(path)[0]
    actual_us = float(row["Latency (μs)"])
    expected_us = (
        _cost_phase_time(cost, "fwd", system, system.gpu.name, 0.0, Dtype.FP8_E4M3)
        + _cost_phase_time(cost, "dx", system, system.gpu.name, 0.0, Dtype.FP8_E4M3)
        + _cost_phase_time(cost, "dw", system, system.gpu.name, 0.0, Dtype.FP8_E4M3)
    ) * 1e6
    bf16_us = (
        _cost_phase_time(cost, "fwd", system, system.gpu.name, 0.0, Dtype.BF16)
        + _cost_phase_time(cost, "dx", system, system.gpu.name, 0.0, Dtype.BF16)
        + _cost_phase_time(cost, "dw", system, system.gpu.name, 0.0, Dtype.BF16)
    ) * 1e6

    assert actual_us == pytest.approx(expected_us, abs=0.01)
    assert actual_us != pytest.approx(bf16_us, rel=1e-3)


def test_excel_operator_time_share_uses_routed_expert_compute_dtype(tmp_path):
    model = _model()
    system = _system()
    strategy = Strategy()
    op = _routed_op()
    graph = Graph(ops=[op], layer_index={0: (0, 1)})
    cost = op_cost(op, model, system)

    path = tmp_path / "report.xlsx"
    export_estimate_excel(
        report=TrainingReport(step_time_ms=100.0, compute_time_ms=0.0),
        graph=graph,
        model=model,
        system=system,
        strategy=strategy,
        op_costs={op.name: cost},
        output_path=path,
    )

    summary = _summary_rows(path)
    actual_ms = float(summary["  MoE/FFN matmul family"][1])
    expected_ms = (
        _cost_phase_time(cost, "fwd", system, system.gpu.name, 0.0, Dtype.FP8_E4M3)
        + _cost_phase_time(cost, "dx", system, system.gpu.name, 0.0, Dtype.FP8_E4M3)
        + _cost_phase_time(cost, "dw", system, system.gpu.name, 0.0, Dtype.FP8_E4M3)
    ) * 1000
    bf16_ms = (
        _cost_phase_time(cost, "fwd", system, system.gpu.name, 0.0, Dtype.BF16)
        + _cost_phase_time(cost, "dx", system, system.gpu.name, 0.0, Dtype.BF16)
        + _cost_phase_time(cost, "dw", system, system.gpu.name, 0.0, Dtype.BF16)
    ) * 1000

    assert actual_ms == pytest.approx(expected_ms, abs=0.01)
    assert actual_ms != pytest.approx(bf16_ms, rel=1e-3)
