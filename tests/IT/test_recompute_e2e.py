"""Recompute E2E tests — full CLI path on DSv4 model.

Validates the final pipeline output: annotations, FLOPs, timing,
memory, and report metrics across none/full/selective policies.

Matches::

    python -m python.zrt --model-id hf_models/deepseek_v4 --layers 4
        --train --hw nvidia_h100_sxm --tp 1
        --recompute-policy none|full|selective
"""
from __future__ import annotations

import pytest

pytestmark = pytest.mark.recompute

_HW, _TP, _EP = "nvidia_h100_sxm", 1, 1
_HIDDEN, _SEQ, _LAYERS, _BATCH = 7168, 128, 4, 1
_BWD_PHASES = {"bwd", "backward", "train_backward"}


def _is_backward_node(node):
    return node.annotations.get("phase", "") in _BWD_PHASES


def _is_forward_node(node):
    return not _is_backward_node(node)


def _forward_nodes(graph):
    return [n for n in graph.nodes.values() if _is_forward_node(n)]


def _backward_nodes(graph):
    return [n for n in graph.nodes.values() if _is_backward_node(n)]


def _base_forward_flops(graph):
    return sum(n.annotations.get("flops", 0) for n in _forward_nodes(graph))


def _recompute_flops_oracle(graph):
    from python.zrt.transform.training.recompute import is_external_recompute_node

    return sum(
        n.annotations.get("flops_fwd", 0) // 2
        for n in _forward_nodes(graph)
        if is_external_recompute_node(n)
    )


def _forward_flops_annotation_sum(graph):
    return sum(n.annotations.get("flops_fwd", 0) for n in _forward_nodes(graph))


def _backward_flops_annotation_sum(graph):
    return sum(n.annotations.get("flops_fwd", 0) for n in _backward_nodes(graph))


def _saved_activation_bytes_oracle(graph, tp=_TP, cp=1):
    fwd_ids = {n.id for n in _forward_nodes(graph)}
    bwd_ids = {n.id for n in _backward_nodes(graph)}
    recomputed = {
        n.id for n in _forward_nodes(graph) if n.annotations.get("recompute")
    }
    saved_bytes = 0
    for edge in graph.edges:
        if edge.src in fwd_ids and edge.dst in bwd_ids and edge.src not in recomputed:
            saved_bytes += getattr(edge.tensor, "mem_bytes", 0)
    return saved_bytes / (max(tp, 1) * max(cp, 1))


def _excel_summary_map(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Training Summary"]
    result = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] not in (None, ""):
            result[str(row[0])] = row[1] if len(row) > 1 else None
    return result


def _sheet_rows(path, sheet_name):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in values):
                continue
            rows.append(dict(zip(header, values)))
        return rows
    finally:
        wb.close()


def _header_index(header, prefix):
    for idx, name in enumerate(header):
        if str(name).startswith(prefix):
            return idx
    raise AssertionError(f"{prefix!r} column not found in {header!r}")


# ═══════════════════════════════════════════════════════════════════════════════
# E2E fixtures (self-contained)
# ═══════════════════════════════════════════════════════════════════════════════

def _capture_model():
    from python.zrt.pipeline import run_trace_phases
    try:
        return run_trace_phases(
            model_id="hf_models/deepseek_v4",
            num_layers=4, batch_size=1, seq_len=128,
            phases=("train_forward", "train_backward"),
        )
    except AssertionError as e:
        if "Mixing fake modes" in str(e):
            pytest.skip("FakeTensorMode already in use by another test module")
        raise

@pytest.fixture(scope="module")
def captured_model():
    return _capture_model()

def _run_estimate(policy, captured_model):
    from python.zrt.transform.analysis import estimate_training_from_graphs
    import python.zrt.hardware.registry as hw_registry
    hw = hw_registry.load(_HW)
    return estimate_training_from_graphs(
        forward_graph=captured_model.graphs["train_forward"],
        backward_graph=captured_model.graphs["train_backward"],
        hw_spec=hw, tp=_TP, ep=_EP, hidden=_HIDDEN, num_layers=_LAYERS,
        seq_len=_SEQ, batch_size=_BATCH,
        recompute_policy=policy, model_id="hf_models/deepseek_v4",
        return_transformed=True,
    )


@pytest.fixture(scope="module")
def none_all(captured_model):
    return _run_estimate("none", captured_model)


@pytest.fixture(scope="module")
def full_all(captured_model):
    return _run_estimate("full", captured_model)


@pytest.fixture(scope="module")
def sel_all(captured_model):
    return _run_estimate("selective", captured_model)


@pytest.fixture(scope="module")
def full_excel(tmp_path_factory, full_all):
    from python.zrt.transform.exporter import export_training_graphs

    report, ctx, transformed = full_all
    out = tmp_path_factory.mktemp("recompute_full_report")
    graph = transformed["unified"]
    paths = export_training_graphs(
        fwd_graph=graph,
        bwd_graph=graph,
        ctx=ctx,
        output_dir=out,
        training_summary=report,
    )
    return paths["excel"], graph, report


@pytest.fixture(scope="module")
def none_excel(tmp_path_factory, none_all):
    from python.zrt.transform.exporter import export_training_graphs

    report, ctx, transformed = none_all
    out = tmp_path_factory.mktemp("recompute_none_report")
    graph = transformed["unified"]
    paths = export_training_graphs(
        fwd_graph=graph,
        bwd_graph=graph,
        ctx=ctx,
        output_dir=out,
        training_summary=report,
    )
    return paths["excel"], graph, report


@pytest.fixture(scope="module")
def full_final_excel(tmp_path_factory, full_all):
    from python.zrt.transform.exporter import export_training_graphs

    report, ctx, transformed = full_all
    graph = transformed["unified"]
    recompute_ms = graph.metadata.get("recompute_compute_ms", 0.0)
    assert recompute_ms > 0
    out = tmp_path_factory.mktemp("recompute_full_final_report")
    paths = export_training_graphs(
        fwd_graph=graph,
        bwd_graph=None,
        ctx=ctx,
        output_dir=out,
        training_summary=report,
    )
    return paths["excel"], graph


# ═══════════════════════════════════════════════════════════════════════════════
# tests
# ═══════════════════════════════════════════════════════════════════════════════

class TestRecomputeE2E:

    # ── annotation ─────────────────────────────────────────────────────────

    def test_none_no_annotations(self, none_all):
        _, _, t = none_all
        rc = [n for n in t["unified"].nodes.values() if n.annotations.get("recompute")]
        assert len(rc) == 0

    def test_full_all_fwd_annotated(self, full_all):
        _, _, t = full_all
        for n in t["unified"].nodes.values():
            phase = n.annotations.get("phase", "fwd")
            if phase not in ("bwd", "backward", "train_backward"):
                assert n.annotations.get("recompute") is True, \
                    f"{n.id} not annotated under full"

    def test_selective_annotates_subset(self, full_all, sel_all):
        _, _, tf = full_all; _, _, ts = sel_all
        cf = sum(1 for n in tf["unified"].nodes.values() if n.annotations.get("recompute"))
        cs = sum(1 for n in ts["unified"].nodes.values() if n.annotations.get("recompute"))
        assert 0 < cs < cf, f"sel={cs} should be between 0 and full={cf}"

    def test_bwd_nodes_excluded(self, full_all):
        _, _, t = full_all
        for n in t["unified"].nodes.values():
            phase = n.annotations.get("phase", "fwd")
            if phase in ("bwd", "backward", "train_backward"):
                assert not n.annotations.get("recompute"), \
                    f"{n.id} bwd node annotated"

    # ── shape ─────────────────────────────────────────────────────────────

    def test_shapes_unchanged(self, none_all, full_all, sel_all):
        ref = {n.id: [t.shape for t in n.outputs] for n in none_all[2]["unified"].nodes.values()}
        for tag, (_, _, t) in ("full", full_all), ("sel", sel_all):
            u = t["unified"]
            for nid, shapes in ref.items():
                if nid in u.nodes:
                    for i, s in enumerate(shapes):
                        if i < len(u.nodes[nid].outputs):
                            assert u.nodes[nid].outputs[i].shape == s, \
                                f"policy={tag} {nid}[{i}] changed"

    # ── FLOPs ─────────────────────────────────────────────────────────────

    def test_recompute_nodes_2x(self, full_all):
        from python.zrt.transform.training.recompute import is_external_recompute_node

        _, _, t = full_all
        for n in t["unified"].nodes.values():
            flops = n.annotations.get("flops", 0)
            ff = n.annotations.get("flops_fwd", 0)
            if is_external_recompute_node(n) and flops > 0:
                assert ff == flops * 2, f"{n.id}: flops_fwd={ff} != 2x{flops}"

    def test_internal_recompute_ops_not_double_charged(self, full_all):
        from python.zrt.transform.training.recompute import has_internal_recompute

        _, _, t = full_all
        internal = [
            n for n in _forward_nodes(t["unified"])
            if n.annotations.get("recompute") and has_internal_recompute(n)
        ]
        if not internal:
            pytest.skip("captured DSv4 E2E graph uses unfused attention in this environment")
        for n in internal:
            assert n.annotations.get("flops_fwd", 0) != n.annotations.get("flops", 0) * 2
            assert n.annotations.get("recompute_flops", 0) == 0

    def test_non_recompute_nodes_1x(self, none_all):
        _, _, t = none_all
        for n in t["unified"].nodes.values():
            flops = n.annotations.get("flops", 0)
            ff = n.annotations.get("flops_fwd", 0)
            if flops > 0:
                assert ff == flops, f"{n.id}: flops_fwd={ff} != flops={flops}"

    def test_metadata_has_recompute_flops(self, full_all, none_all):
        u_full = full_all[2]["unified"]
        u_none = none_all[2]["unified"]
        assert u_full.metadata.get("recompute_flops", 0) > 0
        assert u_none.metadata.get("recompute_flops", 0) == 0

    def test_full_recompute_flops_matches_oracle(self, full_all):
        _, _, t = full_all
        u = t["unified"]
        expected = _recompute_flops_oracle(u)
        assert expected > 0
        assert u.metadata.get("recompute_flops", 0) == expected

    def test_selective_recompute_flops_matches_oracle(self, sel_all):
        _, _, t = sel_all
        u = t["unified"]
        expected = _recompute_flops_oracle(u)
        assert expected > 0
        assert u.metadata.get("recompute_flops", 0) == expected

    @pytest.mark.parametrize("fixture_name", ["full_all", "sel_all"])
    def test_recompute_forward_flops_split_is_exact(self, request, fixture_name):
        _, _, t = request.getfixturevalue(fixture_name)
        u = t["unified"]
        recompute = u.metadata.get("recompute_flops", 0)
        assert u.metadata.get("forward_flops", 0) == _forward_flops_annotation_sum(u)
        assert u.metadata.get("forward_flops", 0) - recompute == _base_forward_flops(u)

    @pytest.mark.parametrize("fixture_name", ["full_all", "sel_all"])
    def test_recompute_training_flops_identity(self, request, fixture_name):
        _, _, t = request.getfixturevalue(fixture_name)
        u = t["unified"]
        assert u.metadata.get("forward_flops", 0) == _forward_flops_annotation_sum(u)
        assert u.metadata.get("backward_flops", 0) == _backward_flops_annotation_sum(u)
        assert u.metadata.get("training_flops", 0) == (
            u.metadata.get("forward_flops", 0) + u.metadata.get("backward_flops", 0)
        )

    # ── timing ────────────────────────────────────────────────────────────

    def _phase_latency_sum(self, g, phase):
        return sum(n.annotations.get("latency_us", 0)
                   for n in g.nodes.values()
                   if n.annotations.get("phase") == phase)

    def _forward_activation_memory_sum(self, g):
        return sum(n.annotations.get("activation_memory_us", 0.0) for n in _forward_nodes(g))

    def test_fwd_activation_save_time_reduced_by_recompute(self, none_all, full_all, sel_all):
        u_none = none_all[2]["unified"]
        u_full = full_all[2]["unified"]
        u_sel  = sel_all[2]["unified"]
        f_none = self._phase_latency_sum(u_none, "fwd")
        f_full = self._phase_latency_sum(u_full, "fwd")
        f_sel  = self._phase_latency_sum(u_sel, "fwd")
        assert f_full < f_none, \
            f"full recompute should reduce fwd activation save time: full={f_full} none={f_none}"
        assert f_full <= f_sel <= f_none, \
            f"fwd latency order broken: none={f_none} sel={f_sel} full={f_full}"

        save_none = self._forward_activation_memory_sum(u_none)
        save_full = self._forward_activation_memory_sum(u_full)
        save_sel = self._forward_activation_memory_sum(u_sel)
        assert save_none > 0
        assert save_full == pytest.approx(0.0)
        assert save_full <= save_sel <= save_none

    def test_bwd_latency_full_greater_than_none(self, none_all, full_all):
        b_none = self._phase_latency_sum(none_all[2]["unified"], "bwd")
        b_full = self._phase_latency_sum(full_all[2]["unified"], "bwd")
        assert b_full > b_none, \
            f"bwd latency not increased: full={b_full} < none={b_none}"

    def test_bwd_latency_sel_between(self, none_all, full_all, sel_all):
        b_none = self._phase_latency_sum(none_all[2]["unified"], "bwd")
        b_full = self._phase_latency_sum(full_all[2]["unified"], "bwd")
        b_sel  = self._phase_latency_sum(sel_all[2]["unified"], "bwd")
        assert b_none <= b_sel <= b_full, \
            f"bwd: none={b_none} sel={b_sel} full={b_full}"

    def test_step_time_full_greater_than_none(self, none_all, full_all):
        r_none, _, _ = none_all
        r_full, _, _ = full_all
        assert r_full.step_time_ms > r_none.step_time_ms, \
            f"step: full={r_full.step_time_ms}ms <= none={r_none.step_time_ms}ms"

    def test_step_time_sel_between(self, none_all, full_all, sel_all):
        r_none, _, _ = none_all
        r_full, _, _ = full_all
        r_sel,  _, _ = sel_all
        assert r_none.step_time_ms <= r_sel.step_time_ms <= r_full.step_time_ms, \
            f"step: none={r_none.step_time_ms} sel={r_sel.step_time_ms} full={r_full.step_time_ms}"

    # ── efficiency ────────────────────────────────────────────────────────

    def test_mfu_full_lower_than_none(self, none_all, full_all):
        r_none, _, _ = none_all
        r_full, _, _ = full_all
        assert r_full.mfu < r_none.mfu, \
            f"MFU: full={r_full.mfu} >= none={r_none.mfu}"

    def test_hfu_exceeds_mfu(self, full_all):
        _, _, t = full_all
        pm = t["unified"].metadata.get("pipeline_metrics")
        if pm is not None:
            assert pm.hfu > pm.mfu, f"HFU={pm.hfu} <= MFU={pm.mfu}"

    def test_none_mfu_equals_hfu(self, none_all):
        _, _, t = none_all
        pm = t["unified"].metadata.get("pipeline_metrics")
        assert pm is not None
        assert pm.hfu == pytest.approx(pm.mfu)

    @pytest.mark.parametrize("fixture_name", ["full_all", "sel_all"])
    def test_recompute_hfu_exceeds_mfu(self, request, fixture_name):
        _, _, t = request.getfixturevalue(fixture_name)
        u = t["unified"]
        pm = u.metadata.get("pipeline_metrics")
        assert pm is not None
        assert u.metadata.get("recompute_flops", 0) > 0
        assert pm.hfu > pm.mfu

    # ── memory ────────────────────────────────────────────────────────────

    def test_activation_memory_reduced(self, none_all, full_all):
        mb_none = none_all[2]["unified"].metadata.get("memory_breakdown")
        mb_full = full_all[2]["unified"].metadata.get("memory_breakdown")
        assert mb_none is not None and mb_full is not None
        assert mb_full.activations < mb_none.activations, \
            f"activations: full={mb_full.activations} >= none={mb_none.activations}"

    # ── per-node memory access ────────────────────────────────────────────

    @pytest.mark.parametrize("fixture_name", ["full_all", "sel_all"])
    def test_recompute_activation_memory_matches_saved_edge_oracle(self, request, fixture_name):
        _, _, t = request.getfixturevalue(fixture_name)
        u = t["unified"]
        mb = u.metadata.get("memory_breakdown")
        assert mb is not None
        assert mb.activations == pytest.approx(_saved_activation_bytes_oracle(u, _TP, 1))

    def test_exported_excel_recompute_time_is_separate(self, full_excel):
        excel_path, graph, report = full_excel
        recompute_compute_ms = graph.metadata.get("recompute_compute_ms", 0.0)
        summary = _excel_summary_map(excel_path)
        assert "Backward compute (ms)" in summary
        assert "Recompute compute (ms)" in summary
        assert float(summary["Recompute compute (ms)"]) > 0
        assert float(summary["Backward compute (ms)"]) == pytest.approx(
            report.bwd_compute_ms, abs=1e-3
        )
        assert float(summary["Recompute compute (ms)"]) == pytest.approx(
            recompute_compute_ms, abs=1e-3
        )
        assert report.compute_time_ms == pytest.approx(
            report.fwd_compute_ms + report.bwd_compute_ms + recompute_compute_ms,
            abs=1e-6,
        )

    def test_exported_excel_recompute_ops_match_oracle(self, full_excel):
        import openpyxl
        from python.zrt.transform.training.recompute import has_internal_recompute

        excel_path, graph, report = full_excel
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        assert "Recompute Ops" in wb.sheetnames
        ws = wb["Recompute Ops"]
        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[0])
        total_row = next(row for row in rows if row and row[0] == "TOTAL")
        flops_idx = _header_index(header, "FLOPs")
        latency_idx = _header_index(header, "Latency")
        assert int(total_row[flops_idx]) == _recompute_flops_oracle(graph)
        assert float(total_row[latency_idx]) / 1000.0 == pytest.approx(
            graph.metadata.get("recompute_compute_ms", 0.0),
            abs=1e-3,
        )
        node_ids = {row[0] for row in rows[1:] if row and row[0] not in (None, "TOTAL")}
        for node_id in node_ids:
            assert not has_internal_recompute(graph.nodes[node_id])

    def test_forward_activation_save_time_is_per_node_final_latency(self, none_all, full_all):
        for graph in (none_all[2]["unified"], full_all[2]["unified"]):
            for node in _forward_nodes(graph):
                activation_us = node.annotations.get("activation_memory_us", 0.0)
                if node.annotations.get("recompute"):
                    assert node.annotations.get("saved_activation_bytes", 0) == 0
                    assert activation_us == pytest.approx(0.0)
                else:
                    assert activation_us >= 0
                assert node.annotations.get("latency_us", 0.0) == pytest.approx(
                    node.annotations.get("base_latency_us", 0.0) + activation_us,
                    abs=1e-6,
                )

    def test_exported_forward_ops_show_reduced_activation(self, full_excel, none_excel):
        excel_path, graph, _report = full_excel
        rows = _sheet_rows(excel_path, "Forward Operators")
        for col in (
            "Activation (B)",
            "Activation Memory (µs)",
            "Final Latency (µs)",
        ):
            assert col in rows[0]
        assert "Checkpoint Memory (µs)" not in rows[0]

        recompute_rows = [
            row for row in rows
            if row["Node ID"] in graph.nodes
            and graph.nodes[row["Node ID"]].annotations.get("recompute")
        ]
        assert recompute_rows
        assert all(float(row["Activation (B)"] or 0) == 0 for row in recompute_rows)
        assert all(float(row["Activation Memory (µs)"] or 0) == 0 for row in recompute_rows)
        for row in recompute_rows[:20]:
            node = graph.nodes[row["Node ID"]]
            assert float(row["Final Latency (µs)"]) == pytest.approx(
                node.annotations.get("latency_us", 0.0),
                abs=1e-3,
            )

        none_path, none_graph, _ = none_excel
        none_rows = _sheet_rows(none_path, "Forward Operators")
        activation_rows = [
            row for row in none_rows
            if row["Node ID"] in none_graph.nodes
            and float(row.get("Activation (B)") or 0) > 0
        ]
        assert activation_rows
        assert any(float(row["Activation Memory (µs)"] or 0) > 0 for row in activation_rows)

    def test_exported_backward_ops_show_recompute_replay(self, full_excel):
        excel_path, graph, _report = full_excel
        rows = _sheet_rows(excel_path, "Backward Operators")
        assert "Recompute Replay (µs)" in rows[0]
        assert "Final Latency (µs)" in rows[0]

        replay_rows = [
            row for row in rows
            if float(row.get("Recompute Replay (µs)") or 0) > 0
        ]
        assert replay_rows
        for row in replay_rows[:20]:
            node = graph.nodes[row["Node ID"]]
            replay = node.annotations.get("recompute_latency_us", 0.0)
            final = node.annotations.get("latency_us", 0.0)
            assert float(row["Recompute Replay (µs)"]) == pytest.approx(replay, abs=1e-3)
            assert float(row["Final Latency (µs)"]) == pytest.approx(final, abs=1e-3)
            assert final == pytest.approx(
                node.annotations.get("base_latency_us", 0.0) + replay,
                abs=1e-6,
            )

    def test_full_training_report_recompute_time_comes_from_graph(self, full_final_excel):
        excel_path, graph = full_final_excel
        summary = _excel_summary_map(excel_path)
        assert "Recompute compute (ms)" in summary
        assert float(summary["Recompute compute (ms)"]) == pytest.approx(
            graph.metadata.get("recompute_compute_ms", 0.0),
            abs=1e-3,
        )

    def test_read_write_bytes_unchanged(self, none_all, full_all):
        u_none = none_all[2]["unified"]
        u_full = full_all[2]["unified"]
        for nid, n_none in u_none.nodes.items():
            if nid not in u_full.nodes:
                continue
            n_full = u_full.nodes[nid]
            assert n_none.annotations.get("read_bytes") == n_full.annotations.get("read_bytes"), \
                f"{nid} read_bytes differs"
            assert n_none.annotations.get("write_bytes") == n_full.annotations.get("write_bytes"), \
                f"{nid} write_bytes differs"
