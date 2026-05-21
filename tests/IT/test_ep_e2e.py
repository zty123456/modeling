"""EP E2E tests — full CLI path on DSv4 model.

Validates the final pipeline output: A2A operators, expert annotations,
shapes, timing, operator counts, and report metrics across EP=1/EP=8.

Matches::

    python -m python.zrt --model-id hf_models/deepseek_v4 --layers 4
        --train --hw nvidia_h100_sxm --tp 8 --ep N
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

pytestmark = pytest.mark.ep

_EP, _TP = 8, 8
_NUM_EXPERTS, _MOE_ACTIVE = 384, 6
_MOE_INTERMEDIATE = 3072
_HIDDEN, _SEQ_LEN, _BATCH = 7168, 128, 1


# ═══════════════════════════════════════════════════════════════════════════════
# fixtures
# ═══════════════════════════════════════════════════════════════════════════════

def _capture_model():
    from python.zrt.pipeline import run_trace_phases
    try:
        return run_trace_phases(
            model_id="hf_models/deepseek_v4",
            num_layers=4, batch_size=1, seq_len=128,
            phases=("train_forward", "train_backward"),
        )
    except AssertionError as e:
        if "Mixing fake modes" in str(e):
            pytest.skip("FakeTensorMode already in use by another test module")
        raise


@pytest.fixture(scope="module")
def captured_model():
    return _capture_model()


def _run_estimate(ep, captured_model):
    from python.zrt.transform.analysis import estimate_training_from_graphs
    import python.zrt.hardware.registry as hw_registry
    hw = hw_registry.load("nvidia_h100_sxm")
    return estimate_training_from_graphs(
        forward_graph=captured_model.graphs["train_forward"],
        backward_graph=captured_model.graphs["train_backward"],
        hw_spec=hw, tp=_TP, ep=ep, hidden=_HIDDEN, num_layers=4,
        seq_len=_SEQ_LEN, batch_size=_BATCH,
        moe_total_experts=_NUM_EXPERTS, moe_active_experts=_MOE_ACTIVE,
        model_id="hf_models/deepseek_v4",
        return_transformed=True,
    )


@pytest.fixture(scope="module")
def ep8_all(captured_model):
    return _run_estimate(_EP, captured_model)


@pytest.fixture(scope="module")
def ep1_all(captured_model):
    return _run_estimate(1, captured_model)


# ═══════════════════════════════════════════════════════════════════════════════
# tests
# ═══════════════════════════════════════════════════════════════════════════════

def _export_training_artifacts(tmp_path: Path, ep_all):
    from python.zrt.transform.exporter import export_training_graphs

    report, ctx, transformed = ep_all
    out = tmp_path / f"ep{ctx.parallel.ep}"
    out.mkdir(parents=True, exist_ok=True)
    graph = transformed["unified"]

    export_paths = export_training_graphs(
        fwd_graph=graph,
        bwd_graph=graph,
        ctx=ctx,
        output_dir=out,
        training_summary=report,
    )

    report_dir = out / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    json_path = report_dir / "deepseek_v4_training_report.json"
    json_path.write_text(json.dumps(report.to_dict(), indent=2), encoding="utf-8")

    return {
        "output_dir": out,
        "excel": export_paths["excel"],
        "json": json_path,
    }


@pytest.fixture(scope="module")
def ep8_artifacts(tmp_path_factory, ep8_all):
    return _export_training_artifacts(tmp_path_factory.mktemp("ep8_artifacts"), ep8_all)


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, object]] = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in values):
                continue
            rows.append(dict(zip(header, values)))
        return rows
    finally:
        wb.close()


def _summary_map(path: Path) -> dict[str, object]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Training Summary"]
        return {
            str(row[0]): row[1]
            for row in ws.iter_rows(values_only=True)
            if row and row[0] not in (None, "")
        }
    finally:
        wb.close()


class TestEPE2E:

    # ── annotation / A2A ───────────────────────────────────────────────────

    def test_capture_succeeded(self, captured_model):
        assert captured_model.graphs["train_forward"].num_nodes() > 0
        assert captured_model.graphs["train_backward"].num_nodes() > 0

    def test_expert_nodes_annotated(self, ep8_all):
        _, _, t = ep8_all
        experts = [n for n in t["unified"].nodes.values() if n.annotations.get("ep_needs_a2a")]
        assert len(experts) > 0
        for n in experts:
            assert n.annotations["ep_experts_local"] == _NUM_EXPERTS // _EP

    def test_a2a_roles(self, ep8_all):
        _, _, t = ep8_all
        a2a = [n for n in t["unified"].nodes.values() if n.op_type == "comm.all_to_all"]
        assert len(a2a) > 0
        assert {n.attrs.get("role") for n in a2a} == {"dispatch", "combine"}

    def test_a2a_tensor_ids_semantic(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if n.op_type != "comm.all_to_all":
                continue
            kw = "dispatch" if n.attrs.get("role") == "dispatch" else "combine"
            assert any(kw in x.id.lower() for x in n.inputs + n.outputs)

    def test_a2a_msg_bytes_formula(self, ep8_all):
        _, _, t = ep8_all
        u = t["unified"]
        s, h = u.metadata.get("seq_len", _SEQ_LEN), u.metadata.get("hidden", _HIDDEN)
        expected = _BATCH * s * h * _MOE_ACTIVE * 2 // _EP
        for n in u.nodes.values():
            if n.op_type == "comm.all_to_all":
                assert n.attrs["msg_bytes"] == expected
                assert n.attrs["msg_bytes_semantics"] == "per_a2a_direction"

    def test_shared_expert_not_epoch_annotated(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if "shared_expert" in n.scope.lower():
                assert "ep_needs_a2a" not in n.annotations

    def test_router_not_epoch_annotated(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if "gate" in n.scope.lower() and "moe" in n.scope.lower():
                assert "ep_needs_a2a" not in n.annotations

    # ── A2A completeness ───────────────────────────────────────────────────

    def test_a2a_has_comm_category(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if n.op_type == "comm.all_to_all":
                assert n.category == "communication"
                assert n.annotations["inserted_by"] == "ep_pass"
                assert n.attrs["group_size"] == _EP
                assert n.attrs["collective"] == "all_to_all"
                assert n.attrs["msg_bytes"] > 0

    def test_a2a_nodes_have_phase(self, ep8_all):
        """After B1 fix, A2A nodes should inherit phase from source."""
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if n.op_type == "comm.all_to_all":
                assert "phase" in n.annotations, \
                    f"A2A {n.id} missing phase annotation"

    # ── GroupedMM ──────────────────────────────────────────────────────────

    def test_grouped_mm_exists(self, ep8_all):
        _, _, t = ep8_all
        grouped = [n for n in t["unified"].nodes.values() if n.op_type == "GroupedMatMul"]
        assert len(grouped) == 16, f"Expected 16 GroupedMM, got {len(grouped)}"

    def test_grouped_mm_per_moe_layer(self, ep8_all):
        _, _, t = ep8_all
        grouped = [n for n in t["unified"].nodes.values() if n.op_type == "GroupedMatMul"]
        role_counts = {}
        for node in grouped:
            role = node.annotations.get("grouped_mm_role")
            role_counts[role] = role_counts.get(role, 0) + 1
        assert role_counts == {
            "gate_up": 4,
            "down": 4,
            "down_bwd": 4,
            "gate_up_bwd": 4,
        }

    def test_grouped_mm_replaces_routed_experts(self, ep8_all):
        _, _, t = ep8_all
        grouped = [n for n in t["unified"].nodes.values() if n.op_type == "GroupedMatMul"]
        assert len(grouped) > 0
        # Only check forward-phase nodes; backward keeps its own expert structure
        for n in t["unified"].nodes.values():
            if n.annotations.get("phase", "fwd") in ("bwd", "backward", "train_backward"):
                continue
            if "shared_expert" in n.scope.lower():
                continue
            if "expert" in n.scope.lower() or "experts." in n.scope.lower():
                assert n.op_type == "GroupedMatMul", \
                    f"{n.id} ({n.scope}, phase={n.annotations.get('phase')}): should be GroupedMM"

    def test_grouped_mm_group_count(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if n.op_type == "GroupedMatMul":
                assert n.inputs[0].shape[0] == _NUM_EXPERTS // _EP

    def test_grouped_mm_token_count(self, ep8_all):
        _, _, t = ep8_all
        expected_M = _BATCH * _SEQ_LEN * _MOE_ACTIVE // _NUM_EXPERTS
        for n in t["unified"].nodes.values():
            if n.op_type == "GroupedMatMul":
                assert n.inputs[0].shape[1] == expected_M

    def test_grouped_mm_shapes_match_dsv4_experts(self, ep8_all):
        _, _, t = ep8_all
        G = _NUM_EXPERTS // _EP
        M = _BATCH * _SEQ_LEN * _MOE_ACTIVE // _NUM_EXPERTS
        for n in t["unified"].nodes.values():
            if n.op_type != "GroupedMatMul" or n.annotations.get("phase") != "fwd":
                continue
            role = n.annotations["grouped_mm_role"]
            if role == "gate_up":
                assert n.inputs[0].shape == (G, M, _HIDDEN)
                assert n.inputs[1].shape == (G, _HIDDEN, _MOE_INTERMEDIATE * 2)
                assert n.outputs[0].shape == (G, M, _MOE_INTERMEDIATE * 2)
            elif role == "down":
                assert n.inputs[0].shape == (G, M, _MOE_INTERMEDIATE)
                assert n.inputs[1].shape == (G, _MOE_INTERMEDIATE, _HIDDEN)
                assert n.outputs[0].shape == (G, M, _HIDDEN)
            else:
                pytest.fail(f"Unknown GroupedMM role: {role}")

    def test_backward_grouped_mm_shapes_match_dsv4_experts(self, ep8_all):
        _, _, t = ep8_all
        G = _NUM_EXPERTS // _EP
        M = _BATCH * _SEQ_LEN * _MOE_ACTIVE // _NUM_EXPERTS
        bwd = [
            n for n in t["unified"].nodes.values()
            if n.op_type == "GroupedMatMul"
            and n.annotations.get("phase") == "bwd"
        ]
        roles = {n.annotations.get("grouped_mm_role") for n in bwd}
        assert {"down_bwd", "gate_up_bwd"} <= roles
        for n in bwd:
            role = n.annotations["grouped_mm_role"]
            if role == "down_bwd":
                assert n.inputs[0].shape == (G, M, _HIDDEN)
                assert n.inputs[1].shape == (G, _HIDDEN, _MOE_INTERMEDIATE)
                assert n.outputs[0].shape == (G, M, _MOE_INTERMEDIATE)
            elif role == "gate_up_bwd":
                assert n.inputs[0].shape == (G, M, _MOE_INTERMEDIATE)
                assert n.inputs[1].shape == (G, _MOE_INTERMEDIATE, _HIDDEN * 2)
                assert n.outputs[0].shape == (G, M, _HIDDEN * 2)

    def test_shared_expert_not_grouped(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if "shared_expert" in n.scope.lower():
                assert "Grouped" not in n.op_type

    def test_swiglu_between_grouped_mm(self, ep8_all):
        _, _, t = ep8_all
        u = t["unified"]
        grouped = [n for n in u.nodes.values() if n.op_type == "GroupedMatMul"]
        for gm in grouped:
            if any("silu" in u.nodes[s].op_type.lower() for s in u.successors(gm.id)):
                return
        pytest.fail("No SwiGLU after any GroupedMM")

    # ── shape ─────────────────────────────────────────────────────────────

    def test_hidden_dim_unchanged(self, ep8_all, ep1_all):
        _, _, t8 = ep8_all; _, _, t1 = ep1_all
        n8 = {n.id: n for n in t8["unified"].nodes.values() if not n.is_comm}
        n1 = {n.id: n for n in t1["unified"].nodes.values() if not n.is_comm}
        for nid in n8.keys() & n1.keys():
            for i in range(min(len(n8[nid].outputs), len(n1[nid].outputs))):
                assert n8[nid].outputs[i].shape == n1[nid].outputs[i].shape

    def test_dispatch_input_shape(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if n.op_type == "comm.all_to_all" and n.attrs.get("role") == "dispatch":
                assert n.inputs[0].shape == (_BATCH, _SEQ_LEN, _HIDDEN)

    def test_combine_output_shape(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            if n.op_type == "comm.all_to_all" and n.attrs.get("role") == "combine":
                assert n.outputs[0].shape == (_BATCH, _SEQ_LEN, _HIDDEN)

    # ── timing ────────────────────────────────────────────────────────────

    def test_step_time_ep8_differs_from_ep1(self, ep8_all, ep1_all):
        r8, _, _ = ep8_all
        r1, _, _ = ep1_all
        assert r8.step_time_ms != r1.step_time_ms, \
            f"EP=8 and EP=1 should differ: both={r8.step_time_ms}ms"

    # ── operator counts ───────────────────────────────────────────────────

    def test_ep1_no_ep(self, ep1_all):
        _, _, t = ep1_all
        for n in t["unified"].nodes.values():
            assert "ep_needs_a2a" not in n.annotations
        assert len([n for n in t["unified"].nodes.values() if n.op_type == "comm.all_to_all"]) == 0

    def test_compute_nodes_ep8_less(self, ep8_all, ep1_all):
        """EP=8 + GroupedMM merges experts → fewer compute nodes."""
        _, _, t8 = ep8_all; _, _, t1 = ep1_all
        assert len(t8["unified"].compute_nodes()) < len(t1["unified"].compute_nodes())

    def test_node_count_diff(self, ep8_all, ep1_all):
        _, _, t8 = ep8_all; _, _, t1 = ep1_all
        diff = t8["unified"].num_nodes() - t1["unified"].num_nodes()
        a2a = len([n for n in t8["unified"].nodes.values() if n.op_type == "comm.all_to_all"])
        # EP=8 adds A2A but GroupedMM reduces nodes. Net effect may be negative.
        assert a2a > 0 and diff != 0, f"a2a={a2a}, diff={diff}"

    def test_a2a_symmetry(self, ep8_all):
        _, _, t = ep8_all
        d = [n for n in t["unified"].nodes.values()
             if n.op_type == "comm.all_to_all" and n.attrs.get("role") == "dispatch"]
        c = [n for n in t["unified"].nodes.values()
             if n.op_type == "comm.all_to_all" and n.attrs.get("role") == "combine"]
        assert len(d) == len(c)
        for dp, cp in zip(d, c):
            assert dp.attrs["msg_bytes"] == cp.attrs["msg_bytes"]

    # ── graph integrity ───────────────────────────────────────────────────

    def test_dag(self, ep8_all):
        _, _, t = ep8_all
        u = t["unified"]
        assert len(u.topo_sort()) == u.num_nodes()

    def test_all_nodes_have_flops_and_stream(self, ep8_all):
        _, _, t = ep8_all
        for n in t["unified"].nodes.values():
            assert "flops" in n.annotations and n.annotations["flops"] >= 0
            assert "stream_id" in n.annotations

    # ── report ────────────────────────────────────────────────────────────

    def test_report_metrics(self, ep8_all):
        r, _, _ = ep8_all
        assert r.step_time_ms > 0 and 0 < r.mfu <= 1.0 and r.training_flops > 0

    def test_step_time_differs(self, ep8_all, ep1_all):
        assert ep8_all[0].step_time_ms != ep1_all[0].step_time_ms

    def test_context(self, ep8_all):
        _, ctx, _ = ep8_all
        assert ctx.parallel.ep == _EP and ctx.parallel.tp == _TP
        assert ctx.profile.num_experts == _NUM_EXPERTS

    # ── exported report artifacts ─────────────────────────────────────────

    def test_exported_training_artifacts_exist(self, ep8_artifacts):
        assert ep8_artifacts["excel"].exists()
        assert ep8_artifacts["json"].exists()

    def test_exported_excel_grouped_mm_has_weight_inputs_and_formula(self, ep8_artifacts):
        rows = _sheet_rows(ep8_artifacts["excel"], "Forward Operators")
        grouped = [r for r in rows if r["Op Type"] == "GroupedMatMul"]
        assert grouped, "Expected GroupedMatMul rows in exported Excel"
        for row in grouped:
            input_shapes = str(row["Input Shapes"])
            assert input_shapes.count("(") >= 2, input_shapes
            assert "2" in str(row["FLOPs Formula (sym)"])
            assert "G" in str(row["FLOPs Formula (sym)"])
            assert int(row["FLOPs"]) > int(row["Activation (B)"])

    def test_exported_excel_grouped_mm_shapes_match_dsv4_experts(self, ep8_artifacts):
        rows = _sheet_rows(ep8_artifacts["excel"], "Forward Operators")
        grouped = {str(r["Node ID"]): r for r in rows if r["Op Type"] == "GroupedMatMul"}
        G = _NUM_EXPERTS // _EP
        M = _BATCH * _SEQ_LEN * _MOE_ACTIVE // _NUM_EXPERTS
        for layer in range(4):
            prefix = f"transformer_layers_{layer}_ffn"
            gate_up = grouped[f"{prefix}_grouped_gate_up"]
            down = grouped[f"{prefix}_grouped_down"]
            assert str((G, M, _HIDDEN)) in str(gate_up["Input Shapes"])
            assert str((G, _HIDDEN, _MOE_INTERMEDIATE * 2)) in str(gate_up["Input Shapes"])
            assert str((G, M, _MOE_INTERMEDIATE * 2)) in str(gate_up["Output Shapes"])
            assert str((G, M, _MOE_INTERMEDIATE)) in str(down["Input Shapes"])
            assert str((G, _MOE_INTERMEDIATE, _HIDDEN)) in str(down["Input Shapes"])
            assert str((G, M, _HIDDEN)) in str(down["Output Shapes"])

    def test_exported_excel_backward_grouped_mm_shapes_and_order(self, ep8_artifacts):
        rows = _sheet_rows(ep8_artifacts["excel"], "Backward Operators")
        ids = [str(r["Node ID"]) for r in rows]
        grouped = {str(r["Node ID"]): r for r in rows if r["Op Type"] == "GroupedMatMul"}
        G = _NUM_EXPERTS // _EP
        M = _BATCH * _SEQ_LEN * _MOE_ACTIVE // _NUM_EXPERTS
        for layer in range(4):
            prefix = f"transformer_layers_{layer}_ffn"
            expected = [
                f"comm_a2a_dispatch_{prefix}_grouped_down_bwd",
                f"{prefix}_grouped_down_bwd",
                f"{prefix}_grouped_gate_up_bwd",
                f"comm_a2a_combine_{prefix}_grouped_gate_up_bwd",
            ]
            positions = [ids.index(x) for x in expected]
            assert positions == sorted(positions), list(zip(expected, positions))
            down = grouped[f"{prefix}_grouped_down_bwd"]
            gate_up = grouped[f"{prefix}_grouped_gate_up_bwd"]
            assert str((G, M, _HIDDEN)) in str(down["Input Shapes"])
            assert str((G, _HIDDEN, _MOE_INTERMEDIATE)) in str(down["Input Shapes"])
            assert str((G, M, _MOE_INTERMEDIATE)) in str(down["Output Shapes"])
            assert str((G, M, _MOE_INTERMEDIATE)) in str(gate_up["Input Shapes"])
            assert str((G, _MOE_INTERMEDIATE, _HIDDEN * 2)) in str(gate_up["Input Shapes"])
            assert str((G, M, _HIDDEN * 2)) in str(gate_up["Output Shapes"])

    def test_exported_excel_ep_forward_order(self, ep8_artifacts):
        rows = _sheet_rows(ep8_artifacts["excel"], "Forward Operators")
        fwd_ep = [
            r for r in rows
            if r["Op Type"] in ("comm.all_to_all", "GroupedMatMul", "aten.silu")
            and (
                r.get("Role") in ("dispatch", "combine")
                or "grouped_" in str(r.get("Node ID", ""))
            )
        ]
        ids = [str(r["Node ID"]) for r in fwd_ep]
        for layer in range(4):
            prefix = f"transformer_layers_{layer}_ffn"
            expected = [
                f"comm_a2a_dispatch_{prefix}_grouped_gate_up",
                f"{prefix}_grouped_gate_up",
                f"{prefix}_grouped_silu",
                f"{prefix}_grouped_down",
                f"comm_a2a_combine_{prefix}_grouped_down",
            ]
            positions = [ids.index(x) for x in expected]
            assert positions == sorted(positions), list(zip(expected, positions))

    def test_exported_excel_ep_a2a_communication(self, ep8_artifacts):
        rows = _sheet_rows(ep8_artifacts["excel"], "Communication Ops")
        ep_rows = [r for r in rows if r["Collective Op"] == "all_to_all"]
        roles = [r["Role"] for r in ep_rows]
        assert roles.count("dispatch") == roles.count("combine") > 0
        routed_tokens_per_ep_rank = (_BATCH * _SEQ_LEN * _MOE_ACTIVE + _EP - 1) // _EP
        expected = routed_tokens_per_ep_rank * _HIDDEN * 2
        for row in ep_rows:
            assert row["Group Size"] == _EP
            assert int(row["Data Volume (bytes)"]) == expected

    def test_exported_excel_training_summary(self, ep8_artifacts):
        summary = _summary_map(ep8_artifacts["excel"])
        assert "TP8" in str(summary["Parallelism"])
        assert "EP8" in str(summary["Parallelism"])
        assert summary["Model"] == "hf_models/deepseek_v4"
        assert summary["Hardware"] == "NVIDIA H100 SXM"
        assert summary["Batch size"] == _BATCH
        assert summary["Sequence length"] == _SEQ_LEN
        assert float(summary["Step latency (ms)"]) > 0
        assert str(summary["MFU"]).endswith("%")

    def test_exported_json_report_metrics(self, ep8_artifacts):
        data = json.loads(ep8_artifacts["json"].read_text(encoding="utf-8"))
        assert data["step_time_ms"] > 0
        assert data["training_flops"] > 0
        assert data["mfu"] > 0
        assert "GroupedMatMul" in data["fused_ops_summary"]
