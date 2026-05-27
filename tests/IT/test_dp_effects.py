"""Integration test: verify how DataParallel (DP) affects TrainingReport fields.

This test runs the training modelling CLI for dp=1, dp=4, dp=8, loads the
generated ``deepseek_v4_training_report.json`` files and asserts expected
relationships:

- optimizer state (opt_state) per-GPU ≈ 1/dp  (ZeRO-1)
- total per-GPU memory drops monotonically as dp increases
- step_time decreases and tokens/sec increases monotonically as dp increases
- dp_hidden_ms and dp_exposed_ms are zero for dp=1, non-zero for dp>1
- dp_comm_total (dp_hidden + dp_exposed) increases monotonically with dp
- optimizer state scales approximately as 1/dp across dp=4 and dp=8

This is a long-running integration test that captures real reports. To avoid
running it by default in fast CI, it is skipped unless the environment
variable ``RUN_DP_TEST`` is set to ``1``.

Run locally (PowerShell):

```powershell
$env:RUN_DP_TEST='1'; pytest tests/IT/test_dp_effects.py -q
```
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import pytest

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class _DPResult:
    """Report dict + output directory, with dict-like access for backward compat."""

    def __init__(self, report: dict, out_dir: Path):
        self._report = report
        self.out_dir = out_dir

    def __getitem__(self, key):
        return self._report[key]

    def get(self, key, default=None):
        return self._report.get(key, default)


def _run_cli_and_load_report(repo_root: Path, outdir: Path, dp: int, timeout: int = 900) -> _DPResult:
    """Run `python -m python.zrt` with given dp and return report + out_dir.

    Raises subprocess.CalledProcessError on failure.
    """
    env = os.environ.copy()
    env["PYTHONPATH"] = str(repo_root / "python")

    cmd = [
        sys.executable,
        "-m",
        "python.zrt",
        "--model-id",
        "hf_models/deepseek_v4",
        "--train",
        "--hw",
        "nvidia_h100_sxm",
        "--dp",
        str(dp),
        "--layers",
        "4",
        "--batch-size",
        "1",
        "--seq-len",
        "128",
        "--output-dir",
        str(outdir),
    ]

    proc = subprocess.run(cmd, check=True, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, timeout=timeout)
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "cli_output.log").write_text(proc.stdout)

    report_path = outdir / "reports" / "deepseek_v4_training_report.json"
    assert report_path.exists(), f"Report not found at {report_path}"
    return _DPResult(json.loads(report_path.read_text()), outdir)


@pytest.fixture(scope="session")
def dp_reports(tmp_path_factory):
    """Session-scoped fixture: run CLI for dp=1, dp=4, dp=8, return all reports."""
    if os.environ.get("RUN_DP_TEST") != "1":
        pytest.skip("Set RUN_DP_TEST=1 to run this long integration test")

    repo_root = Path(__file__).resolve().parents[3]
    tmp_path = tmp_path_factory.mktemp("dp_effects")

    reports = {}
    for dp in (1, 4, 8):
        out_dir = tmp_path / f"out_dp{dp}"
        reports[dp] = _run_cli_and_load_report(repo_root, out_dir, dp=dp)

    return reports


def _load_xlsx_sheets(xlsx_path: Path) -> dict[str, list[dict]]:
    """Load all sheets from training xlsx, return {sheet_name: [row_dict]}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        data = []
        for row in rows[1:]:
            vals = [v for v in row]
            if all(v is None or v == "" for v in vals):
                continue
            data.append(dict(zip(headers, vals)))
        result[sheet_name] = data
    wb.close()
    return result


def _filter_dp_comm(comm_sheet: list[dict]) -> list[dict]:
    """Return only DP gradient reduction communication rows."""
    return [
        r for r in comm_sheet
        if r.get("Role") == "dp_grad_reduce"
        or "dp_grad" in str(r.get("Role", ""))
    ]


@pytest.fixture(scope="session")
def dp_excel_data(dp_reports):
    """Session fixture: load Excel sheets from CLI runs cached by dp_reports."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl required: pip install openpyxl")

    data = {}
    for dp in (1, 4):
        out_dir = dp_reports[dp].out_dir
        xlsx_path = out_dir / "deepseek_v4_training.xlsx"
        assert xlsx_path.exists(), f"Excel not found at {xlsx_path}"
        data[dp] = _load_xlsx_sheets(xlsx_path)

    return data


# ── ZeRO-1 optimizer state scaling ─────────────────────────────────────────

def test_dp_optimizer_state_scales_inverse(dp_reports):
    """Optimizer state per-GPU should roughly scale ~1/dp (ZeRO-1 behaviour)."""
    opt = {}
    for dp in (1, 4, 8):
        mb = dp_reports[dp].get("memory_breakdown_gb")
        assert mb, f"memory_breakdown_gb missing for dp={dp}"
        opt[dp] = mb.get("opt_state")
        assert opt[dp] is not None, f"opt_state missing for dp={dp}"

    # Approximate 1/dp scaling: opt(dp) ≈ opt(1) / dp
    for dp in (4, 8):
        expected = opt[1] / dp
        assert opt[dp] == pytest.approx(expected, rel=0.25), (
            f"opt_state did not scale near 1/dp: dp=1 → {opt[1]}, dp={dp} → {opt[dp]}, "
            f"expected ≈ {expected}"
        )

    # Monotonic decrease
    assert opt[1] > opt[4] > opt[8], (
        f"opt_state should decrease monotonically: dp1={opt[1]}, dp4={opt[4]}, dp8={opt[8]}"
    )


# ── Total memory monotonicity ──────────────────────────────────────────────

def test_dp_total_memory_decreases(dp_reports):
    """Total per-GPU memory should decrease monotonically as dp increases."""
    total = {}
    for dp in (1, 4, 8):
        mb = dp_reports[dp]["memory_breakdown_gb"]
        total[dp] = mb["total"]
        assert total[dp] is not None

    assert total[1] > total[4] > total[8], (
        f"total memory should decrease monotonically: "
        f"dp1={total[1]}, dp4={total[4]}, dp8={total[8]}"
    )


# ── Throughput monotonicity ────────────────────────────────────────────────

def test_dp_throughput_improves(dp_reports):
    """Step time should decrease and tokens/sec should increase monotonically."""
    step_time = {dp: dp_reports[dp]["step_time_ms"] for dp in (1, 4, 8)}

    assert step_time[1] > step_time[4] > step_time[8], (
        f"step_time should decrease monotonically: "
        f"dp1={step_time[1]}, dp4={step_time[4]}, dp8={step_time[8]}"
    )


# ── DP communication accounting ────────────────────────────────────────────

def test_dp_communication_zero_for_dp1(dp_reports):
    """dp_hidden_ms and dp_exposed_ms should be 0 for dp=1 (no DP communication)."""
    rep1 = dp_reports[1]

    dp_hidden1 = rep1.get("dp_hidden_ms")
    dp_exposed1 = rep1.get("dp_exposed_ms")
    assert dp_hidden1 is not None
    assert dp_exposed1 is not None
    assert dp_hidden1 == 0.0, f"dp_hidden should be 0 for dp=1, got {dp_hidden1}"
    assert dp_exposed1 == 0.0, f"dp_exposed should be 0 for dp=1, got {dp_exposed1}"


def test_dp_communication_nonzero_for_dp_gt1(dp_reports):
    """dp_hidden + dp_exposed should be >0 for dp>1 (DP AR/RS communication exists).

    dp_hidden may be 0 (e.g. no bubble to absorb AR) or dp_exposed may be 0
    (e.g. AR fully hidden in bubble), but their sum must reflect the total
    DP communication volume.
    """
    for dp in (4, 8):
        rep = dp_reports[dp]
        dp_hidden = rep.get("dp_hidden_ms")
        dp_exposed = rep.get("dp_exposed_ms")
        dp_comm = rep.get("dp_total_ms", 0.0)
        assert dp_hidden is not None, f"dp_hidden_ms missing for dp={dp}"
        assert dp_exposed is not None, f"dp_exposed_ms missing for dp={dp}"
        assert dp_comm > 0, (
            f"dp_hidden + dp_exposed should be >0 for dp={dp}, "
            f"got hidden={dp_hidden}, exposed={dp_exposed}"
        )


def test_dp_comm_volume_monotonic(dp_reports):
    """Total DP communication volume (dp_hidden + dp_exposed) should increase with dp.

    Larger DP group means more gradient data to reduce, so the total DP
    communication time (exposed + hidden) should grow monotonically.
    Note: on full-mesh topologies the per-step time may not scale linearly
    due to the (N-1)/N ring factor, but the trend should be increasing.
    """
    dp_comm = {}
    for dp in (4, 8):
        rep = dp_reports[dp]
        dp_comm[dp] = rep.get("dp_total_ms", 0.0)

    assert dp_comm[8] > dp_comm[4], (
        f"DP comm volume should increase with dp: "
        f"dp4={dp_comm[4]:.2f}ms, dp8={dp_comm[8]:.2f}ms"
    )


# ── Excel: DP comm operator existence ──────────────────────────────────────

class TestDpCommExistence:
    """DP communication operator count and presence by dp value."""

    def test_count_matches_layers(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        dp_rows = _filter_dp_comm(comm)
        assert len(dp_rows) == 4, (
            f"Expected 4 DP comm ops (one per layer), got {len(dp_rows)}"
        )

    def test_none_for_dp1(self, dp_excel_data):
        comm = dp_excel_data[1].get("Communication Ops", [])
        dp_rows = _filter_dp_comm(comm)
        assert len(dp_rows) == 0, (
            f"Expected 0 DP comm ops for dp=1, got {len(dp_rows)}"
        )


# ── Excel: DP comm operator attributes ─────────────────────────────────────

class TestDpCommAttributes:
    """DP communication node attribute correctness."""

    def test_collective_is_reduce_scatter(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Collective Op") == "reduce_scatter", (
                f"Expected reduce_scatter, got {row.get('Collective Op')}"
            )

    def test_group_size_equals_dp(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Group Size") == 4, (
                f"Expected 4, got {row.get('Group Size')}"
            )

    def test_role_is_dp_grad_reduce(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Role") == "dp_grad_reduce", (
                f"Expected dp_grad_reduce, got {row.get('Role')}"
            )

    def test_scope_pattern(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            scope = str(row.get("Scope", ""))
            assert "data_parallel.grad_reduce.layer_" in scope, (
                f"Unexpected scope: {scope}"
            )

    def test_inserted_by(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            inserted = str(row.get("Inserted By", "")).lower()
            assert "data_parallel" in inserted, (
                f"Expected data_parallel_pass, got {inserted}"
            )

    def test_has_stream_info(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Stream Type"), (
                f"Missing Stream Type on {row.get('Node ID')}"
            )
            assert row.get("Stream ID") is not None, (
                f"Missing Stream ID on {row.get('Node ID')}"
            )

    def test_node_id_pattern(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            nid = str(row.get("Node ID", ""))
            assert nid.startswith("comm_grad_reduce_layer_"), (
                f"Unexpected Node ID: {nid}"
            )


# ── Excel: DP comm scope encodes layer info ────────────────────────────────

class TestDpCommScope:
    """DP comm node scope carries per-layer info even when Layer column is empty
    (injected nodes do not populate `layer`)."""

    def test_scopes_contain_layer_keys(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        scopes = {str(r.get("Scope", "")) for r in _filter_dp_comm(comm)}
        for lk in ("0", "1", "2", "3"):
            assert any(f"layer_{lk}" in s for s in scopes), (
                f"No scope found for layer_{lk} in {scopes}"
            )

    def test_one_comm_per_scope(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        scopes = [str(r.get("Scope", "")) for r in _filter_dp_comm(comm)]
        assert len(scopes) == len(set(scopes)), (
            f"Duplicate scopes: {scopes}"
        )


# ── Excel: grad scale ops (aten.div.Scalar) ────────────────────────────────

class TestDpScaleOps:
    """DP gradient averaging scale nodes (aten.div.Scalar)
    in the Backward Operators sheet."""

    def _scale_rows(self, data: dict) -> list[dict]:
        t_ops = data.get("Backward Operators", [])
        return [
            r for r in t_ops
            if r.get("Op Type") == "aten.div.Scalar"
            and "grad_scale" in str(r.get("Node ID", ""))
        ]

    def test_count_eq_layers(self, dp_excel_data):
        assert len(self._scale_rows(dp_excel_data[4])) == 4, "Expected 4 grad scale ops"

    def test_none_for_dp1(self, dp_excel_data):
        assert len(self._scale_rows(dp_excel_data[1])) == 0, (
            "Expected 0 scale ops for dp=1"
        )

    def test_node_id_pattern(self, dp_excel_data):
        for row in self._scale_rows(dp_excel_data[4]):
            nid = str(row.get("Node ID", ""))
            assert nid.startswith("grad_scale_layer_"), (
                f"Unexpected scale node ID: {nid}"
            )